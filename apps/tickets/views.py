import random, hashlib, os, re, csv, json, io
from datetime import datetime, timedelta, date
from decimal import Decimal, InvalidOperation
from openpyxl import Workbook
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.contrib.auth.decorators import login_required, user_passes_test
from django.core.exceptions import PermissionDenied
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.core.management import call_command
from django.http import JsonResponse, HttpResponse, FileResponse
from django.core.files.base import ContentFile
from django.views.decorators.http import require_POST, require_http_methods
from django.views.decorators.csrf import csrf_exempt
from django.db.models import Count, Avg, Q, F, Value, Prefetch, Sum
from django.db import transaction, IntegrityError
from django.db.models.functions import TruncDate, Concat
from django.utils import timezone
from django.utils.html import strip_tags, escape
from django.urls import reverse
from django.template.loader import render_to_string
from apps.common.utils import send_email_via_brevo, role_of, resolve_sort
from django.conf import settings
from .forms import (
    TicketForm, IncidentReportForm, ServiceRequestForm, CommentForm, AssetForm, MobilizationForm,
    ProcurementRequestForm, TicketResolveForm, AssetReassignForm, AssetScrapRequestForm,
    AssetCheckoutForm, AssetCheckinForm, AssetReturnRequestForm, SLAForm, BusinessCalendarForm, EscalationRuleForm,
    ConnectorEditForm, EscalatedReassignForm, EscalatedReturnForm,
)
from .service_request_fields import build_service_request_details, fields_for_group, display_value_for_field
from .asset_import_transform import transform_raw_rows, resolve_status_hint, parse_track_no_slot
from .asset_name_matching import match_users_by_name
from .models import *
from apps.tickets.models import Asset
from apps.maintenance.models import Vendor, MaintenanceAssetConfirmation, MaintenanceSchedule
from apps.accounts.models import User
from apps.common.models import Notification, AdminActionLog, log_admin_action
from apps.common.permissions import is_admin, is_superadmin, get_sidebar_template, effective_role_name
from bs4 import BeautifulSoup
import bleach


import logging
logger = logging.getLogger(__name__)

# ==========================================================================
# HELPER FUNCTIONS
# ==========================================================================

_COMMENT_ALLOWED_TAGS = [
    'p', 'br', 'strong', 'b', 'em', 'i', 'u', 'ul', 'ol', 'li', 'a',
    'blockquote', 'code', 'pre', 'div', 'span',
]
_COMMENT_ALLOWED_ATTRS = {
    'a': ['href', 'title', 'target', 'rel'],
}
_COMMENT_ALLOWED_PROTOCOLS = ['http', 'https', 'mailto']


def clean_comment_body(body):
    """
    Sanitizes HTML content from the rich text editor before saving, using an
    allowlist (bleach) rather than trying to blocklist dangerous markup —
    strips anything not in _COMMENT_ALLOWED_TAGS/_COMMENT_ALLOWED_ATTRS
    entirely (script tags, event-handler attributes like onerror=,
    javascript:/data: URIs, iframe/style, etc.), then applies the same
    empty-tag cleanup and whitespace collapsing as before.
    Returns None if the cleaned body is empty (used to reject empty comments).
    """
    if not body:
        return None

    sanitized = bleach.clean(
        body,
        tags=_COMMENT_ALLOWED_TAGS,
        attributes=_COMMENT_ALLOWED_ATTRS,
        protocols=_COMMENT_ALLOWED_PROTOCOLS,
        strip=True,
    )

    soup = BeautifulSoup(sanitized, 'html.parser')

    # Remove empty block tags that contain no text and no formatting children
    for tag in soup.find_all():
        if tag.name in ['div', 'p', 'span', 'h1', 'h2', 'h3', 'h4', 'h5', 'h6']:
            if not tag.get_text(strip=True) and not tag.find_all(['strong', 'em', 'br']):
                tag.decompose()

    # Replace <br> with newline characters (they will be turned back into <br> by linebreaksbr filter)
    for br in soup.find_all('br'):
        br.replace_with('\n')

    cleaned = str(soup)
    cleaned = re.sub(r'\n\s*\n+', '\n', cleaned)
    cleaned = cleaned.strip()

    if not cleaned or cleaned == '':
        return None
    return cleaned

# get_sidebar_template is imported from apps.common.permissions (see top of file).

# apps/tickets/views.py - Fix apply_sla

def apply_sla(ticket):
    """
    Sets the response_due_at and resolution_due_at fields on a ticket
    based on the SLA policy configured for its priority.
    Called after ticket creation and when priority changes.

    Due dates walk forward through the SLA's assigned BusinessCalendar
    (work hours/workdays/holidays) via add_business_minutes() when one is
    set, so a P1 filed at 5pm Friday doesn't get a due date that lands
    mid-weekend. An SLA with no calendar attached falls back to flat
    calendar-time addition, same as before.
    """
    try:
        sla = SLA.objects.select_related('calendar').get(priority=ticket.priority)
    except SLA.DoesNotExist:
        return

    # Use the ticket's created_at as the start time
    # If created_at is None or in the future, use timezone.now()
    start_time = ticket.created_at if ticket.created_at and ticket.created_at <= timezone.now() else timezone.now()

    ticket.response_due_at = add_business_minutes(start_time, sla.response_minutes, sla.calendar)
    ticket.resolution_due_at = add_business_minutes(start_time, sla.resolution_minutes, sla.calendar)
    ticket.save(update_fields=['response_due_at', 'resolution_due_at'])

def notify_department_team_leads_pending_review(ticket):
    """Alerts every Team Lead in the requester's department that a service
    request is waiting on their manager-review queue — called whenever a
    ticket enters PENDING_MANAGER_REVIEW (on submission, and again if it
    re-enters review after the requester responds to changes-requested).

    Narrows via the legacy `role` field or the roles M2M (either can lag
    behind a user's actual active role, per report_registry.py's team-scope
    lookup), then resolves each candidate's true active role in Python so a
    stale field doesn't drop a real Team Lead from the notification."""
    from django.db.models import Q
    candidates = User.objects.filter(
        Q(role=User.Role.TEAM_LEAD) | Q(roles__name='TEAM_LEAD'),
        department=ticket.requester.department, is_active=True,
    ).distinct()
    leads = [u for u in candidates if effective_role_name(u) == 'TEAM_LEAD']
    for lead in leads:
        Notification.objects.create(
            recipient=lead,
            role=role_of(lead),
            message=f'Service request {ticket.number} from {ticket.requester.get_full_name()} needs your approval.',
            url=reverse('tickets:manager_review_ticket', kwargs={'pk': ticket.pk}),
            type=Notification.Type.MANAGER_REVIEW,
        )

# Helper function to handle "Other" field logic
def get_other_value(data, select_field, other_field, default_value):
    """Helper to handle 'Other' field logic for asset forms."""
    value = data.get(select_field)
    if value == 'OTHER':
        other_val = data.get(other_field, '').strip()
        return other_val if other_val else default_value
    return value

# Allowed MIME types for file attachments
ALLOWED_MIMES = [
    'image/jpeg', 'image/png', 'image/gif', 'image/webp',
    'application/pdf',
    'application/msword',
    'application/vnd.openxmlformats-officedocument.wordprocessingml.document',
    'application/vnd.ms-excel',
    'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    'text/plain',
    'application/zip',
]
MAX_SIZE_MB = 10

# Magic-byte signatures for the MIME types above that have a reliable one.
# The client-supplied Content-Type header is trivially spoofable (rename a
# script to .jpg, claim image/jpeg) — this checks what the file actually is.
# Types with no reliable signature (text/plain) fall through unsniffed.
_MIME_SIGNATURES = {
    'image/jpeg': [b'\xff\xd8\xff'],
    'image/png': [b'\x89PNG\r\n\x1a\n'],
    'image/gif': [b'GIF87a', b'GIF89a'],
    'image/webp': [b'RIFF'],  # followed by size(4) + 'WEBP', checked below
    'application/pdf': [b'%PDF-'],
    'application/zip': [b'PK\x03\x04', b'PK\x05\x06', b'PK\x07\x08'],
    # Modern Office formats are zip containers under the hood.
    'application/vnd.openxmlformats-officedocument.wordprocessingml.document': [b'PK\x03\x04'],
    'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet': [b'PK\x03\x04'],
    # Legacy Office formats (.doc/.xls) share the same OLE2 container header.
    'application/msword': [b'\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1'],
    'application/vnd.ms-excel': [b'\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1'],
}


def sniffed_mime_matches(uploaded_file, claimed_mime):
    """True if the file's actual bytes are consistent with claimed_mime, or
    claimed_mime has no known signature to check (e.g. text/plain)."""
    signatures = _MIME_SIGNATURES.get(claimed_mime)
    if not signatures:
        return True
    header = uploaded_file.read(16)
    uploaded_file.seek(0)
    if claimed_mime == 'image/webp':
        return header.startswith(b'RIFF') and header[8:12] == b'WEBP'
    return any(header.startswith(sig) for sig in signatures)


def save_attachments(ticket, files, author, comment=None):
    """
    Validates and saves uploaded file attachments.
    - Rejects files that exceed MAX_SIZE_MB or have disallowed/spoofed MIME types.
    - Computes SHA-256 hash for integrity checking.
    - Associates attachments with a ticket and optionally a specific comment.
    Returns (created, rejected) — created is a list of Attachment objects,
    rejected is a list of (filename, reason) tuples for anything skipped, so
    callers can surface exactly what happened instead of silently dropping it.
    """
    created = []
    rejected = []
    for f in files:
        if f.size > MAX_SIZE_MB * 1024 * 1024:
            rejected.append((f.name, f'exceeds the {MAX_SIZE_MB}MB limit'))
            continue
        mime = f.content_type.split(';')[0].strip().lower()
        if mime not in ALLOWED_MIMES:
            rejected.append((f.name, 'file type not allowed'))
            continue
        if not sniffed_mime_matches(f, mime):
            rejected.append((f.name, 'file type not allowed'))
            continue
        sha = hashlib.sha256()
        for chunk in f.chunks():
            sha.update(chunk)
        f.seek(0)
        att = Attachment.objects.create(
            ticket=ticket,
            comment=comment,
            file=f,
            filename=f.name,
            uploaded_by=author,
            content_type=mime,
            size=f.size,
            hash=sha.hexdigest(),
        )
        created.append(att)
    return created, rejected


def restore_kept_draft_attachments(ticket, request):
    """Copies draft attachments the client listed in `keep_draft_attachments`
    (comma-separated TicketDraftAttachment ids, submitted by form_draft.js
    for anything restored from a draft and not removed) onto the new
    ticket. A browser can never resubmit a file it didn't just pick this
    session, so this is the only way a restored-from-draft attachment
    actually reaches the ticket — the file has to already be sitting on the
    server from when it was originally picked and mirrored to the draft.

    Freshly-picked files this session never end up in `keep_draft_attachments`
    (only restored ones the client explicitly tracks by id) even though
    they're also mirrored to the draft in the background for resilience —
    so there's no risk of double-attaching the same file via both this and
    the live `attachments` field above. Whatever's left on the draft (kept
    or not) gets cleaned up regardless when the draft itself is deleted
    right after ticket creation."""
    raw_ids = request.POST.get('keep_draft_attachments', '')
    ids = [int(v) for v in raw_ids.split(',') if v.strip().isdigit()]
    if not ids:
        return
    draft_attachments = TicketDraftAttachment.objects.filter(pk__in=ids, draft__user=request.user)
    for draft_att in draft_attachments:
        draft_att.file.open('rb')
        content = draft_att.file.read()
        draft_att.file.close()
        Attachment.objects.create(
            ticket=ticket,
            file=ContentFile(content, name=draft_att.filename),
            filename=draft_att.filename,
            uploaded_by=request.user,
            content_type=draft_att.content_type,
            size=draft_att.size,
            hash=hashlib.sha256(content).hexdigest(),
        )

# ==========================================================================
# TICKET CREATION & LISTING VIEWS (End Users)
# ==========================================================================

# apps/tickets/views.py
import random
from django.contrib import messages

@login_required
def create_ticket(request):
    """
    Handles creation of a new ticket (incident or service request).
    - GET: displays the appropriate form (incident_form or service_request_form).
    - POST: validates form, generates a unique ticket number, saves ticket,
            attaches files, applies SLA, and redirects to ticket detail page.
    """
    # The form templates submit type as a hidden POST field (see
    # requester/service_request_form.html) — on a real form submission this
    # coincidentally also matched request.GET because the browser posts back
    # to the same "?type=..." URL it loaded, but that isn't guaranteed (a
    # bare test-client POST, a stripped query string, a JS-driven submit),
    # so read POST first and only fall back to GET for the initial page load.
    if request.method == 'POST':
        ticket_type = request.POST.get('type', request.GET.get('type', 'INCIDENT')).upper()
    else:
        ticket_type = request.GET.get('type', 'INCIDENT').upper()
    if ticket_type not in ['INCIDENT', 'SERVICE_REQUEST']:
        ticket_type = 'INCIDENT'

    if ticket_type == 'INCIDENT':
        FormClass = IncidentReportForm
    else:
        FormClass = ServiceRequestForm

    if request.method == 'POST':
        form = FormClass(request.POST)
        service_request_details = {}

        if form.is_valid() and ticket_type == 'SERVICE_REQUEST':
            service_category = form.cleaned_data.get('service_category')
            field_group = service_category.field_group if service_category else ServiceCategory.FieldGroup.GENERAL
            service_request_details, detail_errors = build_service_request_details(field_group, request.POST)
            for message_text in detail_errors.values():
                form.add_error(None, message_text)

        if not form.errors:
            ticket = form.save(commit=False)
            ticket.requester = request.user
            ticket.type = ticket_type
            if ticket_type == 'SERVICE_REQUEST':
                ticket.service_request_details = service_request_details

            notify_admins_new_job_number = None
            if ticket_type == 'SERVICE_REQUEST':
                job_number_selection = request.POST.get('job_number')
                new_job_number_text = (request.POST.get('new_job_number_text') or '').strip()
                if job_number_selection == 'NEW' and new_job_number_text:
                    job_number_obj, job_number_created = JobNumber.objects.get_or_create(
                        number=new_job_number_text,
                        defaults={'is_active': False, 'proposed_by': request.user},
                    )
                    ticket.job_number = job_number_obj
                    if job_number_created:
                        notify_admins_new_job_number = job_number_obj
                elif job_number_selection:
                    ticket.job_number = JobNumber.objects.filter(pk=job_number_selection, is_active=True).first()

                lat = request.POST.get('submission_latitude') or None
                lon = request.POST.get('submission_longitude') or None
                try:
                    ticket.submission_latitude = Decimal(lat) if lat else None
                    ticket.submission_longitude = Decimal(lon) if lon else None
                except (InvalidOperation, ValueError, TypeError):
                    ticket.submission_latitude = ticket.submission_longitude = None
                ticket.submission_location_address = (request.POST.get('submission_location_address') or '').strip()[:255]

            # Generate unique ticket number (e.g., TK#1234 or SRV#5678)
            prefix = 'TK' if ticket.type == Ticket.Type.INCIDENT else 'SRV'
            for _ in range(20):
                suffix = str(random.randint(0, 9999)).zfill(4)
                candidate = f"{prefix}#{suffix}"
                if not Ticket.objects.filter(number=candidate).exists():
                    ticket.number = candidate
                    break
            else:
                import time
                ticket.number = f"{prefix}#{int(time.time()) % 10000:04d}"

            ticket.save()
            if ticket_type == 'SERVICE_REQUEST':
                form.save_m2m()  # persists the `vessels`/`dive_systems` selections

            if notify_admins_new_job_number is not None:
                for admin in User.objects.filter(role=User.Role.ADMIN, is_active=True):
                    Notification.objects.create(
                        recipient=admin,
                        role=role_of(admin),
                        message=(
                            f'{request.user.get_full_name()} submitted service request {ticket.number} for job '
                            f'"{notify_admins_new_job_number.number}", which isn\'t in the system yet. Review and '
                            f'activate it under System Settings → Job Numbers if it should be added.'
                        ),
                        url=reverse('tickets:conversation', args=[ticket.pk]),
                    )

            files = request.FILES.getlist('attachments')
            if files:
                _, rejected = save_attachments(ticket, files, request.user)
                for name, reason in rejected:
                    messages.warning(request, f'"{name}" was not attached — {reason}.')

            restore_kept_draft_attachments(ticket, request)

            # If it's a service request, set status to PENDING_MANAGER_REVIEW
            if ticket.type == Ticket.Type.SERVICE_REQUEST:
                ticket.status = Ticket.Status.PENDING_MANAGER_REVIEW

                if ticket.service_category and ticket.service_category.field_group == ServiceCategory.FieldGroup.ASSET:
                    ticket.is_asset_request = True
                    ticket.is_mobilization_request = request.POST.get('is_mobilization_request') == 'on'

                ticket.save(update_fields=['status', 'is_asset_request', 'is_mobilization_request'])
                notify_department_team_leads_pending_review(ticket)

                messages.success(request, f'Service request {ticket.number} submitted for manager review.')
            else:
                messages.success(request, f'Ticket {ticket.number} created successfully.')

            apply_sla(ticket)

            # Submission is a normal full-page POST-redirect, not AJAX, so
            # there's no client-side "after submit" hook to fire a discard
            # request before the browser navigates away — clear the draft
            # here instead. Explicitly clear each attachment's storage file
            # first — restore_kept_draft_attachments() above already copied
            # anything kept into a real Attachment, so nothing here is still
            # needed, but FileField storage content isn't auto-deleted by
            # cascading the row, so skipping this would leave orphaned
            # blobs (kept-and-copied ones included) sitting in storage.
            draft = TicketDraft.objects.filter(user=request.user, ticket_type=ticket_type).first()
            if draft:
                for draft_att in draft.draft_attachments.all():
                    if draft_att.file:
                        draft_att.file.delete(save=False)
                draft.delete()

            return redirect('tickets:detail', pk=ticket.pk)
        else:
            messages.error(request, 'Please correct the errors below.')
    else:
        form = FormClass(initial={'type': ticket_type})

    template = 'requester/incident_form.html' if ticket_type == 'INCIDENT' else 'requester/service_request_form.html'
    context = {'form': form, 'ticket_type': ticket_type}
    if ticket_type == 'SERVICE_REQUEST':
        from .service_request_fields import DYNAMIC_FIELDS_BY_GROUP
        context['dynamic_fields_by_group'] = DYNAMIC_FIELDS_BY_GROUP
        context['vessels'] = Vessel.objects.filter(is_active=True)
        context['selected_vessels'] = set(request.POST.getlist('vessels')) if request.method == 'POST' else set()
        context['dive_systems'] = DiveSystem.objects.filter(is_active=True)
        context['selected_dive_systems'] = set(request.POST.getlist('dive_systems')) if request.method == 'POST' else set()
        context['job_numbers'] = JobNumber.objects.filter(is_active=True)
    return render(request, template, context)

@login_required
@require_POST
def cancel_ticket(request, pk):
    """
    Allows an end user to cancel a ticket that is still in NEW or TRIAGED status.
    - Closes the ticket and logs the action.
    - If the request is HTMX, returns the updated ticket list partial.
    """
    ticket = get_object_or_404(Ticket, pk=pk)
    if request.user != ticket.requester:
        return HttpResponse(status=403)
    if ticket.status not in [Ticket.Status.NEW, Ticket.Status.TRIAGED]:
        return HttpResponse(status=400)
    ticket.status = Ticket.Status.CLOSED
    ticket.save()
    TicketActivityLog.objects.create(
        ticket=ticket, action='status_changed', actor=request.user,
        details={'from': ticket.status, 'to': Ticket.Status.CLOSED, 'reason': 'Cancelled by requester'}
    )
    messages.success(request, f'Ticket {ticket.number} cancelled.')
    if request.headers.get('HX-Request'):
        tickets = Ticket.objects.filter(requester=request.user).order_by('-created_at')
        status_filter = request.POST.get('current_status', '')
        if status_filter and status_filter.upper() in dict(Ticket.Status.choices):
            tickets = tickets.filter(status=status_filter.upper())
        paginator = Paginator(tickets, 10)
        page_number = request.POST.get('page', 1)
        try:
            page_obj = paginator.page(page_number)
        except (PageNotAnInteger, EmptyPage):
            page_obj = paginator.page(1)
        context = {
            'tickets': page_obj,
            'current_status': status_filter,
            'status_choices': Ticket.Status.choices,
        }
        return render(request, 'partials/ticket_list_partial.html', context)
    return redirect('tickets:my_list')

# apps/tickets/views.py - my_ticket_list

@login_required
def my_ticket_list(request):
    """
    Displays a list of tickets created by the logged‑in end user.
    Supports filtering by status (OPEN/CLOSED or specific status).
    Uses URL parameters for filter persistence.
    """
    order_args, active_sort, sort_options = resolve_sort(request, TICKET_SORT_OPTIONS, '-created_at')
    tickets = Ticket.objects.filter(requester=request.user).order_by(*order_args)

    # Get filter parameters from URL
    status_filter = request.GET.get('status', '')
    base = request.GET.get('base', '')
    
    open_statuses = ['NEW', 'TRIAGED', 'ASSIGNED', 'IN_PROGRESS', 'PENDING_USER', 'PENDING_VENDOR', 'APPROVED']

    if status_filter == 'OPEN':
        tickets = tickets.filter(status__in=open_statuses)
    elif status_filter and status_filter.upper() in dict(Ticket.Status.choices):
        # CLOSED is a real status value (unlike OPEN, which is a synthetic
        # bucket), so it goes through this exact-match branch like RESOLVED
        # does — it must not be bucketed together with RESOLVED, or the
        # "Closed" KPI card's click-through would show Resolved tickets too.
        tickets = tickets.filter(status=status_filter.upper())

    # Pagination
    paginator = Paginator(tickets, 10)
    page_number = request.GET.get('page', 1)
    try:
        page_obj = paginator.page(page_number)
    except (PageNotAnInteger, EmptyPage):
        page_obj = paginator.page(1)

    # Build status choices based on base filter
    all_choices = Ticket.Status.choices
    if base == 'OPEN':
        status_choices = [
            ('NEW', 'New'), ('TRIAGED', 'Triaged'), ('ASSIGNED', 'Assigned'),
            ('IN_PROGRESS', 'In Progress'), ('PENDING_USER', 'Pending User'),
            ('PENDING_VENDOR', 'Pending Vendor'), ('APPROVED', 'Approved')
        ]
    elif base == 'CLOSED':
        status_choices = [('RESOLVED', 'Resolved'), ('CLOSED', 'Closed')]
    else:
        status_choices = all_choices

    base_status = base if base in ['OPEN', 'CLOSED'] else ''
    explicit = request.GET.get('explicit') == '1'

    # Pass selected_status to template
    context = {
        'tickets': page_obj,
        'current_status': status_filter or '',
        'selected_status': status_filter or '',  # For highlighting active chip
        'status_choices': status_choices,
        'sidebar_template': get_sidebar_template(request.user),
        'base_status': base_status,
        'explicit': explicit,
        'sort_options': sort_options,
        'active_sort': active_sort,
    }

    if request.headers.get('HX-Request'):
        return render(request, 'partials/ticket_list_partial.html', context)
    return render(request, 'requester/ticket_list.html', context)

@login_required
def ticket_detail(request, pk):
    """
    Unified ticket conversation page for both requesters and agents.
    - GET: displays the conversation timeline and comment form.
    - POST (HTMX): accepts a new public comment from the requester,
      cleans the HTML body, saves attachments, updates ticket status if needed,
      and returns the updated timeline partial.
    The 'is_agent' flag controls visibility of agent‑only UI elements.
    """
    ticket = get_object_or_404(Ticket, pk=pk)
    if request.user != ticket.requester and effective_role_name(request.user) not in ['AGENT', 'TEAM_LEAD', 'ADMIN', 'SUPERADMIN']:
        return redirect('dashboard')

    # Handle comment submission from requester
    if request.method == 'POST' and request.headers.get('HX-Request'):
        form = CommentForm(request.POST)
        if form.is_valid():
            files = request.FILES.getlist('attachments')
            cleaned_body = clean_comment_body(form.cleaned_data.get('body', ''))
            if cleaned_body is None and not files:
                return HttpResponse('Comment cannot be empty.', status=400)

            comment = form.save(commit=False)
            comment.ticket = ticket
            comment.author = request.user
            comment.visibility = 'PUBLIC'
            comment.body = cleaned_body or ''
            comment.save()

            rejected = []
            if files:
                _, rejected = save_attachments(ticket, files, request.user, comment=comment)

            if ticket.status == Ticket.Status.PENDING_USER:
                old_status = ticket.status
                # PENDING_USER means two different things depending on why it
                # was set: a manager requesting changes on a service request
                # (must go back through manager review, not straight to IT),
                # or an agent proposing a resolution and awaiting confirmation
                # (a reply here just reopens it to IN_PROGRESS). Disambiguate
                # via whichever of those two actions happened most recently.
                last_reason = ticket.activities.filter(
                    action__in=['manager_requested_changes', 'resolution_requested']
                ).first()
                if last_reason and last_reason.action == 'manager_requested_changes':
                    ticket.status = Ticket.Status.PENDING_MANAGER_REVIEW
                else:
                    ticket.status = Ticket.Status.IN_PROGRESS
                ticket.save()
                TicketActivityLog.objects.create(
                    ticket=ticket, action='status_changed', actor=request.user,
                    details={'from': old_status, 'to': ticket.status}
                )
                if ticket.status == Ticket.Status.PENDING_MANAGER_REVIEW:
                    notify_department_team_leads_pending_review(ticket)

            comments = ticket.comments.prefetch_related('attachment_set').all().order_by('created_at')
            initial_attachments = ticket.attachments.filter(comment__isnull=True)
            return render(request, 'partials/conversation_timeline.html', {
                'ticket': ticket,
                'comments': comments,
                'initial_attachments': initial_attachments,
                'attachment_warnings': rejected,
            })
        else:
            return HttpResponse('Please check your comment and try again.', status=422)

    # GET request – render conversation page
    comments = ticket.comments.all().order_by('created_at')
    initial_attachments = ticket.attachments.filter(comment__isnull=True)
    user_attachments = ticket.attachments.filter(uploaded_by__role='END_USER')
    agent_attachments = ticket.attachments.filter(
        uploaded_by__role__in=['AGENT', 'TEAM_LEAD', 'ADMIN', 'SUPERADMIN']
    )
    form = CommentForm()
    is_agent = effective_role_name(request.user) in ['AGENT', 'TEAM_LEAD', 'ADMIN', 'SUPERADMIN']

    return render(request, 'agent/ticket_conversation.html', {
        'ticket': ticket,
        'comments': comments,
        'form': form,
        'initial_attachments': initial_attachments,
        'user_attachments': user_attachments,
        'agent_attachments': agent_attachments,
        'sidebar_template': get_sidebar_template(request.user),
        'is_agent': is_agent,
    })

# ==========================================================================
# AGENT QUEUES & TICKET MANAGEMENT
# ==========================================================================

TICKET_SORT_OPTIONS = {
    '-created_at': (('-created_at',), 'Newest First'),
    '-updated_at': (('-updated_at',), 'Recently Updated'),
    'number': (('number',), 'Number (A-Z)'),
}

@login_required
def unassigned_queue(request):
    """
    Displays all unassigned tickets that are ready for agents to claim.
    """
    if effective_role_name(request.user) not in ['AGENT', 'TEAM_LEAD', 'ADMIN', 'SUPERADMIN'] or request.user.department != 'IT':
        return HttpResponse(status=403)

    # Get all unassigned tickets
    order_args, active_sort, sort_options = resolve_sort(request, TICKET_SORT_OPTIONS, '-created_at')
    tickets = Ticket.objects.filter(
        assigned_to__isnull=True
    ).select_related('requester', 'category').order_by(*order_args)

    # ================================================================
    # 🔥 Filter: Show INCIDENTS + APPROVED SERVICE REQUESTS
    # ================================================================
    tickets = tickets.filter(
        Q(type=Ticket.Type.INCIDENT) |
        (Q(type=Ticket.Type.SERVICE_REQUEST) & Q(status=Ticket.Status.APPROVED))
    )

    # Exclude tickets that shouldn't be in the queue
    tickets = tickets.exclude(
        status__in=[
            Ticket.Status.PENDING_MANAGER_REVIEW,
            Ticket.Status.PENDING_FULFILLMENT,
            Ticket.Status.PENDING_APPROVAL,
            Ticket.Status.RESOLVED,
            Ticket.Status.CLOSED,
        ]
    )

    paginator = Paginator(tickets, 20)
    page_obj = paginator.get_page(request.GET.get('page', 1))

    assignable_agents = User.objects.filter(
        role__in=['AGENT', 'TEAM_LEAD', 'ADMIN', 'SUPERADMIN'],
        department='IT',  # ✅ Only IT department
        is_active=True
    ).only('pk', 'first_name', 'last_name', 'email')

    context = {
        'tickets': page_obj,
        'assignable_agents': assignable_agents,
        'status_choices': Ticket.Status.choices,
        'sidebar_template': get_sidebar_template(request.user),
        'sort_options': sort_options,
        'active_sort': active_sort,
    }
    return render(request, 'agent/unassigned_queue.html', context)


@login_required
def unassigned_pending_count(request):
    """Badge count for the sidebar 'Unassigned' link — same filter as
    unassigned_queue, minus the fields that view only needs for display."""
    if effective_role_name(request.user) not in ['AGENT', 'TEAM_LEAD', 'ADMIN', 'SUPERADMIN'] or request.user.department != 'IT':
        return HttpResponse(status=403)

    count = Ticket.objects.filter(assigned_to__isnull=True).filter(
        Q(type=Ticket.Type.INCIDENT) |
        (Q(type=Ticket.Type.SERVICE_REQUEST) & Q(status=Ticket.Status.APPROVED))
    ).exclude(
        status__in=[
            Ticket.Status.PENDING_MANAGER_REVIEW,
            Ticket.Status.PENDING_FULFILLMENT,
            Ticket.Status.PENDING_APPROVAL,
            Ticket.Status.RESOLVED,
            Ticket.Status.CLOSED,
        ]
    ).count()
    return render(request, 'partials/sidebar_count_badge.html', {'count': count})


@login_required
def assigned_to_me(request):
    """
    Displays all tickets assigned to the logged‑in agent,
    excluding resolved, closed, and pending approval tickets.
    """
    if effective_role_name(request.user) not in ['AGENT', 'TEAM_LEAD', 'ADMIN', 'SUPERADMIN'] or request.user.department != 'IT':
        return HttpResponse(status=403)

    tickets = Ticket.objects.filter(
        assigned_to=request.user
    ).exclude(
        status__in=[Ticket.Status.RESOLVED, Ticket.Status.CLOSED, Ticket.Status.PENDING_APPROVAL, Ticket.Status.PENDING_MANAGER_REVIEW]
    ).select_related('requester', 'category').order_by('-created_at')

    assignable_agents = User.objects.filter(
        role__in=['AGENT', 'TEAM_LEAD', 'ADMIN', 'SUPERADMIN'],
        department='IT',  # ✅ Only IT department
        is_active=True
    ).only('pk', 'first_name', 'last_name', 'email')

    context = {
        'tickets': tickets,
        'assignable_agents': assignable_agents,
        'status_choices': Ticket.Status.choices,
        'sidebar_template': get_sidebar_template(request.user),
    }
    return render(request, 'agent/assigned_to_me.html', context)

from django.template.loader import render_to_string
from django.http import JsonResponse

# apps/tickets/views.py - Update claim_ticket

@login_required
def claim_ticket(request, pk):
    if effective_role_name(request.user) not in ['AGENT', 'TEAM_LEAD', 'ADMIN', 'SUPERADMIN']:
        return JsonResponse({
            'success': False,
            'message': 'You do not have permission to claim tickets.'
        }, status=403)

    source = request.POST.get('source', 'unassigned')

    with transaction.atomic():
        ticket = get_object_or_404(Ticket.objects.select_for_update(), pk=pk)
        claimed = False
        if ticket.assigned_to is None:
            ticket.assigned_to = request.user
            ticket.status = Ticket.Status.ASSIGNED
            ticket.save()
            claimed = True

    if claimed:
        # Log the assignment
        TicketActivityLog.objects.create(
            ticket=ticket, action='assigned', actor=request.user,
            details={'to': request.user.get_full_name(), 'status': ticket.status}
        )
        
        # Create notifications
        Notification.objects.create(
            recipient=request.user,
            role=role_of(request.user),
            message=f"You have claimed ticket {ticket.number}: {ticket.title}",
            url=reverse('tickets:conversation', args=[ticket.pk]),
            type=Notification.Type.TICKET
        )
        
        Notification.objects.create(
            recipient=ticket.requester,
            role=role_of(ticket.requester),
            message=f"Ticket {ticket.number} has been claimed by {request.user.get_full_name()} and is now in progress.",
            url=reverse('tickets:detail', args=[ticket.pk]),
            type=Notification.Type.TICKET
        )
        
        # ALWAYS return JSON for API-like requests
        # Check if the request is from HTMX or a form
        if request.headers.get('HX-Request'):
            # For HTMX requests, return HTML
            assignable_agents = User.objects.filter(
                role__in=['AGENT', 'TEAM_LEAD', 'ADMIN', 'SUPERADMIN']
            ).only('pk', 'first_name', 'last_name', 'email')
            
            unassigned_tickets = Ticket.objects.filter(
                assigned_to__isnull=True
            ).select_related('requester', 'category').exclude(
                status__in=[Ticket.Status.RESOLVED, Ticket.Status.CLOSED, Ticket.Status.PENDING_APPROVAL,
                            Ticket.Status.PENDING_MANAGER_REVIEW, Ticket.Status.PENDING_FULFILLMENT,
                            Ticket.Status.PENDING_VENDOR]
            ).order_by('-created_at')

            if source == 'dashboard':
                unassigned_tickets = unassigned_tickets[:5]
                return JsonResponse({
                    'success': True,
                    'message': f'Successfully claimed ticket {ticket.number}',
                    'unassigned_count': unassigned_tickets.count(),
                })
            else:
                unassigned_tickets = unassigned_tickets[:20]
                return render(request, 'partials/agent_ticket_table.html', {
                    'tickets': unassigned_tickets,
                    'assignable_agents': assignable_agents,
                    'status_choices': Ticket.Status.choices,
                    'source': source,
                })
        else:
            # For JSON requests, return JSON
            return JsonResponse({
                'success': True,
                'message': f'Successfully claimed ticket {ticket.number}',
                'ticket_id': ticket.pk,
                'ticket_number': ticket.number,
            })
    
    # If ticket is already assigned
    return JsonResponse({
        'success': False,
        'message': 'Ticket is already assigned to someone else.'
    }, status=400)

@login_required
def agent_ticket_detail(request, pk):
    """
    Returns a slide‑over panel with ticket details and comments.
    Used when an agent clicks the "eye" icon on a ticket row.
    """ 
    ticket = get_object_or_404(Ticket, pk=pk)
    if effective_role_name(request.user) not in [User.Role.AGENT, User.Role.TEAM_LEAD, User.Role.ADMIN, User.Role.SUPERADMIN]:
        return HttpResponse(status=403)
    comments = ticket.comments.all().order_by('created_at')
    user_attachments = ticket.attachments.filter(uploaded_by__role='END_USER')
    agent_attachments = ticket.attachments.filter(
        uploaded_by__role__in=['AGENT', 'TEAM_LEAD', 'ADMIN', 'SUPERADMIN']
    )
    return render(request, 'partials/ticket_slideover.html', {
        'ticket': ticket,
        'comments': comments,
        'user_attachments': user_attachments,
        'agent_attachments': agent_attachments,
    })

@login_required
def agent_ticket_conversation(request, pk):
    """
    Full‑page conversation view for agents (and admins).
    Renders the same template as ticket_detail but with is_agent=True.
    """
    ticket = get_object_or_404(Ticket, pk=pk)
    if effective_role_name(request.user) not in [User.Role.AGENT, User.Role.TEAM_LEAD, User.Role.ADMIN, User.Role.SUPERADMIN]:
        return HttpResponse(status=403)
    # An unclaimed ticket has no owner, so a reply here would move its
    # status (see add_comment_conversation) while it's still sitting in the
    # unassigned queue — claim it first, then open the conversation. Use
    # the "View Details" slideover (agent_ticket_detail) to preview an
    # unclaimed ticket without entering this page. Asset-request tickets are
    # exempt: they're worked through the fulfillment pool (fulfill_asset_request,
    # mobilization, procurement) and never go through assigned_to at all, so
    # this guard would otherwise lock admins out of fulfilling/commenting on
    # every asset request.
    if ticket.assigned_to_id is None and not ticket.is_asset_request:
        messages.warning(request, f'Claim ticket {ticket.number} before opening its conversation.')
        return redirect('tickets:unassigned')
    comments = ticket.comments.all().order_by('created_at')
    form = CommentForm()
    initial_attachments = ticket.attachments.filter(comment__isnull=True)
    user_attachments = ticket.attachments.filter(uploaded_by__role='END_USER')
    agent_attachments = ticket.attachments.filter(
        uploaded_by__role__in=['AGENT', 'TEAM_LEAD', 'ADMIN', 'SUPERADMIN']
    )
    return render(request, 'agent/ticket_conversation.html', {
        'ticket': ticket,
        'comments': comments,
        'form': form,
        'initial_attachments': initial_attachments,
        'user_attachments': user_attachments,
        'agent_attachments': agent_attachments,
        'sidebar_template': get_sidebar_template(request.user),
        'is_agent': True,
    })

@login_required
@require_POST
def add_comment_conversation(request, pk):
    """
    Handles agent comments (public or internal) on a ticket.
    - Cleans the HTML body using BeautifulSoup.
    - Saves attachments.
    - Updates ticket status to IN_PROGRESS if a public comment is added.
    - Returns the updated conversation timeline partial (HTMX).
    """
    ticket = get_object_or_404(Ticket, pk=pk)
    if effective_role_name(request.user) not in [User.Role.AGENT, User.Role.TEAM_LEAD, User.Role.ADMIN, User.Role.SUPERADMIN]:
        return HttpResponse(status=403)
    if ticket.assigned_to_id is None and not ticket.is_asset_request:
        return HttpResponse('Claim this ticket before replying.', status=403)

    form = CommentForm(request.POST)
    if not form.is_valid():
        return HttpResponse('Please check your comment and try again.', status=422)

    files = request.FILES.getlist('attachments')
    cleaned_body = clean_comment_body(form.cleaned_data.get('body', ''))
    if cleaned_body is None and not files:
        return HttpResponse('Comment cannot be empty.', status=400)

    comment = form.save(commit=False)
    comment.ticket = ticket
    comment.author = request.user
    comment.visibility = request.POST.get('visibility', 'PUBLIC').upper()
    if comment.visibility not in ['PUBLIC', 'INTERNAL']:
        comment.visibility = 'PUBLIC'
    comment.body = cleaned_body or ''
    comment.save()

    rejected = []
    if files:
        _, rejected = save_attachments(ticket, files, request.user, comment=comment)
        comment = TicketComment.objects.get(pk=comment.pk)

    TicketActivityLog.objects.create(
        ticket=ticket, action='commented', actor=request.user,
        details={'visibility': comment.visibility, 'body': comment.body[:200]}
    )

    old_status = ticket.status
    if comment.visibility == 'PUBLIC':
        if old_status in [Ticket.Status.ASSIGNED, Ticket.Status.IN_PROGRESS,
                          Ticket.Status.PENDING_USER, Ticket.Status.NEW, Ticket.Status.TRIAGED]:
            ticket.status = Ticket.Status.IN_PROGRESS
            ticket.save()
            if old_status != ticket.status:
                TicketActivityLog.objects.create(
                    ticket=ticket, action='status_changed', actor=request.user,
                    details={'from': old_status, 'to': ticket.status}
                )

    comments = ticket.comments.prefetch_related('attachment_set').all().order_by('created_at')
    initial_attachments = ticket.attachments.filter(comment__isnull=True)
    return render(request, 'partials/conversation_timeline.html', {
        'ticket': ticket,
        'comments': comments,
        'initial_attachments': initial_attachments,
        'attachment_warnings': rejected,
    })

@login_required
@require_http_methods(['GET', 'POST'])
def resolve_ticket(request, pk):
    """
    Initiates the resolve confirmation flow.
    - GET: Returns the resolve modal (HTMX)
    - POST: Processes the resolution confirmation
    """
    ticket = get_object_or_404(Ticket, pk=pk)
    if effective_role_name(request.user) not in [User.Role.AGENT, User.Role.TEAM_LEAD, User.Role.ADMIN, User.Role.SUPERADMIN]:
        return HttpResponse(status=403)
    
    # Check if ticket is already resolved
    if ticket.status in [Ticket.Status.RESOLVED, Ticket.Status.CLOSED]:
        if request.headers.get('HX-Request'):
            return HttpResponse('<div class="p-4 text-center"><p class="text-text-secondary">This ticket is already resolved.</p></div>')
        messages.warning(request, 'Ticket is already resolved or closed.')
        return redirect('tickets:conversation', pk=ticket.pk)
    
    is_incident = ticket.type == Ticket.Type.INCIDENT

    # GET request - return the modal (HTMX)
    if request.method == 'GET' and request.headers.get('HX-Request'):
        form = TicketResolveForm(is_incident=is_incident)
        return render(request, 'partials/resolve_modal.html', {'ticket': ticket, 'form': form})

    # POST request - process confirmation
    if request.method == 'POST':
        action = request.POST.get('action')

        if action == 'confirm':
            comment = request.POST.get('comment', '').strip()

            if ticket.is_asset_request and ticket.resolution_confirmed_at:
                # The requester already confirmed they received the
                # asset(s) (confirm_resolution) — no need to send it back
                # to them a second time. Resolve directly, no modal round-trip.
                ticket.status = Ticket.Status.RESOLVED
                ticket.resolved_at = timezone.now()
                ticket.save()

                TicketActivityLog.objects.create(
                    ticket=ticket, action='resolved', actor=request.user,
                    details={'comment': comment, 'note': 'receipt already confirmed by requester'}
                )
                TicketComment.objects.create(
                    ticket=ticket, author=request.user, visibility='PUBLIC',
                    body=f"**Resolved**.{' ' + escape(comment) if comment else ''}"
                )
                if request.headers.get('HX-Request'):
                    return HttpResponse(status=204, headers={'HX-Redirect': reverse('tickets:conversation', kwargs={'pk': ticket.pk})})
                messages.success(request, f'Ticket {ticket.number} resolved.')
                return redirect('tickets:conversation', pk=ticket.pk)

            form = TicketResolveForm(request.POST, is_incident=is_incident)
            if not form.is_valid():
                if request.headers.get('HX-Request'):
                    return render(request, 'partials/resolve_modal.html', {'ticket': ticket, 'form': form})
                messages.error(request, 'Please correct the errors below.')
                return redirect('tickets:conversation', pk=ticket.pk)

            comment = form.cleaned_data['comment'].strip()
            root_cause = form.cleaned_data['resolution_root_cause'].strip()
            resolution_steps = form.cleaned_data['resolution_steps'].strip()
            root_cause_categories = form.cleaned_data['resolution_root_cause_category']

            # Create resolution request
            ticket.status = Ticket.Status.PENDING_USER
            ticket.resolution_root_cause = root_cause
            ticket.resolution_steps = resolution_steps
            ticket.resolution_root_cause_category = root_cause_categories
            ticket.save()
            
            # Create activity log
            TicketActivityLog.objects.create(
                ticket=ticket,
                action='resolution_requested',
                actor=request.user,
                details={'comment': comment}
            )
            
            # Add system comment to ticket
            TicketComment.objects.create(
                ticket=ticket,
                author=request.user,
                body=f"**Resolution requested**: Please confirm if this ticket has been resolved.{' ' + escape(comment) if comment else ''}",
                visibility='PUBLIC'
            )
            
            # Send notification to requester
            Notification.objects.create(
                recipient=ticket.requester,
                role=role_of(ticket.requester),
                message=f"Please confirm if ticket {ticket.number} has been resolved.",
                url=reverse('tickets:confirm_resolution', args=[ticket.pk]),
                type=Notification.Type.RESOLUTION_CONFIRMATION
            )
            
            # Send email to requester
            confirm_url = request.build_absolute_uri(
                reverse('tickets:confirm_resolution', args=[ticket.pk])
            )
            html_message = render_to_string('emails/resolution_confirmation.html', {
                'requester_name': ticket.requester.get_full_name() or ticket.requester.email,
                'ticket_number': ticket.number,
                'ticket_title': ticket.title,
                'confirm_url': confirm_url,
                'agent_name': request.user.get_full_name() or request.user.email,
            })
            
            success, result = send_email_via_brevo(
                to_email=ticket.requester.email,
                subject=f"Please confirm resolution for ticket {ticket.number}",
                html_content=html_message,
                from_email=settings.DEFAULT_FROM_EMAIL
            )
            
            if not success:
                logger.error(f"Failed to send resolution confirmation email: {result}")
            
            if request.headers.get('HX-Request'):
                return HttpResponse(status=204, headers={'HX-Redirect': reverse('tickets:conversation', kwargs={'pk': ticket.pk})})
            messages.success(request, f'Resolution confirmation sent to {ticket.requester.get_full_name()}.')
            return redirect('tickets:conversation', pk=ticket.pk)

        elif action == 'cancel':
            return redirect('tickets:conversation', pk=ticket.pk)

    return HttpResponse(status=400)


@login_required
@require_POST
def approve_incident_report(request, pk):
    """
    IT Manager / Head of IT sign-off on a resolved Incident's report.
    Merged approval: this app has no separate IT Manager role distinct from
    Admin, and IT is the highest approval authority for incident reports —
    one approval satisfies both the "Reviewed By" and "Approved By" rows.
    """
    ticket = get_object_or_404(Ticket, pk=pk)

    if effective_role_name(request.user) not in [User.Role.ADMIN, User.Role.SUPERADMIN]:
        return HttpResponse(status=403)

    if ticket.type != Ticket.Type.INCIDENT or ticket.status not in [Ticket.Status.RESOLVED, Ticket.Status.CLOSED]:
        messages.error(request, 'Only resolved or closed Incident tickets can be approved.')
        return redirect('tickets:report_record_detail', report_type='incidents', pk=ticket.pk)

    if ticket.incident_approved_by:
        messages.info(request, f'Ticket {ticket.number} has already been approved.')
        return redirect('tickets:report_record_detail', report_type='incidents', pk=ticket.pk)

    ticket.incident_approved_by = request.user
    ticket.incident_approved_at = timezone.now()
    ticket.save()

    TicketActivityLog.objects.create(
        ticket=ticket,
        action='incident_report_approved',
        actor=request.user,
    )

    messages.success(request, f'Incident report for {ticket.number} approved.')
    return redirect('tickets:report_record_detail', report_type='incidents', pk=ticket.pk)


@login_required
def confirm_resolution(request, pk):
    """
    Page where requester confirms if ticket is resolved.
    """
    ticket = get_object_or_404(Ticket, pk=pk)
    
    # Security: only requester can confirm
    if request.user != ticket.requester:
        return HttpResponse(status=403)
    
    # Check if ticket is in pending user state (waiting for confirmation)
    if ticket.status != Ticket.Status.PENDING_USER:
        messages.warning(request, f'Ticket {ticket.number} is not awaiting resolution confirmation.')
        return redirect('tickets:detail', pk=ticket.pk)

    # Asset-request tickets (mobilization or single-asset) now confirm
    # receipt inline on the ticket conversation page via receipt_confirm_modal
    # — this standalone page is retired for them. A GET here (old bookmark/
    # email link) just redirects; the POST action handlers below are still
    # what the modal's form actually submits to for non-mobilization asset
    # requests, so that logic is untouched.
    if request.method == 'GET' and ticket.is_asset_request:
        messages.info(request, 'Confirm receipt from the ticket page.')
        return redirect('tickets:detail', pk=ticket.pk)

    if request.method == 'POST':
        action = request.POST.get('action')
        reason = request.POST.get('reason', '').strip()

        if action == 'confirm':
            ticket.resolution_confirmed_at = timezone.now()
            ticket.resolution_confirmed_by = request.user

            if ticket.is_asset_request:
                # Confirming receipt isn't the same as the ticket being
                # resolved — an agent still explicitly resolves it
                # afterward (resolve_ticket skips its own confirmation
                # round-trip once it sees resolution_confirmed_at is set).
                ticket.status = Ticket.Status.APPROVED
                ticket.save()

                TicketActivityLog.objects.create(
                    ticket=ticket,
                    action='receipt_confirmed',
                    actor=request.user,
                    details={'confirmed_at': ticket.resolution_confirmed_at.isoformat()}
                )
                messages.success(request, 'Thank you for confirming! Please rate your experience.')
                return redirect('tickets:submit_feedback', pk=ticket.pk)

            # Non-asset (Incident/general) tickets: confirming resolves it
            # directly, same as today.
            ticket.status = Ticket.Status.RESOLVED
            ticket.resolved_at = timezone.now()
            ticket.save()

            TicketActivityLog.objects.create(
                ticket=ticket,
                action='resolution_confirmed',
                actor=request.user,
                details={'confirmed_at': ticket.resolution_confirmed_at.isoformat()}
            )

            TicketComment.objects.create(
                ticket=ticket,
                author=request.user,
                body="**Resolution confirmed**. The issue has been resolved.",
                visibility='PUBLIC',
                is_system_generated=True,
            )

            messages.success(request, 'Thank you for confirming! Please rate your experience.')
            return redirect('tickets:submit_feedback', pk=ticket.pk)

        elif action == 'reopen':
            TicketActivityLog.objects.create(
                ticket=ticket,
                action='resolution_rejected',
                actor=request.user,
                details={'reason': reason}
            )

            if ticket.is_asset_request:
                # Back to the fulfillment queue for correction (wrong item,
                # missing item, etc.) rather than a generic "in progress".
                ticket.status = Ticket.Status.PENDING_FULFILLMENT
                ticket.save()

                TicketComment.objects.create(
                    ticket=ticket,
                    author=request.user,
                    body=f"**Not received**. The requester says the asset(s) weren't received as expected.{' Reason: ' + escape(reason) if reason else ''}",
                    visibility='PUBLIC'
                )

                if ticket.fulfilled_by:
                    Notification.objects.create(
                        recipient=ticket.fulfilled_by,
                        role=role_of(ticket.fulfilled_by),
                        message=f"{ticket.requester.get_full_name()} says request {ticket.number} wasn't received as expected.{' Reason: ' + reason if reason else ''}",
                        url=reverse('tickets:conversation', args=[ticket.pk])
                    )

                messages.info(request, f'Ticket {ticket.number} sent back for fulfillment review.')
                return redirect('tickets:detail', pk=ticket.pk)

            # User says issue is not resolved
            ticket.status = Ticket.Status.IN_PROGRESS
            ticket.save()

            TicketComment.objects.create(
                ticket=ticket,
                author=request.user,
                body=f"**Resolution rejected**. The issue is not fully resolved.{' Reason: ' + escape(reason) if reason else ''}",
                visibility='PUBLIC'
            )

            # Notify the agent who requested resolution
            last_log = TicketActivityLog.objects.filter(
                ticket=ticket,
                action='resolution_requested'
            ).order_by('-created_at').first()

            if last_log and last_log.actor:
                Notification.objects.create(
                    recipient=last_log.actor,
                    role=role_of(last_log.actor),
                    message=f"{ticket.requester.get_full_name()} rejected resolution for ticket {ticket.number}.{' Reason: ' + reason if reason else ''}",
                    url=reverse('tickets:conversation', args=[ticket.pk])
                )

            messages.info(request, f'Ticket {ticket.number} reopened for further investigation.')
            return redirect('tickets:detail', pk=ticket.pk)
        else:
            logger.warning(f"Unknown action: '{action}'")
            messages.error(request, f'Invalid action: {action}')
            return redirect('tickets:confirm_resolution', pk=ticket.pk)
    
    return render(request, 'tickets/confirm_resolution.html', {
        'ticket': ticket,
        'sidebar_template': get_sidebar_template(request.user),
    })


@login_required
def submit_feedback(request, pk):
    """
    Submit 1-5 star feedback for resolved ticket.
    """
    ticket = get_object_or_404(Ticket, pk=pk)
    
    # Security: only requester can submit feedback
    if request.user != ticket.requester:
        return HttpResponse(status=403)
    
    # Check if ticket is resolved and feedback not already submitted
    if ticket.status != Ticket.Status.RESOLVED:
        messages.warning(request, 'Feedback can only be submitted for resolved tickets.')
        return redirect('tickets:detail', pk=ticket.pk)
    
    if ticket.feedback_rating is not None:
        messages.info(request, 'You have already submitted feedback for this ticket.')
        return redirect('tickets:detail', pk=ticket.pk)
    
    if request.method == 'POST':
        rating = request.POST.get('rating')
        comment = request.POST.get('comment', '').strip()
        
        if not rating or int(rating) < 1 or int(rating) > 5:
            messages.error(request, 'Please select a rating from 1 to 5 stars.')
            return render(request, 'tickets/feedback_form.html', {
                'ticket': ticket,
                'sidebar_template': get_sidebar_template(request.user),
                'error': 'Please select a rating.'
            })
        
        ticket.feedback_rating = int(rating)
        ticket.feedback_comment = comment
        ticket.feedback_submitted_at = timezone.now()
        ticket.save()
        
        TicketActivityLog.objects.create(
            ticket=ticket,
            action='feedback_submitted',
            actor=request.user,
            details={'rating': rating, 'comment': comment}
        )
        
        # Notify the agent(s) who worked on this ticket
        assigned_log = TicketActivityLog.objects.filter(
            ticket=ticket,
            action='assigned'
        ).order_by('-created_at').first()
        
        if assigned_log and assigned_log.actor:
            star_emoji = '⭐' * int(rating) + '☆' * (5 - int(rating))
            Notification.objects.create(
                recipient=assigned_log.actor,
                role=role_of(assigned_log.actor),
                message=f"Feedback received for ticket {ticket.number}: {star_emoji} ({rating}/5)",
                url=reverse('tickets:detail', args=[ticket.pk])
            )
        
        messages.success(request, f'Thank you for your feedback! You rated this ticket {rating}/5 stars.')
        return redirect('tickets:detail', pk=ticket.pk)
    
    return render(request, 'tickets/feedback_form.html', {
        'ticket': ticket,
        'sidebar_template': get_sidebar_template(request.user),
    })

# ==========================================================================
# TICKET DETAILS PANEL & METADATA EDITING
# ==========================================================================

@login_required
def ticket_details_panel(request, pk):
    """
    Returns the right‑hand details panel (assignee, metadata, attachments)
    that slides in on the conversation page.
    """
    ticket = get_object_or_404(Ticket, pk=pk)
    followers = User.objects.filter(role__in=['AGENT','TEAM_LEAD'])[:5]
    user_attachments = ticket.attachments.filter(uploaded_by__role='END_USER')
    agent_attachments = ticket.attachments.filter(
        uploaded_by__role__in=['AGENT', 'TEAM_LEAD', 'ADMIN', 'SUPERADMIN']
    )
    return render(request, 'partials/ticket_details_panel.html', {
        'ticket': ticket,
        'followers': followers,
        'user_attachments': user_attachments,
        'agent_attachments': agent_attachments,
    })

@login_required
@require_POST
def edit_subject(request, pk):
    """
    Edits the ticket title inline (subject). Agent-tier roles only.
    Returns the updated subject display partial.
    """
    ticket = get_object_or_404(Ticket, pk=pk)
    if effective_role_name(request.user) not in [User.Role.AGENT, User.Role.TEAM_LEAD, User.Role.ADMIN, User.Role.SUPERADMIN]:
        return HttpResponse(status=403)
    new_title = request.POST.get('title', '').strip()
    if new_title:
        ticket.title = new_title
        ticket.save()
    return render(request, 'partials/subject_display.html', {'ticket': ticket, 'is_agent': True})

# ==========================================================================
# ASSIGNMENT POPOVERS AND ACTIONS
# ==========================================================================

@login_required
def assign_popover(request, pk):
    """
    Returns a popover with a list of assignable agents (Agent, Team Lead, Admin, Superadmin)
    so the agent can reassign the ticket.
    """
    if effective_role_name(request.user) not in ['AGENT', 'TEAM_LEAD', 'ADMIN', 'SUPERADMIN']:
        return HttpResponse(status=403)
    ticket = get_object_or_404(Ticket, pk=pk)
    agents = User.objects.filter(
        role__in=['AGENT', 'TEAM_LEAD', 'ADMIN', 'SUPERADMIN'],
        department='IT',  # ✅ Only IT department
        is_active=True
    )[:10]
    return render(request, 'partials/popovers/assign_popover.html', {'ticket': ticket, 'agents': agents})

@login_required
@require_POST
def assign_to_me(request, pk):
    """
    Assigns the ticket to the current user.
    Returns the updated assignee display partial.
    """
    if effective_role_name(request.user) not in ['AGENT', 'TEAM_LEAD', 'ADMIN', 'SUPERADMIN']:
        return HttpResponse(status=403)
    ticket = get_object_or_404(Ticket, pk=pk)
    old_assignee = ticket.assigned_to
    ticket.assigned_to = request.user
    ticket.save()
    TicketActivityLog.objects.create(
        ticket=ticket, action='assigned', actor=request.user,
        details={'from': old_assignee.get_full_name() if old_assignee else 'Unassigned', 'to': request.user.get_full_name()}
    )
    return render(request, 'partials/ticket_details_assignee.html', {'ticket': ticket})

@login_required
@require_POST
def assign_specific(request, pk, user_pk):
    """
    Assigns the ticket to a specific user (by primary key).
    Returns the updated assignee display partial.
    """
    if effective_role_name(request.user) not in ['AGENT', 'TEAM_LEAD', 'ADMIN', 'SUPERADMIN']:
        return HttpResponse(status=403)
    ticket = get_object_or_404(Ticket, pk=pk)
    agent = get_object_or_404(User, pk=user_pk)
    old_assignee = ticket.assigned_to
    ticket.assigned_to = agent
    ticket.save()
    TicketActivityLog.objects.create(
        ticket=ticket, action='assigned', actor=request.user,
        details={'from': old_assignee.get_full_name() if old_assignee else 'Unassigned', 'to': agent.get_full_name()}
    )
    return render(request, 'partials/ticket_details_assignee.html', {'ticket': ticket})

# ==========================================================================
# MACROS
# ==========================================================================

@login_required
def macro_list(request):
    """
    Returns a dropdown list of macros (predefined reply templates) for agents.
    Used in the conversation composer.
    """
    from .views_macros import visible_macros_for
    macros = visible_macros_for(request.user)
    return render(request, 'partials/macro_dropdown.html', {'macros': macros})

# ==========================================================================
# BULK ACTIONS (for ticket queues)
# ==========================================================================

@login_required
@require_POST
def bulk_action(request):
    """
    Performs bulk status change or assignment on multiple tickets.
    Only Team Leads, Admins, and Superadmins can bulk‑assign.
    Returns the updated agent ticket table partial.
    """
    ticket_ids_str = request.POST.get('ticket_ids', '')
    action = request.POST.get('action', '')
    value = request.POST.get('value', '')
    if not ticket_ids_str or not action:
        return HttpResponse(status=400)

    ids = [int(pk) for pk in ticket_ids_str.split(',') if pk.strip().isdigit()]
    tickets = Ticket.objects.filter(pk__in=ids)

    if action == 'status':
        if value in dict(Ticket.Status.choices):
            for ticket in tickets:
                old_status = ticket.status
                ticket.status = value
                ticket.save()
                TicketActivityLog.objects.create(
                    ticket=ticket, action='status_changed', actor=request.user,
                    details={'from': old_status, 'to': value, 'method': 'bulk'}
                )
                
    elif action == 'assign':
        if effective_role_name(request.user) not in ['TEAM_LEAD', 'ADMIN', 'SUPERADMIN']:
            return HttpResponse(status=403)
        if value.isdigit():
            agent = get_object_or_404(User, pk=int(value))
            actor_name = request.user.get_full_name()
            agent_name = agent.get_full_name()
            
            for ticket in tickets:
                old_assignee = ticket.assigned_to
                old_name = old_assignee.get_full_name() if old_assignee else 'Unassigned'
                
                # Reassign the ticket
                ticket.assigned_to = agent
                ticket.save()
                
                # Create activity log
                TicketActivityLog.objects.create(
                    ticket=ticket, action='assigned', actor=request.user,
                    details={'from': old_name, 'to': agent_name, 'method': 'bulk'}
                )
                
                # ================================================================
                # DEFAULT REASSIGN COMMENT (per ticket)
                # ================================================================
                TicketComment.objects.create(
                    ticket=ticket,
                    author=request.user,
                    body=f"**Bulk reassign**: Ticket reassigned by {escape(actor_name)} from **{escape(old_name)}** to **{escape(agent_name)}**.",
                    visibility='PUBLIC',
                    is_system_generated=True,
                )
            
            # Notify the agent once for all tickets
            Notification.objects.create(
                recipient=agent,
                role=role_of(agent),
                message=f"{len(tickets)} ticket(s) have been reassigned to you by {actor_name}.",
                url=reverse('tickets:unassigned')
            )

    source = request.POST.get('source', 'unassigned')
    if source == 'assigned':
        tickets = Ticket.objects.filter(
            assigned_to=request.user
        ).exclude(status__in=[Ticket.Status.RESOLVED, Ticket.Status.CLOSED, Ticket.Status.PENDING_APPROVAL]
        ).select_related('requester', 'category').order_by('-created_at')
    else:
        tickets = Ticket.objects.filter(
            assigned_to__isnull=True
        ).exclude(status__in=[Ticket.Status.RESOLVED, Ticket.Status.CLOSED]
        ).select_related('requester', 'category').order_by('-created_at')

    assignable_agents = User.objects.filter(
        role__in=['AGENT', 'TEAM_LEAD', 'ADMIN', 'SUPERADMIN'],
        department='IT',  # ✅ Only IT department
        is_active=True
    ).only('pk', 'first_name', 'last_name', 'email')
    return render(request, 'partials/agent_ticket_table.html', {
        'tickets': tickets,
        'assignable_agents': assignable_agents,
        'status_choices': Ticket.Status.choices,
    })

# ==========================================================================
# TEAM LEAD QUEUE & REASSIGNMENT
# ==========================================================================

@login_required
def team_queue(request):
    """
    Team Lead view: shows all tickets assigned to agents in the same department.
    Allows filtering by individual agent.
    """
    if effective_role_name(request.user) != 'TEAM_LEAD' or request.user.department != 'IT':
        return HttpResponse(status=403)
    team_members = User.objects.filter(
        department=request.user.department,
        role='AGENT',
        is_active=True
    )
    # Note: team_members already filtered by department from request.user
    order_args, active_sort, sort_options = resolve_sort(request, TICKET_SORT_OPTIONS, '-created_at')
    tickets = Ticket.objects.filter(
        assigned_to__in=team_members
    ).exclude(status__in=[Ticket.Status.RESOLVED, Ticket.Status.CLOSED]
    ).order_by(*order_args)
    agent_id = request.GET.get('agent')
    if agent_id:
        tickets = tickets.filter(assigned_to_id=agent_id)
    context = {
        'tickets': tickets,
        'team_members': team_members,
        'selected_agent': agent_id,
        'sidebar_template': get_sidebar_template(request.user),
        'sort_options': sort_options,
        'active_sort': active_sort,
    }
    return render(request, 'partials/team_queue.html', context)

@login_required
@require_POST
def team_reassign(request, pk):
    """
    Allows a Team Lead to reassign a ticket to another agent in their team.
    Returns JSON status.
    """
    if effective_role_name(request.user) != 'TEAM_LEAD':
        return HttpResponse(status=403)
    ticket = get_object_or_404(Ticket, pk=pk)
    new_agent_id = request.POST.get('agent_id')
    agent = get_object_or_404(User, pk=new_agent_id, role='AGENT')
    old_assignee = ticket.assigned_to
    
    # Store old assignee name for comment
    old_name = old_assignee.get_full_name() if old_assignee else 'Unassigned'
    new_name = agent.get_full_name()
    actor_name = request.user.get_full_name()
    
    # Reassign the ticket
    ticket.assigned_to = agent
    ticket.save()
    
    # Create activity log
    TicketActivityLog.objects.create(
        ticket=ticket, action='assigned', actor=request.user,
        details={'from': old_name, 'to': new_name}
    )
    
    # ================================================================
    # DEFAULT REASSIGN COMMENT
    # ================================================================
    TicketComment.objects.create(
        ticket=ticket,
        author=request.user,
        body=f"**Ticket reassigned** by {escape(actor_name)} from **{escape(old_name)}** to **{escape(new_name)}**.",
        visibility='PUBLIC',
        is_system_generated=True,
    )

    # Notify the new agent
    Notification.objects.create(
        recipient=agent,
        role=role_of(agent),
        message=f"Ticket {ticket.number} has been reassigned to you by {actor_name}.",
        url=reverse('tickets:conversation', args=[ticket.pk])
    )
    
    return JsonResponse({'status': 'ok'})

# ==========================================================================
# AUDIT LOG
# ==========================================================================

# Category grouping for the Logs page — bucket each recorded action into a
# human-facing category so the page reads as "what area was this in" rather
# than a flat, undifferentiated table. Keep in sync with every
# TicketActivityLog.objects.create(action=...) call site.
LOG_CATEGORY_MAP = {
    'status_changed': 'Ticket Lifecycle',
    'assigned': 'Ticket Lifecycle',
    'commented': 'Ticket Lifecycle',
    'resolution_requested': 'Ticket Lifecycle',
    'resolution_confirmed': 'Ticket Lifecycle',
    'resolution_rejected': 'Ticket Lifecycle',
    'receipt_confirmed': 'Ticket Lifecycle',
    'resolved': 'Ticket Lifecycle',
    'feedback_submitted': 'Ticket Lifecycle',
    'manager_approved': 'Manager Review',
    'manager_rejected': 'Manager Review',
    'manager_requested_changes': 'Manager Review',
    'breached': 'Escalation',
    'escalation_rule_fired': 'Escalation',
    'reassigned_escalated': 'Escalation',
    'returned_to_pool': 'Escalation',
    'remote_session_requested': 'Remote Sessions',
    'remote_session_status_change': 'Remote Sessions',
    'asset_fulfilled': 'Assets',
    'mobilization_fulfilled': 'Assets',
}
LOG_CATEGORY_ORDER = ['Ticket Lifecycle', 'Manager Review', 'Escalation', 'Remote Sessions', 'Assets', 'Other']

# Friendly labels for the JSONField keys stored in TicketActivityLog.details,
# so the Logs page can show "From: Open / To: Resolved" instead of a raw
# dict dump. Unknown keys fall back to a title-cased version of themselves.
LOG_DETAIL_LABELS = {
    'from': 'From', 'to': 'To', 'reason': 'Reason', 'method': 'Method',
    'body': 'Comment', 'visibility': 'Visibility', 'status': 'Status',
    'rating': 'Rating', 'comment': 'Feedback',
}


def _log_category(action):
    return LOG_CATEGORY_MAP.get(action, 'Other')


def _log_detail_items(details):
    """Turn a TicketActivityLog.details dict into readable (label, value) pairs."""
    if not details:
        return []
    items = []
    for key, value in details.items():
        label = LOG_DETAIL_LABELS.get(key, key.replace('_', ' ').title())
        if isinstance(value, str) and len(value) > 160:
            value = value[:160] + '…'
        items.append((label, value))
    return items


def _audit_log_column_help():
    # Reuses the same descriptions as the generic Exportables 'audit-logs'
    # report type (report_registry.py) rather than duplicating them — this
    # view's hand-rolled export writes the same six columns.
    from .report_registry import REPORT_TYPES
    return REPORT_TYPES['audit-logs'].column_help


@login_required
def audit_log(request):
    if effective_role_name(request.user) not in ['TEAM_LEAD', 'ADMIN', 'SUPERADMIN'] or request.user.department != 'IT':
        return HttpResponse(status=403)

    tab = request.GET.get('tab', 'tickets')
    is_admin_tier = effective_role_name(request.user) in ['ADMIN', 'SUPERADMIN']

    # ------------------------------------------------------------------
    # SYSTEM TAB — impersonation events. Admin/Superadmin only; previously
    # this data existed in ImpersonationLog with no UI surfacing it at all.
    # ------------------------------------------------------------------
    if tab == 'system':
        if not is_admin_tier:
            return HttpResponse(status=403)
        from apps.accounts.models import ImpersonationLog
        system_logs = ImpersonationLog.objects.select_related('admin', 'target_user').order_by('-started_at')
        active_sort = request.GET.get('sort', 'newest')
        if active_sort not in ('newest', 'oldest'):
            active_sort = 'newest'
        if active_sort == 'oldest':
            system_logs = system_logs.reverse()
        paginator = Paginator(system_logs, 50)
        page_obj = paginator.get_page(request.GET.get('page', 1))
        context = {
            'tab': 'system',
            'system_logs': page_obj,
            'is_admin_tier': is_admin_tier,
            'sidebar_template': get_sidebar_template(request.user),
            'sort_options': [('newest', 'Newest First'), ('oldest', 'Oldest First')],
            'active_sort': active_sort,
        }
        return render(request, 'partials/audit_log.html', context)

    # ------------------------------------------------------------------
    # TICKETS TAB — ticket activity, grouped by category.
    # ------------------------------------------------------------------
    logs = TicketActivityLog.objects.select_related('ticket', 'actor').all()
    if effective_role_name(request.user) == 'TEAM_LEAD':
        team_members = User.objects.filter(department=request.user.department, role='AGENT')
        logs = logs.filter(
            Q(ticket__assigned_to__in=team_members) | Q(ticket__requester__in=team_members)
        )

    action = request.GET.get('action')
    category = request.GET.get('category')
    ticket_id = request.GET.get('ticket')
    if action:
        logs = logs.filter(action=action)
    if category and category in LOG_CATEGORY_MAP.values():
        matching_actions = [a for a, c in LOG_CATEGORY_MAP.items() if c == category]
        logs = logs.filter(action__in=matching_actions)
    if ticket_id:
        logs = logs.filter(ticket__number__icontains=ticket_id)
    logs = logs.order_by('-created_at')
    active_sort = request.GET.get('sort', 'newest')
    if active_sort not in ('newest', 'oldest'):
        active_sort = 'newest'
    if active_sort == 'oldest':
        logs = logs.reverse()

    # --- Export logic (CSV, JSON, Excel) ---
    export_format = request.GET.get('format')
    if export_format:
        filename = f"logs_{timezone.now().strftime('%Y%m%d_%H%M%S')}"

        # Convert each log to a flat dict for export
        export_data = []
        for log in logs:
            export_data.append({
                'time': log.created_at.strftime('%Y-%m-%d %H:%M:%S'),
                'category': _log_category(log.action),
                'ticket': log.ticket.number if log.ticket else '—',
                'action': log.action,
                'actor': log.actor.get_full_name() if log.actor else 'System',
                'details': str(log.details) if log.details else ''  # Convert dict to string
            })

        # Column picker (CSV/Excel only, see components/export_menu.html) —
        # `cols` is a comma-separated subset of the header labels below,
        # re-filtered/re-ordered against the real list so a tampered value
        # just falls back to every column.
        all_columns = [('Time', 'time'), ('Category', 'category'), ('Ticket', 'ticket'),
                        ('Action', 'action'), ('Actor', 'actor'), ('Details', 'details')]
        columns = all_columns
        if export_format in ('csv', 'excel') and request.GET.get('cols'):
            requested = set(c.strip() for c in request.GET['cols'].split(','))
            selected = [c for c in all_columns if c[0] in requested]
            if selected:
                columns = selected

        if export_format == 'csv':
            response = HttpResponse(content_type='text/csv')
            response['Content-Disposition'] = f'attachment; filename="{filename}.csv"'
            writer = csv.writer(response)
            writer.writerow([label for label, key in columns])
            for row in export_data:
                writer.writerow([row[key] for label, key in columns])
            return response

        elif export_format == 'json':
            response = HttpResponse(json.dumps(export_data, indent=2), content_type='application/json')
            response['Content-Disposition'] = f'attachment; filename="{filename}.json"'
            return response

        elif export_format == 'excel':
            wb = Workbook()
            ws = wb.active
            ws.title = "Logs"
            ws.append([label for label, key in columns])
            for row in export_data:
                ws.append([row[key] for label, key in columns])
            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = f'attachment; filename="{filename}.xlsx"'
            wb.save(response)
            return response

    # --- Pagination, then bucket this page's entries into category groups ---
    paginator = Paginator(logs, 50)
    page_number = request.GET.get('page', 1)
    page_obj = paginator.get_page(page_number)

    grouped = {cat: [] for cat in LOG_CATEGORY_ORDER}
    for log in page_obj:
        log.category = _log_category(log.action)
        log.detail_items = _log_detail_items(log.details)
        grouped[log.category].append(log)
    grouped_logs = [(cat, entries) for cat, entries in grouped.items() if entries]

    base_get = request.GET.copy()
    base_get.pop('page', None)

    context = {
        'tab': 'tickets',
        'logs': page_obj,
        'grouped_logs': grouped_logs,
        'action_choices': sorted(LOG_CATEGORY_MAP.keys()),
        'category_choices': LOG_CATEGORY_ORDER[:-1],  # exclude 'Other' from the filter
        'is_admin_tier': is_admin_tier,
        'base_qs': base_get.urlencode(),
        'audit_log_columns': ['Time', 'Category', 'Ticket', 'Action', 'Actor', 'Details'],
        'audit_log_column_help': _audit_log_column_help(),
        'sidebar_template': get_sidebar_template(request.user),
        'sort_options': [('newest', 'Newest First'), ('oldest', 'Oldest First')],
        'active_sort': active_sort,
    }
    return render(request, 'partials/audit_log.html', context)

# ==========================================================================
# REPORTS DASHBOARD (Charts & KPIs)
# ==========================================================================

@login_required
def reports_dashboard(request):
    """Renders the reports page with SLA compliance, ticket volume, and priority charts."""
    user = request.user
    if effective_role_name(user) not in ('ADMIN', 'SUPERADMIN', 'TEAM_LEAD'):
        return HttpResponse(status=403)
    
    # ================================================================
    # DATE RANGE FILTER
    # ================================================================
    start_date_str = request.GET.get('start_date')
    end_date_str = request.GET.get('end_date')
    
    if start_date_str and end_date_str:
        try:
            from datetime import datetime
            start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
            end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
            date_range_filter = Q(created_at__date__gte=start_date) & Q(created_at__date__lte=end_date)
            date_range_label = f"{start_date.strftime('%b %d, %Y')} – {end_date.strftime('%b %d, %Y')}"
            # Calculate number of days for volume chart
            volume_days = (end_date - start_date).days + 1
            volume_start = start_date
        except ValueError:
            # Invalid date format - fallback to 30 days
            end_date = timezone.now().date()
            start_date = end_date - timedelta(days=29)
            date_range_filter = Q(created_at__date__gte=start_date) & Q(created_at__date__lte=end_date)
            date_range_label = "Last 30 days"
            volume_days = 30
            volume_start = start_date
    else:
        # Default: Last 30 days
        end_date = timezone.now().date()
        start_date = end_date - timedelta(days=29)
        date_range_filter = Q(created_at__date__gte=start_date) & Q(created_at__date__lte=end_date)
        date_range_label = "Last 30 days"
        volume_days = 30
        volume_start = start_date
    
    if effective_role_name(user) == 'TEAM_LEAD':
        team_members = User.objects.filter(department=user.department, role='AGENT')
        ticket_filter = Q(assigned_to__in=team_members) | Q(requester__in=team_members)
    else:
        ticket_filter = Q()
    
    # Apply date range to ticket filter
    ticket_filter = ticket_filter & date_range_filter
    
    # ========== SLA COMPLIANCE ==========
    slas = SLA.objects.all().order_by('priority')
    sla_data = []

    for sla in slas:
        resolved_tickets = Ticket.objects.filter(
            ticket_filter,
            priority=sla.priority,
            status__in=['RESOLVED', 'CLOSED'],
            resolved_at__isnull=False
        )
        
        total = resolved_tickets.count()
        
        if total == 0:
            compliance = 100
            compliant_count = 0
            breached_count = 0
        else:
            compliant_count = 0
            breached_count = 0
            
            for ticket in resolved_tickets:
                if ticket.resolved_at and ticket.created_at:
                    actual_minutes = (ticket.resolved_at - ticket.created_at).total_seconds() / 60
                    
                    if actual_minutes <= sla.resolution_minutes:
                        compliant_count += 1
                    else:
                        breached_count += 1
            
            compliance = round((compliant_count / total) * 100, 1)
        
        sla_data.append({
            'priority': sla.get_priority_display(),
            'compliance': compliance,
            'total': total,
            'compliant': compliant_count,
            'breached': breached_count,
            'sla_minutes': sla.resolution_minutes,
            'sla_display': sla.get_resolution_display(),
        })

    # ========== TICKET VOLUME (Date Range) ==========
    created_qs = Ticket.objects.filter(
        ticket_filter, created_at__date__gte=volume_start
    ).annotate(date=TruncDate('created_at')).values('date').annotate(count=Count('id')).order_by('date')
    
    resolved_qs = Ticket.objects.filter(
        ticket_filter, resolved_at__isnull=False, resolved_at__date__gte=volume_start
    ).annotate(date=TruncDate('resolved_at')).values('date').annotate(count=Count('id')).order_by('date')

    dates, created_counts, resolved_counts = [], [], []
    for i in range(volume_days):
        d = volume_start + timedelta(days=i)
        dates.append(d.strftime('%m/%d'))
        created_counts.append(next((x['count'] for x in created_qs if x['date'] == d), 0))
        resolved_counts.append(next((x['count'] for x in resolved_qs if x['date'] == d), 0))

    # ========== MTTR ==========
    resolved = Ticket.objects.filter(
        ticket_filter, status__in=['RESOLVED', 'CLOSED'], resolved_at__isnull=False
    )
    mttr = resolved.aggregate(avg_mttr=Avg(F('resolved_at') - F('created_at')))['avg_mttr']
    mttr_minutes = round(mttr.total_seconds() / 60) if mttr else 0

    # ========== BACKLOG ==========
    backlog = Ticket.objects.filter(
        ticket_filter,
        status__in=['NEW', 'TRIAGED', 'ASSIGNED', 'IN_PROGRESS', 'PENDING_USER', 'PENDING_VENDOR'],
        created_at__lt=timezone.now() - timedelta(days=7)
    ).count()

    # ========== OPEN BY PRIORITY ==========
    open_by_priority = Ticket.objects.filter(
        ticket_filter,
        status__in=['NEW', 'TRIAGED', 'ASSIGNED', 'IN_PROGRESS', 'PENDING_USER', 'PENDING_VENDOR']
    ).values('priority').annotate(count=Count('id')).order_by('priority')

    open_labels, open_data = [], []
    for p in open_by_priority:
        open_labels.append(dict(Ticket.Priority.choices)[p['priority']])
        open_data.append(p['count'])

    # ========== ASSET METRICS ==========
    
    total_assets = Asset.objects.count()
    
    asset_status_labels = ['Active', 'In Store', 'Maintenance', 'Damaged', 'Scrapped']
    asset_status_counts = [
        Asset.objects.filter(status__in=[Asset.Status.IN_USE, Asset.Status.MOBILIZED]).count(),
        Asset.objects.filter(status=Asset.Status.IN_STORE).count(),
        Asset.objects.filter(status=Asset.Status.MAINTENANCE).count(),
        Asset.objects.filter(status=Asset.Status.DAMAGED).count(),
        Asset.objects.filter(status=Asset.Status.SCRAPPED).count(),
    ]
    
    # Scoped by ticket_filter (department + the page's own date-range
    # picker) — previously always counted all-time, silently ignoring
    # whatever range the user had selected while every other card on this
    # page (SLA, MTTR, Backlog, Open Tickets) did respect it.
    total_asset_requests = Ticket.objects.filter(
        ticket_filter,
        type=Ticket.Type.SERVICE_REQUEST,
        is_asset_request=True
    ).count()

    fulfilled_asset_requests = Ticket.objects.filter(
        ticket_filter,
        type=Ticket.Type.SERVICE_REQUEST,
        is_asset_request=True,
        fulfilled_at__isnull=False
    ).count()

    fulfillment_rate = round((fulfilled_asset_requests / total_asset_requests * 100), 1) if total_asset_requests > 0 else 0

    fulfilled_tickets = Ticket.objects.filter(
        ticket_filter,
        type=Ticket.Type.SERVICE_REQUEST,
        is_asset_request=True,
        fulfilled_at__isnull=False,
        created_at__isnull=False
    )
    
    total_hours = 0
    count = 0
    for ticket in fulfilled_tickets:
        delta = ticket.fulfilled_at - ticket.created_at
        total_hours += delta.total_seconds() / 3600
        count += 1
    
    avg_fulfillment_hours = round(total_hours / count, 1) if count > 0 else 0

    context = {
        'sla_data': sla_data,
        'volume_dates': dates,
        'volume_created': created_counts,
        'volume_resolved': resolved_counts,
        'mttr_minutes': mttr_minutes,
        'backlog': backlog,
        'open_priority_labels': open_labels,
        'open_priority_data': open_data,
        'open_total': sum(open_data),
        'sidebar_template': get_sidebar_template(request.user),
        # Full asset inventory is Admin/Superadmin-only — the KPI cards below
        # stay visible to Team Lead as read-only numbers, just not clickable
        # through to the restricted inventory list.
        'can_view_asset_inventory': effective_role_name(user) in ('ADMIN', 'SUPERADMIN'),
        # Asset metrics
        'total_assets': total_assets,
        'asset_status_labels': asset_status_labels,
        'asset_status_counts': asset_status_counts,
        'total_asset_requests': total_asset_requests,
        'fulfilled_asset_requests': fulfilled_asset_requests,
        'fulfillment_rate': fulfillment_rate,
        'avg_fulfillment_hours': avg_fulfillment_hours,
        # Date range context
        'start_date': start_date,
        'end_date': end_date,
        'date_range_label': date_range_label,
    }
    return render(request, 'dashboards/reports.html', context)


@login_required
def reports_ticket_list(request):
    """Deep-link target for the Reports dashboard's org-wide KPI cards
    (Open Tickets / Backlog) — there's no other view that lists tickets
    across the whole org (existing list views are queue/requester-scoped),
    so this reuses the same department-scoping rule as reports_dashboard."""
    user = request.user
    if effective_role_name(user) not in ('ADMIN', 'SUPERADMIN', 'TEAM_LEAD'):
        return HttpResponse(status=403)

    if effective_role_name(user) == 'TEAM_LEAD':
        team_members = User.objects.filter(department=user.department, role='AGENT')
        ticket_filter = Q(assigned_to__in=team_members) | Q(requester__in=team_members)
    else:
        ticket_filter = Q()

    kpi = request.GET.get('kpi', 'open')
    open_statuses = ['NEW', 'TRIAGED', 'ASSIGNED', 'IN_PROGRESS', 'PENDING_USER', 'PENDING_VENDOR']

    if kpi == 'backlog':
        tickets = Ticket.objects.filter(
            ticket_filter, status__in=open_statuses,
            created_at__lt=timezone.now() - timedelta(days=7)
        )
        heading = 'Backlog (open more than 7 days)'
    else:
        tickets = Ticket.objects.filter(ticket_filter, status__in=open_statuses)
        heading = 'Open Tickets'
    order_args, active_sort, sort_options = resolve_sort(request, TICKET_SORT_OPTIONS, '-created_at')
    tickets = tickets.order_by(*order_args)

    paginator = Paginator(tickets, 20)
    page_obj = paginator.get_page(request.GET.get('page', 1))

    context = {
        'tickets': page_obj,
        'heading': heading,
        'sidebar_template': get_sidebar_template(request.user),
        'sort_options': sort_options,
        'active_sort': active_sort,
    }
    return render(request, 'dashboards/reports_ticket_list.html', context)

# ==========================================================================
# EXTERNAL CRON TRIGGERS (SLA & CLEANUP)
# ==========================================================================

# is_admin is imported from apps.common.permissions (see top of file).

@login_required
@user_passes_test(is_admin)
def trigger_sla_processing(request):
    """
    Admin-only endpoint to trigger SLA processing.
    Protected by authentication and role check.
    """
    try:
        call_command('process_sla')
        return JsonResponse({'status': 'ok', 'message': 'SLA processing triggered successfully.'})
    except Exception:
        logger.exception('trigger_sla_processing failed')
        return JsonResponse({'status': 'error', 'message': 'SLA processing failed. Check server logs for details.'}, status=500)

@login_required
@user_passes_test(is_admin)
def trigger_cleanup(request):
    """
    Admin-only endpoint to trigger cleanup of inactive users.
    Protected by authentication and role check.
    """
    try:
        call_command('cleanup_inactive_users')
        return JsonResponse({'status': 'ok', 'message': 'Cleanup triggered successfully.'})
    except Exception:
        logger.exception('trigger_cleanup failed')
        return JsonResponse({'status': 'error', 'message': 'Cleanup failed. Check server logs for details.'}, status=500)

# If you still need an external trigger (e.g., for cron jobs), use a secure token:
# Option: Use a secure token stored in environment variables
@csrf_exempt
def trigger_sla_processing_external(request):
    """
    External endpoint for cron jobs. Protected by a secure token.
    Token should be set in environment variables, not hardcoded.

    The token is read from the X-SLA-Trigger-Secret header (preferred, not
    logged/leaked via Referer the way a query string is) with a `secret`
    GET param kept as a fallback for backward compatibility with any
    existing callers. Comparison uses secrets.compare_digest to avoid a
    timing side-channel.
    """
    import os
    import secrets as secrets_module
    secret = request.headers.get('X-SLA-Trigger-Secret') or request.GET.get('secret', '')
    expected_secret = os.environ.get('SLA_TRIGGER_SECRET')

    if not expected_secret:
        return JsonResponse({'error': 'SLA trigger not configured'}, status=500)

    if not secrets_module.compare_digest(secret, expected_secret):
        return JsonResponse({'error': 'Unauthorized'}, status=403)

    from apps.tickets.periodic_tasks import run_sla_job_locked

    try:
        # Shared lock with the process_sla step inside run_periodic_jobs
        # (trigger_periodic_jobs_external) — if a deployment runs both this
        # dedicated SLA cron and the general periodic-jobs cron on
        # independent schedules, this stops them from ever running
        # process_sla concurrently. Returns True if it ran, False if
        # skipped because another trigger was already mid-run.
        ran = run_sla_job_locked()
        return JsonResponse({'status': 'ok' if ran else 'skipped'})
    except Exception:
        logger.exception('trigger_sla_processing_external failed')
        return JsonResponse({'error': 'SLA processing failed. Check server logs for details.'}, status=500)


@csrf_exempt
def trigger_periodic_jobs_external(request):
    """
    External endpoint for a scheduled caller with no long-lived process of
    its own (e.g. a Cloudflare Cron Trigger hitting this on an interval)
    to run all periodic jobs — not just SLA. Runs the same job list as
    `run_periodic_tasks`/`scheduler.py` via the shared
    apps.tickets.periodic_tasks.run_periodic_jobs, so this stays in sync
    with that job list automatically instead of needing its own copy.

    Same auth pattern as trigger_sla_processing_external: a secret in the
    X-SLA-Trigger-Secret header (or `secret` query param), compared with
    secrets.compare_digest.
    """
    import os
    import secrets as secrets_module
    from apps.tickets.periodic_tasks import run_periodic_jobs

    secret = request.headers.get('X-SLA-Trigger-Secret') or request.GET.get('secret', '')
    expected_secret = os.environ.get('SLA_TRIGGER_SECRET')

    if not expected_secret:
        return JsonResponse({'error': 'Periodic job trigger not configured'}, status=500)

    if not secrets_module.compare_digest(secret, expected_secret):
        return JsonResponse({'error': 'Unauthorized'}, status=403)

    try:
        run_periodic_jobs()
        return JsonResponse({'status': 'ok'})
    except Exception:
        logger.exception('trigger_periodic_jobs_external failed')
        return JsonResponse({'error': 'Periodic job run failed. Check server logs for details.'}, status=500)

# ==========================================================================
# PLACEHOLDER / STATIC PAGES
# ==========================================================================

@login_required
def catalogue(request):
    if effective_role_name(request.user) not in ['ADMIN', 'SUPERADMIN']: return HttpResponse(status=403)
    return render(request, 'admin/catalogue.html', {'sidebar_template': get_sidebar_template(request.user)})

@login_required
def connectors(request):
    """
    Admin/Superadmin configuration page for remote connectors (Quick Assist, etc.).
    Lists all connectors and allows editing.
    """
    if not is_admin(request.user):
        return HttpResponse(status=403)
    connectors_list = RemoteConnector.objects.all().order_by('name')
    return render(request, 'admin/connectors.html', {
        'connectors': connectors_list,
        'sidebar_template': get_sidebar_template(request.user),
    })

@login_required
@require_http_methods(['GET', 'POST'])
def connector_edit(request, pk):
    """
    Edit a specific remote connector: enable/disable and update instructions.
    """
    if not is_admin(request.user):
        return HttpResponse(status=403)
    connector = get_object_or_404(RemoteConnector, pk=pk)
    if request.method == 'POST':
        form = ConnectorEditForm(request.POST)
        if form.is_valid():
            connector.is_active = form.cleaned_data['is_active']
            connector.instructions_for_requester = form.cleaned_data['instructions_for_requester']
            connector.instructions_for_agent = form.cleaned_data['instructions_for_agent']
            connector.save()
            messages.success(request, f'"{connector.name}" updated successfully.')
            return redirect('tickets:connectors')
    else:
        form = ConnectorEditForm(initial={
            'is_active': connector.is_active,
            'instructions_for_requester': connector.instructions_for_requester,
            'instructions_for_agent': connector.instructions_for_agent,
        })
    return render(request, 'admin/connector_form.html', {
        'connector': connector,
        'form': form,
        'sidebar_template': get_sidebar_template(request.user),
    })

# ==========================================================================
# ASSET MANAGEMENT
# ==========================================================================

# ==========================================================================
# HELPER: Parse date for asset import
# ==========================================================================

def parse_date(value):
    """Parse date from various formats for asset import."""
    if not value:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, str):
        for fmt in ['%Y-%m-%d', '%m/%d/%Y', '%d/%m/%Y', '%b %d, %Y']:
            try:
                return datetime.strptime(value, fmt).date()
            except ValueError:
                continue
    return None

# ==========================================================================
# ASSET LIST
# ==========================================================================

# Worst-case severity tiers for a seat/workspace card's summary badges —
# most severe first. A card shows the single most-severe status/condition
# among its items rather than every item's own badge, so "one damaged UPS"
# is never hidden behind two other perfectly-fine items in the same group.
_ASSET_STATUS_SEVERITY_TIERS = [
    {Asset.Status.DAMAGED},
    {Asset.Status.MAINTENANCE, Asset.Status.REPAIR},
    {Asset.Status.LOST, Asset.Status.STOLEN},
    {Asset.Status.IN_USE, Asset.Status.MOBILIZED},
    {Asset.Status.IN_STORE, Asset.Status.READY},
]
_ASSET_CONDITION_SEVERITY_TIERS = [
    {Asset.Condition.UNUSABLE}, {Asset.Condition.DAMAGED}, {Asset.Condition.POOR},
    {Asset.Condition.FAIR}, {Asset.Condition.GOOD}, {Asset.Condition.EXCELLENT},
]


def _severity_rank(value, tiers):
    for rank, tier in enumerate(tiers):
        if value in tier:
            return rank
    return len(tiers)  # unlisted/unknown values sort as least severe


def _build_asset_groups(asset_qs):
    """Groups an already-ordered Asset queryset/list into workspace/seat
    groups — one person's (or one unmatched-import name's, or one
    department+location pool's) full kit as a single unit, for the
    collapsible seat-card view. Same 3-tier grouping key as the old
    per-row annotation this replaces (assigned user -> unresolved import
    hint -> department+location pool), but now returns real groups (each
    with its own asset list) instead of tagging a flat sequence — needed
    so a card can show its own item count and worst-case status/condition
    before rendering. Must run on the FULL filtered queryset (not a single
    page) so a group can never split across a pagination boundary; the
    view paginates the returned group list itself, not the raw queryset.
    Returns a list of dicts: {label, seat_code, assets, item_count,
    worst_status, worst_condition}, in the queryset's own order."""
    groups = []
    groups_by_key = {}
    for asset in asset_qs:
        if asset.assigned_to_id:
            key = ('user', asset.assigned_to_id)
            label = asset.assigned_to.get_full_name() or asset.assigned_to.email
        elif asset.unresolved_assignee_hint:
            key = ('hint', asset.unresolved_assignee_hint)
            # "Needs review" rather than "Unmatched" — the latter reads like
            # a system error to the non-technical staff this view is for;
            # it just means the imported name hasn't been linked to a login yet.
            label = f'Needs review: {asset.unresolved_assignee_hint}'
        else:
            key = ('pool', asset.department_id, asset.location_id)
            dept = asset.department.name if asset.department else 'No department'
            loc = asset.location.full_name() if asset.location else 'No location'
            label = f'Unassigned — {dept} / {loc}'

        group = groups_by_key.get(key)
        if group is None:
            group = {'key': key, 'label': label, 'seat_code': None, 'assets': []}
            groups_by_key[key] = group
            groups.append(group)
        group['assets'].append(asset)
        # tag_slot_number is shared by every device for the same
        # assigned_to+department pair (see Asset._resolve_tag_slot_number) —
        # the closest existing DB concept to the client's "seat" id (e.g.
        # PLD-003). Null for legacy-format assets, in which case the group
        # just falls back to its owner/hint/pool label above.
        if not group['seat_code'] and asset.tag_slot_number and asset.department_id:
            dept_code = asset.department.tag_code or asset.department.name
            group['seat_code'] = f'{dept_code}-{asset.tag_slot_number:03d}'

    for group in groups:
        group['item_count'] = len(group['assets'])
        group['asset_ids'] = [str(a.pk) for a in group['assets']]
        group['worst_status'] = min(
            group['assets'], key=lambda a: _severity_rank(a.status, _ASSET_STATUS_SEVERITY_TIERS)
        ).status
        conditioned = [a for a in group['assets'] if a.condition]
        group['worst_condition'] = (
            min(conditioned, key=lambda a: _severity_rank(a.condition, _ASSET_CONDITION_SEVERITY_TIERS)).condition
            if conditioned else None
        )
        # Deduplicated, first-seen-order device-type summary for the
        # collapsed card (e.g. "CPU, Monitor, UPS") — cheaper and less
        # fragile than a template-side {% regroup %} (which silently
        # produces wrong groupings if the list isn't already sorted by the
        # grouped attribute).
        category_names = []
        seen = set()
        for a in group['assets']:
            name = a.category.name if a.category_id else 'Other'
            if name not in seen:
                seen.add(name)
                category_names.append(name)
        group['category_summary'] = ', '.join(category_names)
    return groups


def get_asset_kpis():
    """Shared Asset Inventory KPI counts — used by both the Overview
    dashboard and the asset inventory page's stat strip, so the two never
    drift apart. Always global/unfiltered (matches the dashboard's own
    semantics); the inventory page's own filters don't affect these."""
    total_assets = Asset.objects.count()
    # 'ACTIVE' isn't a real Asset.Status value — "active" here means
    # currently in use, matching Asset.is_active's own definition.
    active_assets = Asset.objects.filter(
        status__in=[Asset.Status.IN_USE, Asset.Status.MOBILIZED]
    ).count()
    in_store_assets = Asset.objects.filter(status=Asset.Status.IN_STORE).count()
    maintenance_assets = Asset.objects.filter(status=Asset.Status.MAINTENANCE).count()
    damaged_assets = Asset.objects.filter(status=Asset.Status.DAMAGED).count()
    scrapped_assets = Asset.objects.filter(status=Asset.Status.SCRAPPED).count()

    thirty_days_ago = timezone.now() - timedelta(days=30)
    recently_added = Asset.objects.filter(created_at__gte=thirty_days_ago).count()
    assigned_assets = Asset.objects.filter(assigned_to__isnull=False).count()
    unassigned_assets = total_assets - assigned_assets

    today = timezone.now().date()
    ninety_days_later = today + timedelta(days=90)
    expiring_warranty = Asset.objects.filter(
        warranty_expiry__gte=today,
        warranty_expiry__lte=ninety_days_later,
    ).exclude(
        status__in=[Asset.Status.RETIRED, Asset.Status.SCRAPPED, Asset.Status.LOST, Asset.Status.STOLEN, Asset.Status.DISPOSED]
    ).count()

    return {
        'total_assets': total_assets,
        'active_assets': active_assets,
        'in_store_assets': in_store_assets,
        'maintenance_assets': maintenance_assets,
        'damaged_assets': damaged_assets,
        'scrapped_assets': scrapped_assets,
        'recently_added': recently_added,
        'assigned_assets': assigned_assets,
        'unassigned_assets': unassigned_assets,
        'expiring_warranty': expiring_warranty,
    }


_ASSET_EQUIPMENT_SORT_OPTIONS = {
    'owner': ((
        'assigned_to__last_name', 'assigned_to__first_name',
        'unresolved_assignee_hint', 'department__name', 'location__name',
        'category__name', 'name',
    ), 'Owner (default grouping)'),
    '-created_at': (('-created_at',), 'Recently Added'),
    '-updated_at': (('-updated_at',), 'Recently Updated'),
    'name': (('name',), 'Name (A-Z)'),
}


def _build_equipment_context(request):
    """Context for the Inventory page's Equipment tab — physical assets
    only. Renewable assets (software licenses, subscriptions, support
    contracts) are excluded entirely; see _build_license_context for their
    own tab, since a physical seat card's item count/condition badge is
    meaningless for a subscription."""
    query = request.GET.get('filter_q', '')
    category_filter = request.GET.get('filter_category', '')
    status_filter = request.GET.get('filter_status', '')
    location_filter = request.GET.get('filter_location', '')
    low_stock_filter = request.GET.get('filter_low_stock', '')
    # Defaults on: absent (first load) reads as '1'; only an explicit '0'
    # from the hidden-fallback input (see asset_list.html) turns it off.
    group_by_owner = request.GET.get('filter_group_by_owner', '1') != '0'

    assets_list = Asset.objects.select_related(
        'category', 'assigned_to', 'checked_out_to', 'department', 'location'
    ).exclude(category__is_renewable=True)
    # Default ordering matches prior behavior exactly (grouped-by-owner vs.
    # flat-by-recently-added); a `sort=` override on top of that — e.g.
    # `-updated_at` — is what lets an admin float a just-edited asset back
    # to page 1 instead of losing it in whatever page they last had open.
    default_sort_key = 'owner' if group_by_owner else '-created_at'
    order_args, active_sort, sort_options = resolve_sort(request, _ASSET_EQUIPMENT_SORT_OPTIONS, default_sort_key)
    assets_list = assets_list.order_by(*order_args)

    if query:
        assets_list = assets_list.filter(
            Q(name__icontains=query) |
            Q(tracking_id__icontains=query) |
            Q(serial_number__icontains=query) |
            Q(model__icontains=query) |
            Q(manufacturer__icontains=query) |
            Q(assigned_to__first_name__icontains=query) |
            Q(assigned_to__last_name__icontains=query) |
            Q(checked_out_to__first_name__icontains=query) |
            Q(checked_out_to__last_name__icontains=query) |
            Q(unresolved_assignee_hint__icontains=query)
        )
    if category_filter:
        assets_list = assets_list.filter(category_id=category_filter)
    if status_filter:
        assets_list = assets_list.filter(status=status_filter)
    if location_filter:
        assets_list = assets_list.filter(location_id=location_filter)
    if low_stock_filter:
        assets_list = assets_list.filter(
            category__is_consumable=True,
            low_stock_threshold__isnull=False,
            quantity_in_stock__lte=F('low_stock_threshold'),
        )

    page_number = request.GET.get('page')
    if group_by_owner:
        # Group on the FULL filtered/ordered queryset, then paginate the
        # resulting group list itself — a seat's items can no longer split
        # across a page boundary the way the old per-page annotation could.
        groups = _build_asset_groups(assets_list)
        paginator = Paginator(groups, 10)
    else:
        paginator = Paginator(assets_list, 10)
    page_obj = paginator.get_page(page_number)

    users = User.objects.filter(is_active=True).order_by('first_name', 'last_name')

    return {
        'assets': page_obj,
        'group_by_owner': group_by_owner,
        'group_by_owner_param': '1' if group_by_owner else '0',
        'users': users,
        'categories': AssetCategory.objects.filter(is_renewable=False).order_by('name'),
        'status_choices': Asset.Status.choices,
        'status_values': [v for v, _ in Asset.Status.choices],
        'location_choices': [(loc.pk, loc.full_name()) for loc in Location.objects.filter(is_active=True)],
        'location_values': [loc.pk for loc in Location.objects.filter(is_active=True)],
        'query': query,
        'selected_category': category_filter,
        'selected_status': status_filter,
        'selected_location': location_filter,
        'selected_low_stock': low_stock_filter,
        'today': timezone.now().date(),
        'sort_options': sort_options,
        'active_sort': active_sort,
    }


_ASSET_LICENSE_SORT_OPTIONS = {
    'renewal': ((F('next_renewal_date').asc(nulls_last=True), 'name'), 'Renewal Date (default)'),
    '-updated_at': (('-updated_at',), 'Recently Updated'),
    'name': (('name',), 'Name (A-Z)'),
}


def _build_license_context(request):
    """Context for the Inventory page's Licenses & Subscriptions tab —
    renewable assets only (software licenses, SaaS subscriptions, support
    contracts), sorted by renewal urgency rather than grouped by owner.
    Budget/renewal-urgency aggregates are deliberately global/unfiltered
    (same convention as get_asset_kpis' dashboard-equivalent counts) so the
    summary strip doesn't jump around as someone types a search."""
    query = request.GET.get('filter_q', '')
    category_filter = request.GET.get('filter_category', '')
    due_soon_filter = request.GET.get('filter_due_soon', '')
    today = timezone.now().date()

    license_qs = Asset.objects.filter(category__is_renewable=True).select_related(
        'category', 'assigned_to', 'department', 'renewal_vendor'
    )
    if query:
        license_qs = license_qs.filter(
            Q(name__icontains=query) |
            Q(renewal_vendor__name__icontains=query) |
            Q(assigned_to__first_name__icontains=query) |
            Q(assigned_to__last_name__icontains=query)
        )
    if category_filter:
        license_qs = license_qs.filter(category_id=category_filter)
    if due_soon_filter:
        license_qs = license_qs.filter(next_renewal_date__isnull=False, next_renewal_date__lte=today + timedelta(days=30))
    order_args, active_sort, sort_options = resolve_sort(request, _ASSET_LICENSE_SORT_OPTIONS, 'renewal')
    license_qs = license_qs.order_by(*order_args)

    paginator = Paginator(license_qs, 10)
    page_obj = paginator.get_page(request.GET.get('page'))

    renewable_qs = Asset.objects.filter(category__is_renewable=True)
    costed_count = renewable_qs.filter(renewal_cost__isnull=False).count()
    total_count = renewable_qs.count()
    due_soon_count = renewable_qs.filter(
        next_renewal_date__isnull=False, next_renewal_date__gte=today, next_renewal_date__lte=today + timedelta(days=30)
    ).count()
    overdue_count = renewable_qs.filter(next_renewal_date__lt=today).count()

    # Currency is set per-asset (see Asset.renewal_currency), not org-wide, so
    # a plain Sum('renewal_cost') across assets would silently add e.g. a
    # $100 line and a ₦20,000 line into a meaningless "20100". Group by
    # currency instead: one total per currency, and only collapse to a single
    # labeled figure when every costed asset actually agrees on one.
    currency_breakdown = list(
        renewable_qs.filter(renewal_cost__isnull=False)
        .values('renewal_currency').annotate(total=Sum('renewal_cost')).order_by('-total')
    )
    budget_mixed_currencies = len(currency_breakdown) > 1
    if len(currency_breakdown) == 1:
        budget_currency = currency_breakdown[0]['renewal_currency']
        total_cost = currency_breakdown[0]['total']
    else:
        budget_currency = ''
        total_cost = None

    # Per-renewal audit trail — every actual renewal (auto or manual), for
    # a spend history under the table (distinct from the forward-looking
    # "what's currently costed" KPI strip above; this is "what did we
    # actually pay, and when" for a budget summary at year-end).
    budget_renewal_logs = AssetLog.objects.filter(
        action=AssetLog.Action.RENEWED, asset__category__is_renewable=True
    ).select_related('asset', 'actor').order_by('-created_at')[:100]

    return {
        'assets': page_obj,
        'categories': AssetCategory.objects.filter(is_renewable=True).order_by('name'),
        'query': query,
        'selected_category': category_filter,
        'selected_due_soon': due_soon_filter,
        'today': today,
        'license_total_cost': total_cost,
        'license_costed_count': costed_count,
        'license_total_count': total_count,
        'license_due_soon_count': due_soon_count,
        'license_overdue_count': overdue_count,
        'license_budget_currency': budget_currency,
        'license_budget_mixed_currencies': budget_mixed_currencies,
        'license_budget_breakdown': currency_breakdown,
        'budget_renewal_logs': budget_renewal_logs,
        'sort_options': sort_options,
        'active_sort': active_sort,
    }


@login_required
def assets(request):
    """Full asset inventory — Admin/Superadmin only. Every other role only
    ever sees assets assigned to them, via my_assets. Two tabs sharing this
    one view/URL: Equipment (physical assets, default) and Licenses &
    Subscriptions (renewable assets) — see _build_equipment_context/
    _build_license_context."""
    if effective_role_name(request.user) not in ('ADMIN', 'SUPERADMIN'):
        return HttpResponse(status=403)

    active_tab = request.GET.get('filter_tab', 'equipment')
    if active_tab not in ('equipment', 'licenses'):
        active_tab = 'equipment'

    context = _build_license_context(request) if active_tab == 'licenses' else _build_equipment_context(request)
    context['active_tab'] = active_tab
    context['sidebar_template'] = get_sidebar_template(request.user)
    # Needed on both tabs — the tab switcher's Licenses badge count must
    # show even while viewing Equipment. Cheap count query either way.
    context['license_total_count'] = Asset.objects.filter(category__is_renewable=True).count()
    context.update(get_asset_kpis())

    from .report_registry import REPORT_TYPES
    context['export_columns'] = REPORT_TYPES['assets'].columns
    context['export_column_help'] = REPORT_TYPES['assets'].column_help

    if request.headers.get('HX-Request'):
        # The tab switcher (asset_tabs.html) targets #assetPanel and needs
        # the whole tabs+filter-form+results block re-rendered; every other
        # HTMX trigger on this page (search/category/pagination/reset)
        # targets #assetTableContainer nested inside that block and only
        # wants the results — re-rendering the full panel there would
        # nest a second copy of the filter form/tabs inside that small
        # container.
        if request.headers.get('HX-Target') == 'assetPanel':
            return render(request, 'partials/asset_panel.html', context)
        table_template = 'partials/asset_license_table.html' if active_tab == 'licenses' else 'partials/asset_table.html'
        return render(request, table_template, context)

    return render(request, 'tickets/asset_list.html', context)


# ==========================================================================
# MY ASSETS — assets allocated to the logged-in user, any role. Separate
# from `assets()` above (IT-only asset management) — this is a read-mostly
# view of "what's assigned to me", including any maintenance scheduled or
# awaiting the owner's confirmation on those assets.
# ==========================================================================

def _my_assets_pending_confirmations_q(user):
    """Rows the given user is eligible to confirm as an asset OWNER — used
    by both my_assets() and my_assets_pending_count(). Deliberately does NOT
    include the department-Team-Lead fallback or the Admin/Superadmin
    override (see apps.maintenance.views.can_confirm_asset_maintenance) —
    that's a separate category, handled by
    _shared_asset_pending_confirmations_q below."""
    return MaintenanceAssetConfirmation.objects.filter(
        asset__assigned_to=user, status=MaintenanceAssetConfirmation.Status.PENDING,
    )


def _shared_asset_pending_confirmations_q(user):
    """Rows the IT Team Lead is eligible to confirm as the fallback for
    shared, ownerless pool inventory — regardless of which department code
    the asset itself is tagged with, since unassigned equipment is IT-
    managed inventory (see apps.maintenance.views.can_confirm_asset_maintenance).
    Empty for anyone who isn't currently acting as the IT Team Lead.
    Admin/Superadmin's override isn't surfaced here — they already have
    full IT-internal access via the maintenance schedule detail page."""
    if effective_role_name(user) != 'TEAM_LEAD' or user.department != 'IT':
        return MaintenanceAssetConfirmation.objects.none()
    return MaintenanceAssetConfirmation.objects.filter(
        asset__assigned_to__isnull=True,
        status=MaintenanceAssetConfirmation.Status.PENDING,
    )


def _shared_asset_resolved_confirmations_q(user):
    """The history counterpart to _shared_asset_pending_confirmations_q — a
    resolved (CONFIRMED/DISPUTED) row drops off the pending list the moment
    it's acted on, so without this the IT Team Lead has nowhere to see that
    a shared asset was ever maintained at all."""
    if effective_role_name(user) != 'TEAM_LEAD' or user.department != 'IT':
        return MaintenanceAssetConfirmation.objects.none()
    return MaintenanceAssetConfirmation.objects.filter(
        asset__assigned_to__isnull=True,
    ).exclude(status=MaintenanceAssetConfirmation.Status.PENDING)


@login_required
def my_assets(request):
    """Assets assigned to the current user, with any upcoming/in-progress
    maintenance and any completion awaiting their confirmation surfaced
    inline — open to every role, since asset ownership isn't role-bound."""
    assets_list = Asset.objects.filter(assigned_to=request.user).select_related('category').prefetch_related(
        Prefetch(
            'maintenance_confirmations',
            queryset=MaintenanceAssetConfirmation.objects.select_related('schedule', 'confirmed_by'),
        ),
        'maintenance_schedules',
    ).order_by('name')

    rows = []
    for asset in assets_list:
        confirmations = list(asset.maintenance_confirmations.all())
        pending_confirmations = [
            c for c in confirmations if c.status == MaintenanceAssetConfirmation.Status.PENDING
        ]
        # Most recent resolved confirmation — the maintenance history trail
        # a requester sees for their own asset, independent of whether a
        # newer maintenance cycle is currently pending confirmation above.
        confirmed_history = sorted(
            (c for c in confirmations if c.status == MaintenanceAssetConfirmation.Status.CONFIRMED and c.confirmed_at),
            key=lambda c: c.confirmed_at, reverse=True,
        )
        upcoming_schedules = [
            s for s in asset.maintenance_schedules.all()
            if s.status in (MaintenanceSchedule.Status.SCHEDULED, MaintenanceSchedule.Status.IN_PROGRESS)
        ]
        rows.append({
            'asset': asset,
            'pending_confirmations': pending_confirmations,
            'upcoming_schedules': upcoming_schedules,
            'last_maintained': confirmed_history[0] if confirmed_history else None,
            'open_history': asset._open_checkout_history(),
        })

    shared_asset_confirmations = list(
        _shared_asset_pending_confirmations_q(request.user).select_related('asset', 'schedule')
    )
    shared_asset_history = list(
        _shared_asset_resolved_confirmations_q(request.user)
        .select_related('asset', 'schedule', 'confirmed_by')
        .order_by('-confirmed_at')[:20]
    )

    context = {
        'rows': rows,
        'shared_asset_confirmations': shared_asset_confirmations,
        'shared_asset_history': shared_asset_history,
        'today': timezone.now().date(),
        'sidebar_template': get_sidebar_template(request.user),
    }
    return render(request, 'tickets/my_assets.html', context)


@login_required
def my_assets_pending_count(request):
    """Badge count for the sidebar 'My Assets' link — mirrors the pattern
    used by remote_session_pending_count. Includes the IT Team Lead's
    shared-asset fallback confirmations alongside their own owned assets, so
    the badge matches everything my_assets() actually surfaces to them."""
    count = (
        _my_assets_pending_confirmations_q(request.user).count()
        + _shared_asset_pending_confirmations_q(request.user).count()
    )
    return render(request, 'partials/my_assets_badge.html', {'count': count})


# ==========================================================================
# DEMOBILIZATION (requester self-report — mirrors My Assets / asset return
# request. Open to every role, since being a ticket requester isn't
# role-bound, same posture as my_assets above.)
# ==========================================================================

def _demobilization_eligible_q(user):
    """Items mobilized to this user's own tickets, confirmed received, and
    not yet demobilized by an admin — the pool eligible for self-report."""
    return MobilizationItem.objects.filter(
        mobilization__ticket__requester=user,
        acknowledged_at__isnull=False,
        demobilized_at__isnull=True,
    ).select_related('asset', 'mobilization').order_by('mobilization_id', 'asset__name')


@login_required
def demobilization_list(request):
    """Two-tab view of the requester's own mobilized assets: 'Active' —
    grouped by job, anything ready to report or awaiting admin
    confirmation — and 'History' — everything already demobilized, also
    grouped by job (so items from different jobs never run together),
    with a search box to narrow within/across groups. Kept as separate
    tabs (rather than mixed into one list) so the action-needed items
    never get buried under an ever-growing return log; History persists
    forever so the requester always has a timestamped record to point to."""
    active_items = list(_demobilization_eligible_q(request.user))
    history_items = list(
        MobilizationItem.objects.filter(
            mobilization__ticket__requester=request.user,
            demobilized_at__isnull=False,
        ).select_related('asset', 'mobilization', 'mobilization__ticket').order_by('-demobilized_at')
    )

    active_groups_map = {}
    for item in active_items:
        mob = item.mobilization
        bucket = active_groups_map.setdefault(mob.pk, {'mobilization': mob, 'ready': [], 'pending': []})
        if item.return_requested_at:
            bucket['pending'].append(item)
        else:
            bucket['ready'].append(item)
    active_groups = sorted(active_groups_map.values(), key=lambda g: g['mobilization'].mobilized_at, reverse=True)
    active_count = sum(len(g['ready']) + len(g['pending']) for g in active_groups)

    history_groups_map = {}
    for item in history_items:
        mob = item.mobilization
        if mob.pk not in history_groups_map:
            history_groups_map[mob.pk] = {
                'mobilization': mob,
                'items': [],
                'vessel_names': list(mob.vessels.values_list('name', flat=True)),
                'dive_system_names': list(mob.dive_systems.values_list('name', flat=True)),
            }
        history_groups_map[mob.pk]['items'].append(item)
    history_groups = sorted(
        history_groups_map.values(), key=lambda g: g['items'][0].demobilized_at, reverse=True
    )

    context = {
        'active_groups': active_groups,
        'active_count': active_count,
        'history_groups': history_groups,
        'history_count': len(history_items),
        'sidebar_template': get_sidebar_template(request.user),
    }
    return render(request, 'tickets/demobilization_list.html', context)


@login_required
def demobilization_pending_count(request):
    """Badge count for the sidebar 'Demobilization' link — items ready to
    report that haven't been reported yet (the nudge, same semantics as
    my_assets_pending_count)."""
    count = _demobilization_eligible_q(request.user).filter(return_requested_at__isnull=True).count()
    return render(request, 'partials/sidebar_count_badge.html', {'count': count})


@login_required
@require_POST
def mobilization_items_request_demobilize_batch(request):
    """Requester reports a batch of items as sent back, sharing one notes
    field across the batch. One bad/ineligible id is skipped rather than
    failing the whole batch."""
    item_ids = request.POST.getlist('item_ids')
    notes = request.POST.get('notes', '').strip()

    if not item_ids:
        messages.error(request, 'Select at least one item to report.')
        return redirect('tickets:demobilization_list')

    reported, skipped = 0, 0
    for item_pk in item_ids:
        try:
            item = MobilizationItem.objects.select_related('mobilization', 'asset').get(pk=item_pk)
            item.request_demobilization(actor=request.user, notes=notes)
            reported += 1
        except (MobilizationItem.DoesNotExist, ValueError):
            skipped += 1

    if reported:
        messages.success(request, f'Reported {reported} asset(s) as demobilized — awaiting admin confirmation.')
    if skipped:
        messages.warning(request, f'{skipped} item(s) could not be reported (already actioned or not yours).')
    return redirect('tickets:demobilization_list')


@login_required
@require_POST
def mobilization_item_cancel_demobilize_request(request, item_pk):
    item = get_object_or_404(MobilizationItem.objects.select_related('mobilization'), pk=item_pk)
    try:
        item.cancel_demobilization_request(actor=request.user)
    except ValueError as e:
        messages.error(request, str(e))
        return redirect('tickets:demobilization_list')
    messages.success(request, f'Demobilization request for "{item.asset.name}" cancelled.')
    return redirect('tickets:demobilization_list')


@login_required
def pending_demobilizations_list(request):
    """Admin-side queue: items the requester has self-reported as returned
    but an admin hasn't yet confirmed physical receipt — mirrors
    pending_asset_returns_list."""
    if effective_role_name(request.user) not in ['ADMIN', 'SUPERADMIN']:
        return HttpResponse(status=403)

    items_qs = MobilizationItem.objects.filter(
        return_requested_at__isnull=False, demobilized_at__isnull=True
    ).select_related('asset', 'mobilization', 'return_requested_by').order_by('return_requested_at')

    paginator = Paginator(items_qs, 10)
    page_obj = paginator.get_page(request.GET.get('page'))

    return render(request, 'tickets/pending_demobilizations_list.html', {
        'items': page_obj,
        'sidebar_template': get_sidebar_template(request.user),
    })


@login_required
def pending_demobilizations_count(request):
    if effective_role_name(request.user) not in ['ADMIN', 'SUPERADMIN']:
        return HttpResponse(status=403)

    count = MobilizationItem.objects.filter(
        return_requested_at__isnull=False, demobilized_at__isnull=True
    ).count()
    return render(request, 'partials/sidebar_count_badge.html', {'count': count})


# ==========================================================================
# ASSET CREATE PAGE (Dedicated Page)
# ==========================================================================

@login_required
@require_http_methods(['GET', 'POST'])
def asset_create_page(request):
    if effective_role_name(request.user) not in ['ADMIN', 'SUPERADMIN']:
        return HttpResponse(status=403)

    if request.method == 'POST':
        form = AssetForm(request.POST)
        if form.is_valid():
            asset = form.save(commit=False)
            asset.tracking_id = None
            asset.save()
            
            # Create CREATED log
            AssetLog.objects.create(
                asset=asset,
                action=AssetLog.Action.CREATED,
                actor=request.user,
                details={'name': asset.name, 'category': asset.category.name if asset.category else None}
            )
            
            # Create ASSIGNED log if assigned to someone
            if asset.assigned_to:
                AssetLog.objects.create(
                    asset=asset,
                    action=AssetLog.Action.ASSIGNED,
                    actor=request.user,
                    details={
                        'from': None,
                        'to': asset.assigned_to.get_full_name() if asset.assigned_to else None,
                        'comment': 'Initial assignment'
                    }
                )
            
            messages.success(request, f'Asset "{asset.name}" created successfully!')
            return redirect('tickets:assets')
        else:
            messages.error(request, 'Please correct the errors below.')
    else:
        form = AssetForm()

    context = {
        'form': form,
        'asset': None,
        'action': 'create',
        'users': User.objects.filter(is_active=True).order_by('first_name', 'last_name'),
        'categories': AssetCategory.objects.all().order_by('name'),
        'status_choices': Asset.Status.choices,
        'status_values': [v for v, _ in Asset.Status.choices],
        'sidebar_template': get_sidebar_template(request.user),
    }
    return render(request, 'tickets/asset_form_page.html', context)


# ==========================================================================
# ASSET EDIT PAGE (Dedicated Page)
# ==========================================================================

@login_required
@require_http_methods(['GET', 'POST'])
def asset_edit_page(request, pk):
    if effective_role_name(request.user) not in ['ADMIN', 'SUPERADMIN']:
        return HttpResponse(status=403)

    asset = get_object_or_404(Asset, pk=pk)

    if request.method == 'POST':
        form = AssetForm(request.POST, instance=asset)
        if form.is_valid():
            # assigned_to is disabled on this form for an existing asset
            # (see AssetForm.__init__) — custody changes only happen
            # through assign_to()/release() now (Checkout/Reassign/Check-in),
            # so this save can never touch who has the asset.
            asset = form.save()
            asset.refresh_low_stock_alert()

            AssetLog.objects.create(
                asset=asset,
                action=AssetLog.Action.UPDATED,
                actor=request.user,
                details={'source': 'edit_page'}
            )

            messages.success(request, f'Asset "{asset.name}" updated successfully!')
            
            # Preserve filters/sort/page when redirecting back — param names
            # here must match what _build_equipment_context actually reads
            # (filter_q/filter_category/etc., not q/category/etc.), otherwise
            # this silently no-ops and every edit bounces back to page 1 of
            # the default view regardless of where the admin edited from.
            source = request.GET.get('source', 'list')
            redirect_url = reverse('tickets:assets')
            if source == 'list':
                preserved_params = ['filter_q', 'filter_category', 'filter_status', 'filter_location',
                                     'filter_group_by_owner', 'filter_tab', 'sort', 'page']
                params = [f'{name}={request.GET[name]}' for name in preserved_params if request.GET.get(name)]
                if params:
                    redirect_url += '?' + '&'.join(params)

            return redirect(redirect_url)
        else:
            messages.error(request, 'Please correct the errors below.')
    else:
        form = AssetForm(instance=asset)

    context = {
        'form': form,
        'asset': asset,
        'action': 'edit',
        'users': User.objects.filter(is_active=True).order_by('first_name', 'last_name'),
        'categories': AssetCategory.objects.all().order_by('name'),
        'status_choices': Asset.Status.choices,
        'status_values': [v for v, _ in Asset.Status.choices],
        'sidebar_template': get_sidebar_template(request.user),
    }
    return render(request, 'tickets/asset_form_page.html', context)

# ==========================================================================
# ASSET REASSIGN
# ==========================================================================

@login_required
@require_POST
def asset_reassign(request, pk):
    if effective_role_name(request.user) not in ('ADMIN', 'SUPERADMIN'):
        return HttpResponse(status=403)

    asset = get_object_or_404(Asset, pk=pk)

    # Prevent the action rather than fail after the fact: Reassign only
    # makes sense when someone actually has the asset. An asset with no
    # current holder should be Checked Out, not "re"-assigned. A currently-
    # mobilized asset must be demobilized first — reassigning it would clear
    # status/holder out from under the still-open MobilizationItem, leaving
    # it orphaned (see Asset.mobilization_blocked_reason).
    if not asset.can_reassign:
        blocked = asset.mobilization_blocked_reason
        if blocked:
            message = f'{blocked} before reassigning.'
        else:
            message = f'"{asset.name}" cannot be reassigned — it has no current holder. Use Checkout instead.'
        if request.headers.get('HX-Request'):
            return HttpResponse(message, status=400)
        messages.error(request, message)
        return redirect('tickets:assets')

    form = AssetReassignForm(request.POST)
    if not form.is_valid():
        if request.headers.get('HX-Request'):
            return render(request, 'partials/asset_reassign_modal.html', {'asset': asset, 'form': form})
        messages.error(request, 'Please correct the errors below.')
        return redirect('tickets:assets')

    new_user = form.cleaned_data['assigned_to']
    comment = form.cleaned_data['comment']

    old_holder = asset.assigned_to or asset.checked_out_to
    old_name = old_holder.get_full_name() if old_holder else 'Unassigned'
    actor_name = request.user.get_full_name()
    new_name = new_user.get_full_name() if new_user else 'Unassigned'
    new_user_id = new_user.pk if new_user else None

    # release()/assign_to() are the single source of truth for who has this
    # asset — routing reassignment through them (instead of setting
    # assigned_to directly) keeps checked_out_to/status/AssetCheckoutHistory
    # in lockstep with it, so reassigning can no longer leave the asset
    # claiming two different current holders at once. The ValueError catch
    # sits outside the atomic block so a failed assign_to() rolls back the
    # release() that just happened, rather than leaving the asset unassigned.
    # can_reassign no longer guarantees old_holder is set — it also allows
    # reassigning an asset whose status implies a holder but has none (e.g.
    # an imported row left unassigned). release() and the notification
    # below both tolerate old_holder being None; release() still runs
    # unconditionally to normalize status/holder fields before assign_to().
    try:
        with transaction.atomic():
            asset.release(actor=request.user, return_reason=Asset.ReturnReason.OTHER, return_comment=comment or 'Reassigned')
            if new_user:
                asset.assign_to(new_user, actor=request.user, notes=comment, previous_holder_name=old_name)
    except ValueError as e:
        if request.headers.get('HX-Request'):
            return HttpResponse(str(e), status=400)
        messages.error(request, str(e))
        return redirect('tickets:assets')

    # The old holder didn't initiate this (unlike a self-requested return),
    # so unlike request_return()/release()'s other callers, nothing else
    # tells them the asset left their hands — do it here. assign_to()
    # already notifies the new recipient internally, so no duplicate needed.
    # Skipped entirely when there was no old holder to begin with (e.g.
    # claiming an imported asset that was never actually assigned to anyone).
    if old_holder:
        Notification.objects.create(
            recipient=old_holder, role=role_of(old_holder),
            message=f'"{asset.name}" ({asset.tracking_id}) has been reassigned away from you'
                    f'{" to " + new_name if new_user else ""}.',
            url='/tickets/my-assets/',
        )

    # Add comment to asset notes (user can edit the default comment)
    if new_user_id:
        comment_body = f"**Asset reassigned** by {actor_name} from **{old_name}** to **{new_name}**.\n\n**Reason:** {comment}"
    else:
        comment_body = f"**Asset unassigned** by {actor_name} from **{old_name}**.\n\n**Reason:** {comment}"
    
    if asset.notes:
        asset.notes = f"{asset.notes}\n\n{comment_body}"
    else:
        asset.notes = comment_body
    asset.save(update_fields=['notes'])

    if new_user_id:
        success_message = f'"{asset.name}" reassigned to {new_name}.'
    else:
        success_message = f'"{asset.name}" unassigned.'

    if request.headers.get('HX-Request'):
        messages.success(request, success_message)
        return HttpResponse(status=204, headers={'HX-Redirect': reverse('tickets:assets')})
    messages.success(request, success_message)
    return redirect('tickets:assets')

@login_required
def asset_reassign_modal(request, pk):
    """Returns the asset reassign modal with pre-filled comment."""
    if effective_role_name(request.user) not in ('ADMIN', 'SUPERADMIN'):
        return HttpResponse(status=403)

    asset = get_object_or_404(Asset, pk=pk)
    if not asset.can_reassign:
        return HttpResponse(
            f'<div class="p-4 text-center text-text-secondary">"{asset.name}" has no current holder, so it can\'t be reassigned — use Checkout instead.</div>',
            status=400,
        )
    users = User.objects.filter(is_active=True).order_by('first_name', 'last_name')

    # Generate default comment
    old_name = asset.assigned_to.get_full_name() if asset.assigned_to else 'Unassigned'
    default_comment = f"Reassigning asset from {old_name} to [new user]."
    form = AssetReassignForm(initial={'comment': default_comment})

    return render(request, 'partials/asset_reassign_modal.html', {
        'asset': asset,
        'users': users,
        'default_comment': default_comment,
        'form': form,
    })

# ==========================================================================
# ASSET DETAIL
# ==========================================================================

@login_required
def asset_detail(request, pk):
    if effective_role_name(request.user) not in ('ADMIN', 'SUPERADMIN'):
        return HttpResponse(status=403)
    
    asset = get_object_or_404(Asset, pk=pk)
    logs = asset.logs.all()[:10]  # Recent activity
    renewal_logs = asset.logs.filter(action=AssetLog.Action.RENEWED)[:10] if asset.is_renewable else []

    from apps.maintenance.models import AssetBackupStatus
    import json

    return render(request, 'tickets/asset_detail.html', {
        'asset': asset,
        'logs': logs,
        'renewal_logs': renewal_logs,
        'attachments': asset.attachments.all(),
        'maintenance_confirmations': asset.maintenance_confirmations.select_related(
            'schedule', 'confirmed_by'
        ).order_by('-technician_completed_at'),
        'checkout_history': asset.checkout_history.select_related(
            'checked_out_by', 'checked_out_to'
        ).order_by('-checked_out_at'),
        'backup_status_choices': json.dumps(list(AssetBackupStatus.Status.choices)),
        'today': timezone.now().date(),
        'sidebar_template': get_sidebar_template(request.user),
    })


@login_required
@require_POST
def asset_mark_renewed(request, pk):
    """Admin action: the license/subscription was actually paid/renewed —
    advances next_renewal_date, resets reminder flags, logs an audit entry."""
    if effective_role_name(request.user) not in ['ADMIN', 'SUPERADMIN']:
        return HttpResponse(status=403)

    asset = get_object_or_404(Asset, pk=pk)
    if not asset.is_renewable:
        messages.error(request, 'This asset is not in a renewable category.')
        return redirect('tickets:asset_detail', pk=asset.pk)
    if not asset.renewal_interval_months:
        messages.error(request, 'Set a renewal interval (months) before marking this as renewed.')
        return redirect('tickets:asset_detail', pk=asset.pk)

    new_cost_raw = request.POST.get('new_cost', '').strip()
    new_cost = None
    if new_cost_raw:
        try:
            new_cost = Decimal(new_cost_raw)
        except InvalidOperation:
            messages.error(request, 'Enter a valid cost, or leave it blank to keep the current cost.')
            return redirect('tickets:asset_detail', pk=asset.pk)

    date_type = request.POST.get('date_type', 'LAST')
    if date_type not in ('LAST', 'NEXT'):
        date_type = 'LAST'
    renewal_date_raw = request.POST.get('renewal_date', '').strip()
    renewal_date = None
    if renewal_date_raw:
        renewal_date = parse_date(renewal_date_raw)
        if renewal_date is None:
            messages.error(request, 'Enter a valid date.')
            return redirect('tickets:asset_detail', pk=asset.pk)

    asset.mark_renewed(request.user, new_cost=new_cost, renewal_date=renewal_date, date_type=date_type)
    messages.success(request, f'"{asset.name}" renewed — next renewal {asset.next_renewal_date}.')
    return redirect('tickets:asset_detail', pk=asset.pk)


@login_required
@require_POST
def asset_adjust_stock(request, pk):
    """Admin action: audited correction of a consumable's quantity_in_stock
    (stocktake correction, shrinkage, breakage found in storage) — see
    Asset.adjust_stock()."""
    if effective_role_name(request.user) not in ['ADMIN', 'SUPERADMIN']:
        return HttpResponse(status=403)

    asset = get_object_or_404(Asset, pk=pk)
    if not asset.is_consumable:
        messages.error(request, 'Stock adjustment only applies to consumable assets.')
        return redirect('tickets:asset_detail', pk=asset.pk)

    reason = request.POST.get('reason', '').strip()
    if not reason:
        messages.error(request, 'A reason is required to adjust stock.')
        return redirect('tickets:asset_detail', pk=asset.pk)

    try:
        new_quantity = int(request.POST.get('new_quantity', ''))
    except (TypeError, ValueError):
        messages.error(request, 'Enter a valid quantity.')
        return redirect('tickets:asset_detail', pk=asset.pk)
    if new_quantity < 0:
        messages.error(request, 'Quantity cannot be negative.')
        return redirect('tickets:asset_detail', pk=asset.pk)

    asset.adjust_stock(new_quantity, reason, request.user)
    messages.success(request, f'Stock for "{asset.name}" adjusted to {new_quantity}.')
    return redirect('tickets:asset_detail', pk=asset.pk)


@login_required
@require_POST
def asset_attachment_upload(request, pk):
    """Attach an optional file (license agreement, invoice, contract, etc.)
    to an asset — free-form, not required by any workflow."""
    if effective_role_name(request.user) not in ['ADMIN', 'SUPERADMIN']:
        return HttpResponse(status=403)

    asset = get_object_or_404(Asset, pk=pk)
    uploaded = request.FILES.get('file')
    if not uploaded:
        messages.error(request, 'Choose a file to upload.')
        return redirect('tickets:asset_detail', pk=asset.pk)

    if uploaded.size > MAX_SIZE_MB * 1024 * 1024:
        messages.error(request, f'"{uploaded.name}" is larger than the {MAX_SIZE_MB}MB limit.')
        return redirect('tickets:asset_detail', pk=asset.pk)

    mime = (uploaded.content_type or '').split(';')[0].strip().lower()
    if mime not in ALLOWED_MIMES or not sniffed_mime_matches(uploaded, mime):
        messages.error(request, f'"{uploaded.name}" is not an allowed file type.')
        return redirect('tickets:asset_detail', pk=asset.pk)

    AssetAttachment.objects.create(
        asset=asset,
        file=uploaded,
        filename=uploaded.name,
        uploaded_by=request.user,
        content_type=uploaded.content_type or '',
        size=uploaded.size,
    )
    messages.success(request, f'"{uploaded.name}" attached to "{asset.name}".')
    return redirect('tickets:asset_detail', pk=asset.pk)


@login_required
@require_POST
def asset_attachment_delete(request, pk):
    if effective_role_name(request.user) not in ['ADMIN', 'SUPERADMIN']:
        return HttpResponse(status=403)

    attachment = get_object_or_404(AssetAttachment, pk=pk)
    asset_pk = attachment.asset_id
    attachment.file.delete(save=False)
    attachment.delete()
    messages.success(request, 'Attachment removed.')
    return redirect('tickets:asset_detail', pk=asset_pk)


# ==========================================================================
# ASSET SCRAP REQUEST
# ==========================================================================

# apps/tickets/views.py - Fix asset_scrap_request

@login_required
@require_POST
def asset_scrap_request(request, pk):
    if effective_role_name(request.user) not in ('ADMIN', 'SUPERADMIN'):
        return HttpResponse(status=403)

    asset = get_object_or_404(Asset, pk=pk)

    # Check if asset is already scrapped or damaged
    if asset.status == Asset.Status.SCRAPPED:
        return JsonResponse({'error': 'Asset already scrapped.'}, status=400)

    if asset.status == Asset.Status.DAMAGED:
        return JsonResponse({'error': 'Asset already marked as damaged.'}, status=400)

    # A mobilized asset must be demobilized first — this view clears
    # checked_out_to/assigned_to and flips status directly (mobilization
    # doesn't route through release()), so scrap-requesting it while the
    # MobilizationItem is still open would orphan that row exactly like an
    # un-guarded reassign would (see Asset.mobilization_blocked_reason).
    blocked = asset.mobilization_blocked_reason
    if blocked:
        return JsonResponse({'error': f'{blocked} before requesting scrap.'}, status=400)

    form = AssetScrapRequestForm(request.POST)
    if not form.is_valid():
        if request.headers.get('HX-Request'):
            return render(request, 'partials/scrap_request_modal.html', {'asset': asset, 'form': form})
        messages.error(request, 'Please correct the errors below.')
        return redirect('tickets:assets')
    comment = form.cleaned_data['comment']

    # Record what to restore to if this scrap request is later rejected —
    # previously hardcoded to IN_STORE on reject, silently "healing" an
    # asset that had e.g. been in MAINTENANCE before someone additionally
    # flagged it for scrap. Also release it from whoever currently has it:
    # scrapping (or requesting to) a checked-out asset used to leave the
    # AssetCheckoutHistory row open and checked_out_to/assigned_to pointing
    # at that person forever, misrepresenting a to-be-scrapped item as
    # still being in someone's possession.
    previous_status = asset.status
    previous_holder = asset.checked_out_to or asset.assigned_to
    if previous_holder:
        open_history = asset.checkout_history.filter(checked_in_at__isnull=True).first()
        if open_history:
            open_history.checked_in_by = request.user
            open_history.checked_in_at = timezone.now()
            open_history.return_reason = Asset.ReturnReason.DAMAGED
            open_history.return_comment = f'Released for scrap request: {comment}' if comment else 'Released for scrap request'
            open_history.save()
        asset.checked_out_to = None
        asset.assigned_to = None

    asset.status = Asset.Status.DAMAGED
    asset.status_updated_at = timezone.now()
    asset.status_updated_by = request.user
    asset.save()

    AssetLog.objects.create(
        asset=asset,
        action=AssetLog.Action.SCRAP_REQUESTED,
        actor=request.user,
        details={
            'comment': comment,
            'previous_status': previous_status,
            'released_from': previous_holder.get_full_name() if previous_holder else None,
        }
    )

    if request.headers.get('HX-Request'):
        messages.success(request, f'Scrap requested for "{asset.name}" — pending approval.')
        return HttpResponse(status=204, headers={'HX-Redirect': reverse('tickets:assets')})
    messages.success(request, f'Scrap requested for "{asset.name}" — pending approval.')
    return redirect('tickets:assets')

@login_required
def scrap_request_modal(request, pk):
    """Returns the scrap request modal content."""
    if effective_role_name(request.user) not in ('ADMIN', 'SUPERADMIN'):
        return HttpResponse(status=403)

    asset = get_object_or_404(Asset, pk=pk)
    form = AssetScrapRequestForm()
    return render(request, 'partials/scrap_request_modal.html', {'asset': asset, 'form': form})

# ==========================================================================
# ASSET SCRAP APPROVE
# ==========================================================================

@login_required
@require_POST
def asset_scrap_approve(request, pk):
    if effective_role_name(request.user) not in ['ADMIN', 'SUPERADMIN']:
        return HttpResponse(status=403)

    asset = get_object_or_404(Asset, pk=pk)
    action = request.POST.get('action')  # 'approve' or 'reject'

    requester = asset.pending_scrap_requested_by
    if not requester:
        return HttpResponse('This asset has no pending scrap request.', status=400)
    if requester.id == request.user.id:
        return HttpResponse('A different Admin/Superadmin must approve or reject this request — the requester cannot approve their own.', status=403)

    if action == 'approve':
        asset.status = Asset.Status.SCRAPPED
        asset.scrap_approved = True
        asset.scrap_approved_at = timezone.now()
        asset.scrap_approved_by = request.user
        asset.save()
        AssetLog.objects.create(
            asset=asset,
            action=AssetLog.Action.SCRAP_APPROVED,
            actor=request.user,
            details={'comment': request.POST.get('comment', '')}
        )
        messages.success(request, f'Scrap approved for "{asset.name}" — asset marked as scrapped.')
    else:
        # Reject: restore whatever status the asset was actually in before
        # the scrap request marked it DAMAGED (e.g. MAINTENANCE), rather
        # than always dropping it straight back to IN_STORE regardless of
        # what it was doing before. Custody was already released at
        # request time (see asset_scrap_request) and isn't restored here —
        # the asset just goes back into the pool at its prior status.
        last_request_log = asset.logs.filter(action=AssetLog.Action.SCRAP_REQUESTED).order_by('-created_at').first()
        previous_status = (last_request_log.details or {}).get('previous_status') if last_request_log else None
        asset.status = previous_status if previous_status in Asset.Status.values else Asset.Status.IN_STORE
        asset.status_updated_at = timezone.now()
        asset.status_updated_by = request.user
        asset.save()
        AssetLog.objects.create(
            asset=asset,
            action=AssetLog.Action.SCRAP_REJECTED,
            actor=request.user,
            details={'comment': request.POST.get('comment', '')}
        )
        messages.success(request, f'Scrap rejected for "{asset.name}" — returned to inventory.')

    return redirect('tickets:assets')

@login_required
def scrap_approve_modal(request, pk):
    """Returns the scrap approve modal content."""
    if effective_role_name(request.user) not in ['ADMIN', 'SUPERADMIN']:
        return HttpResponse(status=403)
    
    asset = get_object_or_404(Asset, pk=pk)
    requester = asset.pending_scrap_requested_by
    if not requester:
        return HttpResponse('This asset has no pending scrap request.', status=400)
    if requester.id == request.user.id:
        return HttpResponse('A different Admin/Superadmin must approve or reject this request — the requester cannot approve their own.', status=400)
    return render(request, 'partials/scrap_approve_modal.html', {'asset': asset})


# ==========================================================================
# ASSET CALCULATE WARRANTY
# ==========================================================================

@login_required
def asset_calculate_warranty(request):
    """Calculate warranty expiry date based on purchase date and duration."""
    purchase_date_str = request.GET.get('purchase_date')
    duration_years = request.GET.get('warranty_duration', 0)
    expiry_date_str = ''
    
    import logging
    logger = logging.getLogger(__name__)
    logger.info(f"Warranty calculation: purchase_date={purchase_date_str}, duration={duration_years}")
    
    try:
        duration_years = int(duration_years)
        if duration_years > 0 and purchase_date_str:
            from datetime import datetime
            purchase_date = datetime.strptime(purchase_date_str, '%Y-%m-%d').date()
            # Calculate expiry by adding years
            try:
                expiry_date = purchase_date.replace(year=purchase_date.year + duration_years)
                expiry_date_str = expiry_date.strftime('%Y-%m-%d')
                logger.info(f"Warranty calculation result: {expiry_date_str}")
            except ValueError:
                # Handle Feb 29 edge case - approximate by adding days
                days = 365 * duration_years
                expiry_date = purchase_date + timedelta(days=days)
                expiry_date_str = expiry_date.strftime('%Y-%m-%d')
                logger.info(f"Warranty calculation (approx): {expiry_date_str}")
    except (ValueError, TypeError, OverflowError) as e:
        logger.error(f"Warranty calculation error: {e}")
        pass

    return render(request, 'partials/warranty_expiry_input.html', {'value': expiry_date_str})

# ==========================================================================
# REMOTE SESSION REQUESTS (Quick Assist integration)
# ==========================================================================

@login_required
@require_POST
def request_remote_session(request, pk):
    """
    Initiates a remote session request from an agent to the ticket requester.
    - Creates a RemoteSession record (status=REQUESTED).
    - Adds a public comment on the ticket.
    - Sends an in‑app notification to the requester.
    - Sends an email to the requester with a link to accept the session.
    - Logs the action in the activity log.
    Only agents, team leads, admins, and superadmins can call this.
    """
    ticket = get_object_or_404(Ticket, pk=pk)
    if effective_role_name(request.user) not in [User.Role.AGENT, User.Role.TEAM_LEAD, User.Role.ADMIN, User.Role.SUPERADMIN]:
        return JsonResponse({'error': 'You do not have permission to request a remote session.'}, status=403)

    # Get the first active remote connector (e.g., Quick Assist)
    connector = RemoteConnector.objects.filter(is_active=True).first()
    if not connector:
        # This request is always hx-post/hx-swap="none" (see ticket_conversation.html),
        # so a messages.error()+redirect is never actually rendered to the user —
        # return JSON so the caller's toast can show the real reason.
        return JsonResponse({'error': 'No active remote connector configured. Please contact your administrator.'}, status=400)

    # Check if there's already a pending or active session for this ticket
    existing = RemoteSession.objects.filter(ticket=ticket, status__in=[RemoteSession.Status.REQUESTED, RemoteSession.Status.ACCEPTED, RemoteSession.Status.STARTED]).first()
    if existing:
        return JsonResponse({'error': 'A remote session is already pending or in progress.'}, status=400)
    
    session = RemoteSession.objects.create(
        ticket=ticket,
        requester=ticket.requester,
        agent=request.user,
        connector=connector,
        status=RemoteSession.Status.REQUESTED
    )

    # Add a public comment to the ticket timeline
    TicketComment.objects.create(
        ticket=ticket,
        author=request.user,
        body=f"Remote session requested via {escape(connector.name)}. Please check your notifications to accept.",
        visibility='PUBLIC',
        is_system_generated=True,
        system_icon='monitor',
    )

    # Send in‑app notification
    Notification.objects.create(
        recipient=ticket.requester,
        role=role_of(ticket.requester),
        message=f"Remote session requested for ticket {ticket.number}. Click to accept.",
        url=reverse('tickets:remote_session_detail', args=[session.pk]),
        type=Notification.Type.REMOTE_SESSION
    )

    # send_mail section

    accept_url = request.build_absolute_uri(reverse('tickets:remote_session_detail', args=[session.pk]))
    reject_url = request.build_absolute_uri(reverse('tickets:remote_session_detail', args=[session.pk])) + '?action=reject'

    html_message = render_to_string('emails/remote_session_request.html', {
        'requester_name': ticket.requester.get_full_name() or ticket.requester.email,
        'ticket_number': ticket.number,
        'accept_url': accept_url,
        'reject_url': reject_url,
    })
    plain_message = strip_tags(html_message)

    success, result = send_email_via_brevo(
        to_email=ticket.requester.email,
        subject=f"Remote Session Request – Ticket {ticket.number}",
        html_content=html_message,
        from_email=settings.DEFAULT_FROM_EMAIL
    )

    if not success:
        logger.error(f"Failed to send remote session email: {result}")
    
    TicketActivityLog.objects.create(
        ticket=ticket,
        action='remote_session_requested',
        actor=request.user,
        details={'connector': connector.name, 'session_id': session.pk}
    )
    
    return JsonResponse({'session_id': session.pk, 'status': 'requested'})

@login_required
def remote_session_detail(request, session_pk):
    session = get_object_or_404(RemoteSession, pk=session_pk)
    user = request.user
    is_overseer = effective_role_name(user) in [User.Role.ADMIN, User.Role.SUPERADMIN]
    if user != session.requester and user != session.agent and not is_overseer:
        return HttpResponse(status=403)

    if user == session.agent:
        instructions = session.connector.instructions_for_agent
        role = 'agent'
    elif user == session.requester:
        instructions = session.connector.instructions_for_requester
        role = 'requester'
    else:
        # Admin/Superadmin viewing a session they're not party to — read-only oversight,
        # same as remote_sessions_list already shows them every session.
        instructions = session.connector.instructions_for_requester
        role = 'observer'

    code_error = None

    if request.method == 'POST':
        new_status = request.POST.get('status')
        if new_status in dict(RemoteSession.STATUS_CHOICES):
            old_status = session.status
            # Handle REJECT from requester — only meaningful before the session has
            # actually started, matching every other transition's old_status guard
            # (previously unguarded, so a stray POST could "reject" an ended session).
            if new_status == RemoteSession.Status.REJECTED and role == 'requester' and old_status in (RemoteSession.Status.REQUESTED, RemoteSession.Status.ACCEPTED):
                session.status = RemoteSession.Status.REJECTED
                session.save()
                # Notify agent
                Notification.objects.create(
                    recipient=session.agent,
                    role=role_of(session.agent),
                    message=f"{session.requester.get_full_name()} rejected the remote session for ticket {session.ticket.number}.",
                    url=reverse('tickets:remote_session_detail', args=[session.pk]),
                    type=Notification.Type.REMOTE_SESSION
                )
                TicketComment.objects.create(
                    ticket=session.ticket,
                    author=user,
                    body=f"{escape(session.requester.get_full_name())} declined the remote session request.",
                    visibility='PUBLIC',
                    is_system_generated=True,
                    system_icon='monitor',
                )
                TicketActivityLog.objects.create(
                    ticket=session.ticket,
                    action='remote_session_status_change',
                    actor=user,
                    details={'from': old_status, 'to': RemoteSession.Status.REJECTED, 'session_id': session.pk}
                )
                return redirect('tickets:remote_session_detail', session_pk=session.pk)
            
            # Handle ACCEPT from requester
            elif new_status == RemoteSession.Status.ACCEPTED and role == 'requester' and old_status == RemoteSession.Status.REQUESTED:
                session.status = RemoteSession.Status.ACCEPTED
                session.save()
                # Notify agent
                Notification.objects.create(
                    recipient=session.agent,
                    role=role_of(session.agent),
                    message=f"{session.requester.get_full_name()} accepted the remote session for ticket {session.ticket.number}.",
                    url=reverse('tickets:remote_session_detail', args=[session.pk]),
                    type=Notification.Type.REMOTE_SESSION
                )
                TicketComment.objects.create(
                    ticket=session.ticket,
                    author=user,
                    body=f"{escape(session.requester.get_full_name())} accepted the remote session request.",
                    visibility='PUBLIC',
                    is_system_generated=True,
                    system_icon='monitor',
                )
                TicketActivityLog.objects.create(
                    ticket=session.ticket,
                    action='remote_session_status_change',
                    actor=user,
                    details={'from': old_status, 'to': RemoteSession.Status.ACCEPTED, 'session_id': session.pk}
                )
                return redirect('tickets:remote_session_detail', session_pk=session.pk)
            
            # Handle START with code from agent
            elif new_status == RemoteSession.Status.STARTED and role == 'agent' and old_status == RemoteSession.Status.ACCEPTED:
                code = request.POST.get('quick_assist_code', '').strip()
                if not code or len(code) < 6:
                    code_error = 'Enter the 6-digit code shown in Quick Assist before starting the session.'
                else:
                    session.session_code = code
                    session.status = RemoteSession.Status.STARTED
                    session.started_at = timezone.now()
                    session.save()
                    # Send automatic public comment on ticket
                    TicketComment.objects.create(
                        ticket=session.ticket,
                        author=request.user,
                        body=f"Remote session code: {escape(code)}. Please use this code in Quick Assist to start the session.",
                        visibility='PUBLIC',
                        is_system_generated=True,
                        system_icon='monitor',
                    )
                    # Send email to requester
                    html_message = render_to_string('emails/remote_session_code.html', {
                        'requester_name': session.requester.get_full_name() or session.requester.email,
                        'ticket_number': session.ticket.number,
                        'code': code,
                    })
                    plain_message = strip_tags(html_message)

                    success, result = send_email_via_brevo(
                        to_email=session.requester.email,
                        subject=f"Remote Session Code – Ticket {session.ticket.number}",
                        html_content=html_message,
                        from_email=settings.DEFAULT_FROM_EMAIL
                    )

                    if not success:
                        logger.error(f"Failed to send remote session code email: {result}")
                        
                    TicketActivityLog.objects.create(
                        ticket=session.ticket,
                        action='remote_session_status_change',
                        actor=user,
                        details={'from': old_status, 'to': RemoteSession.Status.STARTED, 'session_id': session.pk, 'code': code}
                    )
                    return redirect('tickets:remote_session_detail', session_pk=session.pk)
            
            # Handle END from agent
            elif new_status == RemoteSession.Status.ENDED and role == 'agent' and old_status == RemoteSession.Status.STARTED:
                session.status = RemoteSession.Status.ENDED
                session.ended_at = timezone.now()
                session.save()
                # Previously the requester was never told the session ended — they'd
                # only find out by refreshing. Notify + email them like every other
                # transition does.
                Notification.objects.create(
                    recipient=session.requester,
                    role=role_of(session.requester),
                    message=f"The remote session for ticket {session.ticket.number} has ended.",
                    url=reverse('tickets:remote_session_detail', args=[session.pk]),
                    type=Notification.Type.REMOTE_SESSION
                )
                TicketComment.objects.create(
                    ticket=session.ticket,
                    author=user,
                    body="The remote session has ended.",
                    visibility='PUBLIC',
                    is_system_generated=True,
                    system_icon='monitor',
                )
                html_message = render_to_string('emails/remote_session_ended.html', {
                    'requester_name': session.requester.get_full_name() or session.requester.email,
                    'ticket_number': session.ticket.number,
                })
                success, result = send_email_via_brevo(
                    to_email=session.requester.email,
                    subject=f"Remote Session Ended – Ticket {session.ticket.number}",
                    html_content=html_message,
                    from_email=settings.DEFAULT_FROM_EMAIL
                )
                if not success:
                    logger.error(f"Failed to send remote session ended email: {result}")
                TicketActivityLog.objects.create(
                    ticket=session.ticket,
                    action='remote_session_status_change',
                    actor=user,
                    details={'from': old_status, 'to': RemoteSession.Status.ENDED, 'session_id': session.pk}
                )
                return redirect('tickets:remote_session_detail', session_pk=session.pk)
    
    context = {
        'session': session,
        'instructions': instructions,
        'role': role,
        'sidebar_template': get_sidebar_template(request.user),
        'code_error': code_error,
        # The "Reject" button in the request email links here with ?action=reject —
        # previously nothing read this param at all, so the link silently did
        # nothing beyond opening the normal page. We still require an explicit
        # click to actually reject (a GET request shouldn't change state), but
        # highlight the reject option so the email's promise is at least honored.
        'highlight_reject': request.GET.get('action') == 'reject',
    }
    return render(request, 'tickets/remote_session_detail.html', context)

@login_required
def remote_session_pending_count(request):
    """
    Returns the count of pending remote sessions for the current user.
    For agents: sessions they requested with status REQUESTED or ACCEPTED.
    For requesters: sessions requested for their tickets with status REQUESTED.
    """
    user = request.user
    if effective_role_name(user) in [User.Role.ADMIN, User.Role.SUPERADMIN]:
        # Admins see all pending sessions
        count = RemoteSession.objects.filter(status__in=[RemoteSession.Status.REQUESTED, RemoteSession.Status.ACCEPTED]).count()
    elif effective_role_name(user) in [User.Role.AGENT, User.Role.TEAM_LEAD]:
        count = RemoteSession.objects.filter(agent=user, status__in=[RemoteSession.Status.REQUESTED, RemoteSession.Status.ACCEPTED]).count()
    else:
        # End users: sessions requested for their tickets
        count = RemoteSession.objects.filter(requester=user, status=RemoteSession.Status.REQUESTED).count()
    return render(request, 'partials/remote_session_badge.html', {'count': count})

@login_required
def remote_sessions_list(request):
    """
    List all remote sessions relevant to the logged‑in user.
    - Agents see sessions they initiated (as agent).
    - Requesters see sessions requested for their tickets.
    - Admins/Superadmins see all sessions (optional, but we'll filter by role).
    """
    user = request.user
    order_args, active_sort, sort_options = resolve_sort(request, {
        '-created_at': (('-created_at',), 'Newest First'),
        '-updated_at': (('-updated_at',), 'Recently Updated'),
    }, '-created_at')
    base_qs = RemoteSession.objects.select_related('ticket', 'requester', 'agent', 'connector')
    if effective_role_name(user) in [User.Role.ADMIN, User.Role.SUPERADMIN]:
        sessions = base_qs.order_by(*order_args)
    elif effective_role_name(user) in [User.Role.AGENT, User.Role.TEAM_LEAD] and user.department == 'IT':
        sessions = base_qs.filter(agent=user).order_by(*order_args)
    else:
        sessions = base_qs.filter(requester=user).order_by(*order_args)

    # Pagination
    paginator = Paginator(sessions, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    context = {
        'sessions': page_obj,
        'sidebar_template': get_sidebar_template(request.user),
        'sort_options': sort_options,
        'active_sort': active_sort,
    }
    if request.headers.get('HX-Request'):
        return render(request, 'partials/remote_sessions_grid.html', context)
    return render(request, 'tickets/remote_sessions_list.html', context)

def _build_escalated_reassign_form(agents, agent_workload, data=None):
    """Builds the reassign form with per-agent workload counts in the option
    labels — label_from_instance is overridden per-instance since the
    workload dict is only known at request time."""
    form = EscalatedReassignForm(data, agents=agents)
    form.fields['agent_id'].label_from_instance = lambda obj: (
        f"{obj.get_full_name()} ({agent_workload.get(obj.pk, 0)} open tickets)"
        if obj.pk in agent_workload else obj.get_full_name()
    )
    return form


@login_required
def escalated_tickets(request):
    if effective_role_name(request.user) not in [User.Role.TEAM_LEAD, User.Role.ADMIN, User.Role.SUPERADMIN] or request.user.department != 'IT':
        return HttpResponse(status=403)

    order_args, active_sort, sort_options = resolve_sort(request, TICKET_SORT_OPTIONS, '-created_at')
    tickets = Ticket.objects.filter(status=Ticket.Status.ESCALATED).order_by(*order_args)

    # If Team Lead, filter by their department
    if effective_role_name(request.user) == User.Role.TEAM_LEAD:
        tickets = tickets.filter(assigned_to__department=request.user.department)

    # Agents in the same department (for reassign)
    agents = User.objects.filter(
        department=request.user.department,
        role=User.Role.AGENT,
        is_active=True
    )

    # Workload: open tickets per agent
    open_statuses = ['NEW', 'TRIAGED', 'ASSIGNED', 'IN_PROGRESS', 'PENDING_USER', 'PENDING_VENDOR']
    workload_counts = dict(
        Ticket.objects.filter(assigned_to__in=agents, status__in=open_statuses)
        .values_list('assigned_to').annotate(total=Count('id')).values_list('assigned_to', 'total')
    )
    agent_workload = {agent.pk: workload_counts.get(agent.pk, 0) for agent in agents}

    # Reason macros
    reassign_reasons = Macro.objects.filter(type=Macro.Type.REASSIGN_REASON)
    return_reasons = Macro.objects.filter(type=Macro.Type.RETURN_REASON)

    context = {
        'tickets': tickets,
        'assignable_agents': agents,
        'agent_workload': agent_workload,
        'reassign_reasons': reassign_reasons,
        'return_reasons': return_reasons,
        'reassign_form': _build_escalated_reassign_form(agents, agent_workload),
        'return_form': EscalatedReturnForm(),
        'sidebar_template': get_sidebar_template(request.user),
        'sort_options': sort_options,
        'active_sort': active_sort,
    }
    return render(request, 'team_lead/escalated_tickets.html', context)


@login_required
def escalated_pending_count(request):
    """Badge count for the sidebar 'Escalated' link — same filter as
    escalated_tickets, minus the fields that view only needs for display."""
    if effective_role_name(request.user) not in [User.Role.TEAM_LEAD, User.Role.ADMIN, User.Role.SUPERADMIN] or request.user.department != 'IT':
        return HttpResponse(status=403)

    tickets = Ticket.objects.filter(status=Ticket.Status.ESCALATED)
    if effective_role_name(request.user) == User.Role.TEAM_LEAD:
        tickets = tickets.filter(assigned_to__department=request.user.department)
    return render(request, 'partials/sidebar_count_badge.html', {'count': tickets.count()})


@login_required
@require_POST
def reassign_escalated(request, pk):
    if effective_role_name(request.user) not in [User.Role.TEAM_LEAD, User.Role.ADMIN, User.Role.SUPERADMIN]:
        return HttpResponse(status=403)
    
    ticket = get_object_or_404(Ticket, pk=pk, status=Ticket.Status.ESCALATED)

    agents = User.objects.filter(department=request.user.department, role=User.Role.AGENT, is_active=True)
    open_statuses = ['NEW', 'TRIAGED', 'ASSIGNED', 'IN_PROGRESS', 'PENDING_USER', 'PENDING_VENDOR']
    agent_workload = {
        agent.pk: Ticket.objects.filter(assigned_to=agent, status__in=open_statuses).count()
        for agent in agents
    }
    form = _build_escalated_reassign_form(agents, agent_workload, data=request.POST)
    if not form.is_valid():
        if request.headers.get('HX-Request'):
            return render(request, 'partials/escalated_reassign_form_fields.html', {
                'form': form,
                'reassign_reasons': Macro.objects.filter(type=Macro.Type.REASSIGN_REASON),
            })
        for error in form.errors.values():
            messages.error(request, error.as_text())
        return redirect('tickets:escalated_tickets')

    agent = form.cleaned_data['agent_id']
    comment = form.cleaned_data['comment']

    # Store old assignee name
    old_name = ticket.assigned_to.get_full_name() if ticket.assigned_to else 'Unassigned'
    new_name = agent.get_full_name()
    actor_name = request.user.get_full_name()
    
    # Reassign the ticket
    ticket.assigned_to = agent
    ticket.status = Ticket.Status.ASSIGNED
    ticket.save()
    
    # Create activity log
    TicketActivityLog.objects.create(
        ticket=ticket,
        action='reassigned_escalated',
        actor=request.user,
        details={'to': agent.get_full_name(), 'comment': comment}
    )
    
    # ================================================================
    # DEFAULT REASSIGN COMMENT
    # ================================================================
    reassign_body = f"**Escalated ticket reassigned** by {escape(actor_name)} from **{escape(old_name)}** to **{escape(new_name)}**."
    if comment:
        reassign_body += f"\n\n**Reason:** {escape(comment)}"
    
    TicketComment.objects.create(
        ticket=ticket,
        author=request.user,
        body=reassign_body,
        visibility='PUBLIC'
    )
    
    # Notify agent
    Notification.objects.create(
        recipient=agent,
        role=role_of(agent),
        message=f"Ticket {ticket.number} has been reassigned to you by {request.user.get_full_name()}.",
        url=reverse('tickets:detail', args=[ticket.pk])
    )
    messages.success(request, f'Ticket {ticket.number} reassigned to {new_name}.')
    if request.headers.get('HX-Request'):
        return HttpResponse(status=204, headers={'HX-Redirect': reverse('tickets:escalated_tickets')})
    return redirect('tickets:escalated_tickets')

@login_required
def escalated_reassign_modal(request, pk):
    """Returns the escalated ticket reassign modal with pre-filled comment."""
    if effective_role_name(request.user) not in [User.Role.TEAM_LEAD, User.Role.ADMIN, User.Role.SUPERADMIN]:
        return HttpResponse(status=403)
    
    ticket = get_object_or_404(Ticket, pk=pk, status=Ticket.Status.ESCALATED)
    
    # Get agents in the same department
    agents = User.objects.filter(
        department=request.user.department, 
        role=User.Role.AGENT, 
        is_active=True
    )
    
    # Generate default comment
    old_name = ticket.assigned_to.get_full_name() if ticket.assigned_to else 'Unassigned'
    default_comment = f"Reassigning escalated ticket from {old_name} to [new agent]."
    
    return render(request, 'partials/escalated_reassign_modal.html', {
        'ticket': ticket,
        'agents': agents,
        'default_comment': default_comment,
    })

@login_required
@require_POST
def return_escalated_to_pool(request, pk):
    if effective_role_name(request.user) not in [User.Role.TEAM_LEAD, User.Role.ADMIN, User.Role.SUPERADMIN]:
        return HttpResponse(status=403)
    ticket = get_object_or_404(Ticket, pk=pk, status=Ticket.Status.ESCALATED)

    form = EscalatedReturnForm(request.POST)
    if not form.is_valid():
        if request.headers.get('HX-Request'):
            return render(request, 'partials/escalated_return_form_fields.html', {
                'form': form,
                'return_reasons': Macro.objects.filter(type=Macro.Type.RETURN_REASON),
            })
        for error in form.errors.values():
            messages.error(request, error.as_text())
        return redirect('tickets:escalated_tickets')

    comment = form.cleaned_data['comment']
    ticket.assigned_to = None
    ticket.status = Ticket.Status.NEW
    ticket.save()
    TicketActivityLog.objects.create(
        ticket=ticket,
        action='returned_to_pool',
        actor=request.user,
        details={'comment': comment}
    )
    messages.success(request, f'Ticket {ticket.number} returned to the pool.')
    if request.headers.get('HX-Request'):
        return HttpResponse(status=204, headers={'HX-Redirect': reverse('tickets:escalated_tickets')})
    return redirect('tickets:escalated_tickets')


# ==========================================================================
# ATTACHMENT PREVIEW AND DOWNLOAD
# ==========================================================================

_OFFICE_CONTENT_TYPES = {
    'application/msword',
    'application/vnd.openxmlformats-officedocument.wordprocessingml.document',
    'application/vnd.ms-excel',
    'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    'application/vnd.ms-powerpoint',
    'application/vnd.openxmlformats-officedocument.presentationml.presentation',
}


def _generate_attachment_preview_pdf(attachment):
    """Converts an Office attachment to PDF via LibreOffice and caches the
    result on attachment.preview_pdf (reuses documents_display's conversion
    pipeline rather than duplicating it). Returns the FieldFile to serve, or
    None if conversion failed."""
    if attachment.preview_pdf:
        return attachment.preview_pdf

    from django.core.files import File as DjangoFile
    from apps.documents_display.utils import convert_office_to_pdf, _local_path_for

    input_path, input_is_temp = _local_path_for(attachment.file)
    try:
        pdf_path = convert_office_to_pdf(input_path)
    finally:
        if input_is_temp:
            try:
                os.remove(input_path)
            except OSError:
                pass

    if not pdf_path or not os.path.exists(pdf_path):
        return None

    base_name = os.path.splitext(attachment.filename)[0]
    with open(pdf_path, 'rb') as f:
        attachment.preview_pdf.save(f'{base_name}_preview.pdf', DjangoFile(f), save=True)

    import shutil
    shutil.rmtree(os.path.dirname(pdf_path), ignore_errors=True)
    return attachment.preview_pdf


@login_required
def attachment_preview(request, pk):
    """
    Returns a modal with a preview of the attachment.
    Supports images, PDF, Office documents (converted to PDF via
    LibreOffice), video, audio, and text files — everything a browser can
    natively render. Anything else falls back to a plain download link,
    since there's no way to meaningfully render an arbitrary binary format
    in a browser without one.
    """
    attachment = get_object_or_404(Attachment, pk=pk)
    ticket = attachment.ticket

    # Permission check: same as download
    if request.user != ticket.requester and effective_role_name(request.user) not in ['AGENT', 'TEAM_LEAD', 'ADMIN', 'SUPERADMIN']:
        return HttpResponse(status=403)

    # Get file URL for embedding
    file_url = request.build_absolute_uri(attachment.file.url)
    content_type = attachment.content_type or ''
    filename = attachment.filename

    # Determine preview type
    preview_type = 'unknown'
    embed_url = None
    text_content = None

    if content_type.startswith('image/'):
        preview_type = 'image'
        embed_url = file_url
    elif content_type == 'application/pdf':
        # Browsers render PDFs natively — no third-party viewer needed.
        preview_type = 'pdf'
        embed_url = file_url
    elif content_type in _OFFICE_CONTENT_TYPES:
        preview_pdf = _generate_attachment_preview_pdf(attachment)
        if preview_pdf:
            preview_type = 'pdf'
            embed_url = request.build_absolute_uri(preview_pdf.url)
        else:
            preview_type = 'conversion_failed'
    elif content_type.startswith('video/'):
        preview_type = 'video'
        embed_url = file_url
    elif content_type.startswith('audio/'):
        preview_type = 'audio'
        embed_url = file_url
    elif content_type.startswith('text/') or filename.endswith(('.txt', '.csv', '.log', '.py', '.js', '.html', '.css')):
        preview_type = 'text'
        # Fetch the file content (for small text files only)
        try:
            with attachment.file.open('r') as f:
                text_content = f.read()
                # Limit size to prevent huge files
                if len(text_content) > 100000:  # 100KB limit
                    text_content = "File too large to preview as text."
        except Exception:
            text_content = "Could not read file content."

    context = {
        'attachment': attachment,
        'preview_type': preview_type,
        'embed_url': embed_url,
        'text_content': text_content,
        'file_url': file_url,
    }
    return render(request, 'tickets/attachment_preview.html', context)

@login_required
def attachment_download(request, pk):
    """
    Serves a file attachment. Only the ticket requester or agents/leads/admins can download.
    """
    attachment = get_object_or_404(Attachment, pk=pk)
    ticket = attachment.ticket
    if request.user != ticket.requester and effective_role_name(request.user) not in ['AGENT', 'TEAM_LEAD', 'ADMIN', 'SUPERADMIN']:
        raise PermissionDenied
    response = FileResponse(attachment.file.open('rb'), content_type=attachment.content_type or 'application/octet-stream')
    response['Content-Disposition'] = f'attachment; filename="{attachment.filename}"'
    return response


@login_required
def requester_profile_modal(request, pk):
    """Returns a modal with the full profile of a ticket requester — the
    "See all" expansion of the Lead Information section on the ticket
    details panel, which previously had no destination at all."""
    profile_user = get_object_or_404(User, pk=pk)

    # Permission check: same shape as attachment access — the requester
    # themselves, or any support-staff role.
    if request.user != profile_user and effective_role_name(request.user) not in ['AGENT', 'TEAM_LEAD', 'ADMIN', 'SUPERADMIN']:
        return HttpResponse(status=403)

    context = {'profile_user': profile_user}
    return render(request, 'partials/requester_profile_modal.html', context)

# ==========================================================================
# SLA MANAGEMENT (Admin & Superadmin only)
# ==========================================================================

# is_admin is imported from apps.common.permissions (see top of file).

@login_required
@user_passes_test(is_admin)
def sla_list(request):
    slas = SLA.objects.all().order_by('priority')
    calendars = BusinessCalendar.objects.annotate(sla_count=Count('sla')).order_by('name')
    rules = EscalationRule.objects.all().order_by('priority', 'timer_type', 'threshold_percent')
    context = {
        'slas': slas,
        'calendars': calendars,
        'rules': rules,
        'priority_choices': Ticket.Priority.choices,
        'sidebar_template': get_sidebar_template(request.user),
        'sla_form': SLAForm(),
        'calendar_form': BusinessCalendarForm(initial={'workdays': ['0', '1', '2', '3', '4']}),
        'rule_form': EscalationRuleForm(),
    }
    return render(request, 'admin/sla_management.html', context)

@login_required
@user_passes_test(is_admin)
@require_POST
def sla_create(request):
    form = SLAForm(request.POST)
    if not form.is_valid():
        if request.headers.get('HX-Request'):
            return render(request, 'partials/sla_form.html', {'form': form})
        messages.error(request, 'Please correct the errors below.')
        return redirect('tickets:sla_management')

    response_total = (form.cleaned_data['response_hours'] or 0) * 60 + (form.cleaned_data['response_minutes'] or 0)
    resolution_total = (form.cleaned_data['resolution_hours'] or 0) * 60 + (form.cleaned_data['resolution_minutes'] or 0)
    calendar = form.cleaned_data['calendar_id']

    try:
        sla, created = SLA.objects.update_or_create(
            priority=form.cleaned_data['priority'],
            defaults={
                'response_minutes': response_total if response_total > 0 else 60,  # Default 1 hour
                'resolution_minutes': resolution_total if resolution_total > 0 else 480,  # Default 8 hours
                'calendar': calendar,
            }
        )
    except Exception as e:
        logger.error(f"Failed to save SLA policy: {e}")
        if request.headers.get('HX-Request'):
            form.add_error(None, 'Could not save the SLA policy. Please try again.')
            return render(request, 'partials/sla_form.html', {'form': form})
        messages.error(request, 'Could not save the SLA policy. Please try again.')
        return redirect('tickets:sla_management')

    if created:
        messages.success(request, f'SLA policy for {sla.get_priority_display()} created successfully.')
        log_admin_action(request.user, AdminActionLog.Category.SLA_CONFIG, 'Created SLA policy', sla.get_priority_display())
    else:
        messages.success(request, f'SLA policy for {sla.get_priority_display()} updated successfully.')
        log_admin_action(
            request.user, AdminActionLog.Category.SLA_CONFIG, 'Updated SLA policy', sla.get_priority_display(),
            details=f'Response: {sla.response_minutes}min, Resolution: {sla.resolution_minutes}min, Calendar: {sla.calendar}',
        )
    if request.headers.get('HX-Request'):
        return HttpResponse(status=204, headers={'HX-Redirect': reverse('tickets:sla_management')})
    return redirect('tickets:sla_management')

@login_required
@user_passes_test(is_admin)
@require_POST
def sla_delete(request, pk):
    sla = get_object_or_404(SLA, pk=pk)
    priority_display = sla.get_priority_display()
    sla.delete()
    log_admin_action(request.user, AdminActionLog.Category.SLA_CONFIG, 'Deleted SLA policy', priority_display)
    messages.success(request, f'SLA policy for {priority_display} deleted.')
    return redirect('tickets:sla_management')

@login_required
def sla_badge(request, pk):
    ticket = get_object_or_404(Ticket, pk=pk)
    return render(request, 'partials/sla_badge.html', {'ticket': ticket})

@login_required
@user_passes_test(is_admin)
@require_POST
def calendar_create(request):
    form = BusinessCalendarForm(request.POST)
    if not form.is_valid():
        if request.headers.get('HX-Request'):
            return render(request, 'partials/calendar_form.html', {'form': form})
        messages.error(request, 'Please correct the errors below.')
        return redirect('tickets:sla_management')

    name = form.cleaned_data['name']
    try:
        BusinessCalendar.objects.create(
            name=name,
            workdays=form.cleaned_data['workdays'],
            work_start=form.cleaned_data['work_start'],
            work_end=form.cleaned_data['work_end'],
            holidays=form.cleaned_data['holidays'],
        )
    except Exception as e:
        logger.error(f"Failed to create business calendar: {e}")
        if request.headers.get('HX-Request'):
            form.add_error(None, 'Could not create the business calendar. Please check the values and try again.')
            return render(request, 'partials/calendar_form.html', {'form': form})
        messages.error(request, 'Could not create the business calendar. Please check the values and try again.')
        return redirect('tickets:sla_management')

    log_admin_action(request.user, AdminActionLog.Category.SLA_CONFIG, 'Created business calendar', name)
    messages.success(request, f'Business calendar "{name}" created successfully.')
    if request.headers.get('HX-Request'):
        return HttpResponse(status=204, headers={'HX-Redirect': reverse('tickets:sla_management')})
    return redirect('tickets:sla_management')

@login_required
@user_passes_test(is_admin)
@require_POST
def rule_create(request):
    form = EscalationRuleForm(request.POST)
    if not form.is_valid():
        if request.headers.get('HX-Request'):
            return render(request, 'partials/rule_form.html', {'form': form})
        messages.error(request, 'Please correct the errors below.')
        return redirect('tickets:sla_management')

    try:
        rule = EscalationRule.objects.create(
            priority=form.cleaned_data['priority'],
            timer_type=form.cleaned_data['timer_type'],
            threshold_percent=form.cleaned_data['threshold_percent'],
            action_type=form.cleaned_data['action_type'],
            notify_role=form.cleaned_data['notify_role'] or None,
            reassign_to_role=form.cleaned_data['reassign_to_role'] or None,
        )
    except Exception as e:
        logger.error(f"Failed to create escalation rule: {e}")
        if request.headers.get('HX-Request'):
            form.add_error(None, 'Could not create the escalation rule. Please check the values and try again.')
            return render(request, 'partials/rule_form.html', {'form': form})
        messages.error(request, 'Could not create the escalation rule. Please check the values and try again.')
        return redirect('tickets:sla_management')

    log_admin_action(
        request.user, AdminActionLog.Category.SLA_CONFIG, 'Created escalation rule',
        f'{rule.get_priority_display()} / {rule.get_timer_type_display()}',
    )
    messages.success(request, 'Escalation rule created successfully.')
    if request.headers.get('HX-Request'):
        return HttpResponse(status=204, headers={'HX-Redirect': reverse('tickets:sla_management')})
    return redirect('tickets:sla_management')

@login_required
@user_passes_test(is_admin)
def calendar_add_modal(request):
    """Blank calendar form (partials/calendar_form.html, no edit_pk) —
    fetched via HTMX by the Add button so it always replaces whatever was
    last swapped into the modal (which might be a stale Edit form) with a
    guaranteed-fresh create form."""
    return render(request, 'partials/calendar_form.html', {'form': BusinessCalendarForm(initial={'workdays': ['0', '1', '2', '3', '4']})})


@login_required
@user_passes_test(is_admin)
def calendar_edit_modal(request, pk):
    """Returns the calendar form pre-filled for editing an existing
    calendar (partials/calendar_form.html, shared with Add)."""
    cal = get_object_or_404(BusinessCalendar, pk=pk)
    form = BusinessCalendarForm(initial={
        'name': cal.name,
        'workdays': cal.workdays,
        'work_start': cal.work_start,
        'work_end': cal.work_end,
        'holidays': ', '.join(cal.holidays),
    })
    return render(request, 'partials/calendar_form.html', {'form': form, 'edit_pk': cal.pk})


@login_required
@user_passes_test(is_admin)
@require_POST
def calendar_edit(request, pk):
    cal = get_object_or_404(BusinessCalendar, pk=pk)
    form = BusinessCalendarForm(request.POST)
    if not form.is_valid():
        if request.headers.get('HX-Request'):
            return render(request, 'partials/calendar_form.html', {'form': form, 'edit_pk': cal.pk})
        messages.error(request, 'Please correct the errors below.')
        return redirect('tickets:sla_management')

    cal.name = form.cleaned_data['name']
    cal.workdays = form.cleaned_data['workdays']
    cal.work_start = form.cleaned_data['work_start']
    cal.work_end = form.cleaned_data['work_end']
    cal.holidays = form.cleaned_data['holidays']
    cal.save()

    log_admin_action(request.user, AdminActionLog.Category.SLA_CONFIG, 'Updated business calendar', cal.name)
    messages.success(request, f'Business calendar "{cal.name}" updated successfully.')
    if request.headers.get('HX-Request'):
        return HttpResponse(status=204, headers={'HX-Redirect': reverse('tickets:sla_management')})
    return redirect('tickets:sla_management')


@login_required
@user_passes_test(is_admin)
def rule_add_modal(request):
    """Blank escalation rule form — see calendar_add_modal for why this
    exists alongside rule_edit_modal."""
    return render(request, 'partials/rule_form.html', {'form': EscalationRuleForm()})


@login_required
@user_passes_test(is_admin)
def rule_edit_modal(request, pk):
    """Returns the escalation rule form pre-filled for editing an existing
    rule (partials/rule_form.html, shared with Add)."""
    rule = get_object_or_404(EscalationRule, pk=pk)
    form = EscalationRuleForm(initial={
        'priority': rule.priority,
        'timer_type': rule.timer_type,
        'threshold_percent': rule.threshold_percent,
        'action_type': rule.action_type,
        'notify_role': rule.notify_role or '',
        'reassign_to_role': rule.reassign_to_role or '',
    })
    return render(request, 'partials/rule_form.html', {'form': form, 'edit_pk': rule.pk})


@login_required
@user_passes_test(is_admin)
@require_POST
def rule_edit(request, pk):
    rule = get_object_or_404(EscalationRule, pk=pk)
    form = EscalationRuleForm(request.POST)
    if not form.is_valid():
        if request.headers.get('HX-Request'):
            return render(request, 'partials/rule_form.html', {'form': form, 'edit_pk': rule.pk})
        messages.error(request, 'Please correct the errors below.')
        return redirect('tickets:sla_management')

    rule.priority = form.cleaned_data['priority']
    rule.timer_type = form.cleaned_data['timer_type']
    rule.threshold_percent = form.cleaned_data['threshold_percent']
    rule.action_type = form.cleaned_data['action_type']
    rule.notify_role = form.cleaned_data['notify_role'] or None
    rule.reassign_to_role = form.cleaned_data['reassign_to_role'] or None
    rule.save()

    log_admin_action(
        request.user, AdminActionLog.Category.SLA_CONFIG, 'Updated escalation rule',
        f'{rule.get_priority_display()} / {rule.get_timer_type_display()}',
    )
    messages.success(request, 'Escalation rule updated successfully.')
    if request.headers.get('HX-Request'):
        return HttpResponse(status=204, headers={'HX-Redirect': reverse('tickets:sla_management')})
    return redirect('tickets:sla_management')


@login_required
@user_passes_test(is_admin)
@require_POST
def rule_delete(request, pk):
    rule = get_object_or_404(EscalationRule, pk=pk)
    rule_label = f'{rule.get_priority_display()} / {rule.get_timer_type_display()}'
    rule.delete()
    log_admin_action(request.user, AdminActionLog.Category.SLA_CONFIG, 'Deleted escalation rule', rule_label)
    messages.success(request, 'Escalation rule deleted.')
    return redirect('tickets:sla_management')

@login_required
@user_passes_test(is_admin)
@require_POST
def calendar_delete(request, pk):
    cal = get_object_or_404(BusinessCalendar, pk=pk)
    cal_name = cal.name
    cal.delete()
    log_admin_action(request.user, AdminActionLog.Category.SLA_CONFIG, 'Deleted business calendar', cal_name)
    messages.success(request, f'Business calendar "{cal_name}" deleted.')
    return redirect('tickets:sla_management')

@login_required
@user_passes_test(is_admin)
def resolved_service_requests(request):
    """Admin-only transparency view: every service request that has reached
    a resolved outcome, org-wide. Distinct from Reports/Exportables, which
    are analytics/export tooling rather than a plain audit list."""
    order_args, active_sort, sort_options = resolve_sort(request, TICKET_SORT_OPTIONS, '-updated_at')
    tickets = Ticket.objects.filter(
        type=Ticket.Type.SERVICE_REQUEST,
        status__in=[Ticket.Status.RESOLVED, Ticket.Status.CLOSED],
    ).select_related('requester', 'assigned_to', 'category').order_by(*order_args)

    department = request.GET.get('department', '').strip()
    if department:
        tickets = tickets.filter(requester__department=department)

    q = request.GET.get('q', '').strip()
    if q:
        tickets = tickets.filter(
            Q(number__icontains=q) | Q(title__icontains=q) |
            Q(requester__first_name__icontains=q) | Q(requester__last_name__icontains=q) |
            Q(requester__email__icontains=q)
        )

    paginator = Paginator(tickets, 20)
    page_obj = paginator.get_page(request.GET.get('page', 1))

    context = {
        'tickets': page_obj,
        'department_choices': User.DEPARTMENT_CHOICES,
        'department_filter': department,
        'q': q,
        'sidebar_template': get_sidebar_template(request.user),
        'sort_options': sort_options,
        'active_sort': active_sort,
    }
    if request.headers.get('HX-Request'):
        return render(request, 'partials/resolved_service_requests_table.html', context)
    return render(request, 'admin/resolved_service_requests.html', context)

# ==========================================================================
# MANAGER WORKFLOW (Team Lead reviews service requests)
# ==========================================================================

@login_required
def manager_review_queue(request):
    """Team Lead view – list service requests pending manager review."""
    if effective_role_name(request.user) != User.Role.TEAM_LEAD:
        return HttpResponse(status=403)

    order_args, active_sort, sort_options = resolve_sort(request, TICKET_SORT_OPTIONS, '-created_at')
    tickets = Ticket.objects.filter(
        status=Ticket.Status.PENDING_MANAGER_REVIEW,
        requester__department=request.user.department
    ).order_by(*order_args)

    context = {
        'tickets': tickets,
        'sidebar_template': get_sidebar_template(request.user),
        'sort_options': sort_options,
        'active_sort': active_sort,
    }
    return render(request, 'team_lead/manager_review_queue.html', context)


@login_required
def manager_review_ticket(request, pk):
    """Team Lead review page for a single service request."""
    if effective_role_name(request.user) != User.Role.TEAM_LEAD:
        return HttpResponse(status=403)

    ticket = get_object_or_404(Ticket, pk=pk)

    # Security: ensure ticket belongs to Team Lead's department
    if ticket.requester.department != request.user.department:
        return HttpResponse(status=403)

    # Only allow review of PENDING_MANAGER_REVIEW tickets
    if ticket.status != Ticket.Status.PENDING_MANAGER_REVIEW:
        messages.warning(request, f'Ticket {ticket.number} is not pending manager review.')
        return redirect('tickets:manager_review_queue')

    if request.method == 'POST':
        action = request.POST.get('action', '').strip()
        comment = request.POST.get('comment', '').strip()

        # Comment is required for REJECT and REQUEST_CHANGES only
        if action in ['reject', 'request_changes'] and not comment:
            messages.error(request, f'Please provide a comment explaining why you are {action.replace("_", " ")} this request.')
            return redirect('tickets:manager_review_ticket', pk=pk)

        if action == 'approve':
            # ================================================================
            # APPROVAL - Comment is Optional
            # ================================================================
            if ticket.is_asset_request:
                ticket.status = Ticket.Status.PENDING_FULFILLMENT
                ticket.save()
                
                TicketActivityLog.objects.create(
                    ticket=ticket,
                    action='manager_approved',
                    actor=request.user,
                    details={'comment': comment if comment else 'No comment provided', 'routed_to': 'PENDING_FULFILLMENT'}
                )
                
                admins = User.objects.filter(role=User.Role.ADMIN, is_active=True)
                for admin in admins:
                    Notification.objects.create(
                        recipient=admin,
                        role=role_of(admin),
                        message=f'Asset request {ticket.number} from {ticket.requester.get_full_name()} needs fulfillment.',
                        url=reverse('tickets:conversation', args=[ticket.pk])
                    )
                
                Notification.objects.create(
                    recipient=ticket.requester,
                    role=role_of(ticket.requester),
                    message=f'Your asset request {ticket.number} has been approved by your manager and is pending fulfillment.',
                    url=reverse('tickets:detail', args=[ticket.pk])
                )
                
                messages.success(request, f'Asset request {ticket.number} approved. An admin will fulfill it shortly.')
            else:
                ticket.status = Ticket.Status.APPROVED
                ticket.save()
                
                TicketActivityLog.objects.create(
                    ticket=ticket,
                    action='manager_approved',
                    actor=request.user,
                    details={'comment': comment if comment else 'No comment provided', 'routed_to': 'APPROVED'}
                )
                
                Notification.objects.create(
                    recipient=ticket.requester,
                    role=role_of(ticket.requester),
                    message=f'Your service request {ticket.number} has been approved.',
                    url=reverse('tickets:detail', args=[ticket.pk])
                )
                
                agents = User.objects.filter(role__in=[User.Role.AGENT, User.Role.TEAM_LEAD])
                for agent in agents:
                    Notification.objects.create(
                        recipient=agent,
                        role=role_of(agent),
                        message=f'New approved ticket {ticket.number}: {ticket.title}',
                        url=reverse('tickets:detail', args=[ticket.pk])
                    )
                
                messages.success(request, f'Ticket {ticket.number} approved and sent to agent queue.')

        elif action == 'reject':
            ticket.status = Ticket.Status.CLOSED
            ticket.save()
            TicketActivityLog.objects.create(
                ticket=ticket,
                action='manager_rejected',
                actor=request.user,
                details={'comment': comment}
            )
            Notification.objects.create(
                recipient=ticket.requester,
                role=role_of(ticket.requester),
                message=f'Your service request {ticket.number} was rejected by your manager. Reason: {comment}',
                url=reverse('tickets:detail', args=[ticket.pk])
            )
            messages.info(request, f'Ticket {ticket.number} rejected.')

        elif action == 'request_changes':
            ticket.status = Ticket.Status.PENDING_USER
            ticket.save()
            TicketActivityLog.objects.create(
                ticket=ticket,
                action='manager_requested_changes',
                actor=request.user,
                details={'comment': comment}
            )
            # Post the reason as a real comment — previously it only lived in
            # the activity log and the notification text, so the requester's
            # ticket page showed nothing explaining what to change.
            comment_body = clean_comment_body(f'<p><strong>Changes requested:</strong> {comment}</p>')
            if comment_body:
                TicketComment.objects.create(
                    ticket=ticket, author=request.user, visibility='PUBLIC', body=comment_body
                )
            Notification.objects.create(
                recipient=ticket.requester,
                role=role_of(ticket.requester),
                message=f'Changes requested for ticket {ticket.number} by your manager: {comment}',
                url=reverse('tickets:detail', args=[ticket.pk])
            )
            messages.info(request, f'Changes requested on ticket {ticket.number}.')

        else:
            messages.error(request, f'Invalid action: "{action}"')
            return redirect('tickets:manager_review_ticket', pk=pk)

        return redirect('tickets:manager_review_queue')

    # GET – render review page
    comments = ticket.comments.all().order_by('created_at')
    initial_attachments = ticket.attachments.filter(comment__isnull=True)
    attachments = ticket.attachments.all().order_by('uploaded_at')

    context = {
        'ticket': ticket,
        'comments': comments,
        'initial_attachments': initial_attachments,
        'attachments': attachments,
        'sidebar_template': get_sidebar_template(request.user),
    }
    return render(request, 'team_lead/manager_review_ticket.html', context)

@login_required
def manager_review_count(request):
    if effective_role_name(request.user) != User.Role.TEAM_LEAD:
        return HttpResponse('')
    count = Ticket.objects.filter(
        status=Ticket.Status.PENDING_MANAGER_REVIEW,
        requester__department=request.user.department
    ).count()
    return render(request, 'partials/manager_review_badge.html', {'count': count})


@login_required
def manager_review_history(request):
    """Team Lead view - every service request this Team Lead has personally
    approved, most recent first, with their approval comment and where it
    was routed (straight to the agent queue, or to Admin for fulfillment)."""
    if effective_role_name(request.user) != User.Role.TEAM_LEAD:
        return HttpResponse(status=403)

    logs = TicketActivityLog.objects.filter(
        action='manager_approved', actor=request.user
    ).select_related('ticket', 'ticket__requester').order_by('-created_at')

    paginator = Paginator(logs, 15)
    page_number = request.GET.get('page', 1)
    try:
        page_obj = paginator.page(page_number)
    except (PageNotAnInteger, EmptyPage):
        page_obj = paginator.page(1)

    context = {
        'page_obj': page_obj,
        'sidebar_template': get_sidebar_template(request.user),
    }
    return render(request, 'team_lead/manager_review_history.html', context)


# ==========================================================================
# ASSET IMPORT
# ==========================================================================

def _find_assigned_to_by_name_or_email(name_or_email):
    """Exact email or exact full-name match (case-insensitive) only —
    ambiguous/no matches return (None, warning) rather than guessing via a
    fuzzy substring match. Name matching also tolerates first/last being
    swapped (see asset_name_matching.match_users_by_name) since imported
    sheets and the system's own first_name/last_name split don't always
    agree on which is which. Shared by the legacy row-mapping helper below
    and asset_import_commit."""
    name_or_email = str(name_or_email).strip()
    if not name_or_email:
        return None, None
    if '@' in name_or_email:
        matches = list(User.objects.filter(email__iexact=name_or_email))
    else:
        matches = list(match_users_by_name(name_or_email))
    if len(matches) == 1:
        return matches[0], None
    if len(matches) > 1:
        return None, f"'{name_or_email}' matched multiple users, left unassigned."
    return None, f"no user found matching '{name_or_email}', left unassigned."


def _resolve_department_match(name):
    """Existing AssetDepartment matching `name` by exact name or tag_code
    (case-insensitive) — or None if none exists and one would need to be
    created. Deliberately exact-only, same as _find_assigned_to_by_name_or_
    email: a fuzzy match here is exactly how 'Account' silently became a
    duplicate of 'Accounting' instead of asking. Shared by
    asset_import_preview (to flag the miss before committing) and
    asset_import_commit (to actually resolve it)."""
    name = (name or '').strip()
    if not name:
        return None
    return (
        AssetDepartment.objects.filter(name__iexact=name).first()
        or AssetDepartment.objects.filter(tag_code__iexact=name).first()
    )


def _resolve_location_match(name):
    """Existing top-level Location matching `name` (case-insensitive) — or
    None if one would need to be created. See _resolve_department_match."""
    name = (name or '').strip()
    if not name:
        return None
    return Location.objects.filter(name__iexact=name, parent__isnull=True).first()


def _resolve_category_match(name):
    """Existing top-level AssetCategory matching `name` by exact name or
    tag_code (case-insensitive) — or None if one would need to be created.
    See _resolve_department_match. AssetCategory already had pre-existing
    near-duplicates before this import ever ran (e.g. 'Laptop'/'Laptops',
    'Monitor'/'Monitor x2') — this only prevents new ones from imports
    going forward, it doesn't retroactively clean up those."""
    name = (name or '').strip()
    if not name:
        return None
    return (
        AssetCategory.objects.filter(name__iexact=name, parent__isnull=True).first()
        or AssetCategory.objects.filter(tag_code__iexact=name, parent__isnull=True).first()
    )


def _read_raw_rows(file):
    """Returns a list of row tuples (values only) for either a CSV or
    Excel upload, in original sheet order — the shape asset_import_transform
    .transform_raw_rows() expects. CSV rows are read plainly (no DictReader)
    so a CSV export of the same raw, section-header-having layout is
    normalized identically to an Excel one."""
    file_name = file.name.lower()
    if file_name.endswith('.csv'):
        decoded = file.read().decode('utf-8')
        # io.StringIO (not decoded.splitlines()) so csv.reader sees the
        # real line breaks itself — splitlines() would pre-split a quoted
        # multi-line field (e.g. a Comments cell) before csv.reader gets a
        # chance to respect the quoting, silently merging it into one line.
        return [tuple(row) for row in csv.reader(io.StringIO(decoded))]
    import openpyxl
    wb = openpyxl.load_workbook(file)
    ws = wb.active
    return list(ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True))


@login_required
@require_POST
def asset_import(request):
    """Step 1: upload a CSV/Excel file, run it through the transform step,
    and stage the result as an AssetImportBatch for review — no Asset rows
    are created here. URL name kept as 'asset_import' (unchanged from the
    old single-step flow) so the existing Import button/form doesn't need
    updating."""
    if effective_role_name(request.user) not in ['ADMIN', 'SUPERADMIN']:
        return HttpResponse(status=403)

    file = request.FILES.get('file')
    if not file:
        messages.error(request, 'Please select a file to import.')
        return redirect('tickets:assets')

    MAX_IMPORT_SIZE = 5 * 1024 * 1024  # 5MB
    if file.size > MAX_IMPORT_SIZE:
        messages.error(request, 'File too large. Maximum size is 5MB.')
        return redirect('tickets:assets')

    file_name = file.name.lower()
    if not file_name.endswith(('.csv', '.xlsx', '.xls')):
        messages.error(request, 'Please upload a CSV or Excel file.')
        return redirect('tickets:assets')

    try:
        raw_rows = _read_raw_rows(file)
        file.seek(0)
        normalized_rows = transform_raw_rows(raw_rows)
    except Exception as e:
        messages.error(request, f'Error reading file: {str(e)}')
        return redirect('tickets:assets')

    if not normalized_rows:
        messages.warning(request, 'No asset rows were recognized in that file. Please check the column headers.')
        return redirect('tickets:assets')

    batch = AssetImportBatch.objects.create(
        uploaded_file=file,
        uploaded_by=request.user,
        normalized_data=normalized_rows,
        row_count=len(normalized_rows),
    )
    return redirect('tickets:asset_import_preview', pk=batch.pk)


@login_required
def asset_import_preview(request, pk):
    """Step 2: review the transformed data before anything is created.
    Annotates each row with what asset_import_commit would actually do
    (resolved status, any assigned-to lookup problem) so issues are visible
    up front rather than discovered mid-commit."""
    if effective_role_name(request.user) not in ['ADMIN', 'SUPERADMIN']:
        return HttpResponse(status=403)

    batch = get_object_or_404(AssetImportBatch, pk=pk)
    if batch.status != AssetImportBatch.Status.PENDING_REVIEW:
        messages.info(request, f'This import batch has already been {batch.get_status_display().lower()}.')
        return redirect('tickets:assets')

    preview_rows = []
    unmatched_dept_names = set()
    unmatched_loc_names = set()
    unmatched_category_names = set()
    for row in batch.normalized_data:
        assigned_to_name = row.get('assigned_to_name', '')
        assigned_to, assign_warning = (None, None)
        if assigned_to_name:
            assigned_to, assign_warning = _find_assigned_to_by_name_or_email(assigned_to_name)

        department_name = (row.get('department_name') or '').strip()
        if department_name and _resolve_department_match(department_name) is None:
            unmatched_dept_names.add(department_name)

        location_name = (row.get('location_name') or '').strip()
        if location_name and _resolve_location_match(location_name) is None:
            unmatched_loc_names.add(location_name)

        category_name = (row.get('category_name') or '').strip()
        if category_name and _resolve_category_match(category_name) is None:
            unmatched_category_names.add(category_name)

        preview_rows.append({
            **row,
            'resolved_status': resolve_status_hint(row.get('status_hint', '')),
            'assigned_to_resolved': assigned_to.get_full_name() if assigned_to else None,
            'assign_warning': assign_warning,
        })

    return render(request, 'tickets/asset_import_preview.html', {
        'batch': batch,
        'preview_rows': preview_rows,
        # Sheet department/location/category text that doesn't exactly
        # match an existing row — surfaced here so an admin can map it to
        # the right existing one (or explicitly confirm it's new) instead
        # of the commit step silently creating a duplicate (see the
        # 'Account' vs 'Accounting' incident this was built to prevent).
        'unmatched_departments': sorted(unmatched_dept_names),
        'unmatched_locations': sorted(unmatched_loc_names),
        'unmatched_categories': sorted(unmatched_category_names),
        'existing_departments': AssetDepartment.objects.all().order_by('name'),
        'existing_locations': Location.objects.filter(parent__isnull=True).order_by('name'),
        'existing_categories': AssetCategory.objects.filter(parent__isnull=True).order_by('name'),
        'sidebar_template': get_sidebar_template(request.user),
    })


@login_required
@require_POST
def asset_import_commit(request, pk):
    """Step 3: actually create the Asset rows from the reviewed batch."""
    if effective_role_name(request.user) not in ['ADMIN', 'SUPERADMIN']:
        return HttpResponse(status=403)

    batch = get_object_or_404(AssetImportBatch, pk=pk)
    if batch.status != AssetImportBatch.Status.PENDING_REVIEW:
        messages.error(request, f'This import batch has already been {batch.get_status_display().lower()}.')
        return redirect('tickets:assets')

    imported = 0
    warnings = []
    errors = []

    # Admin's choices from the preview page for any department/location/
    # category text that didn't exactly match an existing row —
    # 'dept_map:<sheet text>' / 'loc_map:<sheet text>' / 'cat_map:<sheet
    # text>' -> existing row's pk, or blank/absent to confirm creating a
    # new one under that name. Keyed by the literal sheet text (not an
    # index) so it's unambiguous regardless of row order.
    dept_overrides = {}
    loc_overrides = {}
    cat_overrides = {}
    for key, value in request.POST.items():
        if not value:
            continue
        if key.startswith('dept_map:'):
            dept_overrides[key[len('dept_map:'):]] = value
        elif key.startswith('loc_map:'):
            loc_overrides[key[len('loc_map:'):]] = value
        elif key.startswith('cat_map:'):
            cat_overrides[key[len('cat_map:'):]] = value

    for row_idx, row in enumerate(batch.normalized_data, start=1):
        try:
            name = row.get('name', '')
            if not name:
                continue

            category = None
            category_name = (row.get('category_name') or '').strip()
            if category_name:
                override_pk = cat_overrides.get(category_name)
                category = AssetCategory.objects.filter(pk=override_pk).first() if override_pk else None
                if not category:
                    category = _resolve_category_match(category_name)
                if not category:
                    category = AssetCategory.objects.create(name=category_name)

            location = None
            location_name = (row.get('location_name') or '').strip()
            if location_name:
                override_pk = loc_overrides.get(location_name)
                location = Location.objects.filter(pk=override_pk).first() if override_pk else None
                if not location:
                    location = _resolve_location_match(location_name)
                if not location:
                    location = Location.objects.create(name=location_name, parent=None)

            department = None
            department_name = (row.get('department_name') or '').strip()
            if department_name:
                override_pk = dept_overrides.get(department_name)
                department = AssetDepartment.objects.filter(pk=override_pk).first() if override_pk else None
                if not department:
                    department = _resolve_department_match(department_name)
                if not department:
                    department = AssetDepartment.objects.create(name=department_name)

            assigned_to = None
            unresolved_assignee_hint = ''
            assigned_to_name = row.get('assigned_to_name', '')
            if assigned_to_name:
                assigned_to, assign_warning = _find_assigned_to_by_name_or_email(assigned_to_name)
                if assign_warning:
                    warnings.append(f"Row {row_idx}: {assign_warning}")
                    # Keep the raw name rather than discarding it — lets an
                    # admin later match/create the account and assign this
                    # asset properly instead of it becoming untraceable.
                    unresolved_assignee_hint = assigned_to_name.strip()[:150]

            tag_slot_number = parse_track_no_slot(row.get('track_no', ''))

            asset = Asset(
                name=name,
                category=category,
                location=location,
                department=department,
                status=resolve_status_hint(row.get('status_hint', '')),
                notes=row.get('notes', ''),
                assigned_to=assigned_to,
                unresolved_assignee_hint=unresolved_assignee_hint,
            )
            # The org's own physical tag, already affixed to real hardware
            # — preserved as-is rather than regenerated, same principle as
            # tracking_id never being touched for existing assets.
            given_tag = row.get('tracking_id', '')
            if given_tag:
                asset.tracking_id = given_tag
                if tag_slot_number is not None:
                    asset.tag_slot_number = tag_slot_number
            asset.save()

            AssetLog.objects.create(
                asset=asset,
                action=AssetLog.Action.CREATED,
                actor=request.user,
                details={'name': asset.name, 'category': asset.category.name if asset.category else None, 'source': 'import'}
            )
            if asset.assigned_to:
                AssetLog.objects.create(
                    asset=asset,
                    action=AssetLog.Action.ASSIGNED,
                    actor=request.user,
                    details={'to': asset.assigned_to.get_full_name(), 'source': 'import'}
                )

            imported += 1
        except Exception as e:
            errors.append(f"Row {row_idx}: {str(e)}")

    batch.status = AssetImportBatch.Status.COMMITTED
    batch.committed_at = timezone.now()
    batch.save(update_fields=['status', 'committed_at'])

    if imported > 0:
        messages.success(request, f'Successfully imported {imported} asset(s).')
    if warnings:
        for w in warnings[:10]:
            messages.warning(request, w)
    if errors:
        messages.warning(request, f'{len(errors)} error(s) occurred.')
    if not imported and not errors:
        messages.warning(request, 'No assets were imported.')

    return redirect('tickets:assets')


@login_required
@require_POST
def asset_import_discard(request, pk):
    """Cancel a pending review without creating anything. The uploaded
    file itself is kept (not deleted) per the same policy as a committed
    batch — only the batch's status changes."""
    if effective_role_name(request.user) not in ['ADMIN', 'SUPERADMIN']:
        return HttpResponse(status=403)

    batch = get_object_or_404(AssetImportBatch, pk=pk)
    if batch.status == AssetImportBatch.Status.PENDING_REVIEW:
        batch.status = AssetImportBatch.Status.DISCARDED
        batch.save(update_fields=['status'])
        messages.info(request, 'Import discarded — no assets were created.')
    return redirect('tickets:assets')


# ==========================================================================
# ASSET FULFILLMENT (Admin only)
# ==========================================================================

@login_required
def fulfill_asset_modal(request, pk):
    """Returns the fulfillment modal for an asset request."""
    if effective_role_name(request.user) not in ['ADMIN', 'SUPERADMIN']:
        return HttpResponse(status=403)
    
    ticket = get_object_or_404(
        Ticket, pk=pk,
        status__in=[Ticket.Status.PENDING_FULFILLMENT, Ticket.Status.PENDING_VENDOR]
    )

    # Get available assets (IN_STORE or unassigned READY)
    available_assets = Asset.objects.filter(
        status__in=[Asset.Status.IN_STORE, Asset.Status.READY],
        assigned_to__isnull=True
    ).select_related('category').order_by('name')

    # Show what's already mobilized (out) for this ticket's job/vessel(s)/
    # dive system(s), so the admin can check existing issue before approving more.
    already_mobilized = MobilizationItem.objects.none()
    vessel_ids = list(ticket.vessels.values_list('id', flat=True))
    system_ids = list(ticket.dive_systems.values_list('id', flat=True))
    if ticket.job_number_id or vessel_ids or system_ids:
        job_filter = Q()
        if ticket.job_number_id:
            job_filter |= Q(mobilization__job_number_id=ticket.job_number_id)
        if vessel_ids:
            job_filter |= Q(mobilization__vessels__id__in=vessel_ids)
        if system_ids:
            job_filter |= Q(mobilization__dive_systems__id__in=system_ids)
        already_mobilized = MobilizationItem.objects.filter(
            job_filter, demobilized_at__isnull=True
        ).select_related('asset', 'mobilization').distinct()

    open_procurement = ticket.procurement_requests.filter(
        status__in=[AssetProcurementRequest.Status.REQUESTED, AssetProcurementRequest.Status.ORDERED]
    ).select_related('vendor', 'category').first()

    return render(request, 'admin/fulfill_asset_modal.html', {
        'ticket': ticket,
        'available_assets': available_assets,
        'already_mobilized': already_mobilized,
        'open_procurement': open_procurement,
        'procurement_form': ProcurementRequestForm(),
    })


def _mark_asset_ticket_fulfilled(ticket, request, summary):
    """Shared terminal step for both single-asset (_fulfill_ticket_with_asset)
    and mobilization (mobilization_create/_maybe_fulfill_mobilization_ticket)
    fulfillment: instead of resting at APPROVED forever, ask the requester to
    confirm the asset(s) arrived — reuses the PENDING_USER -> confirm_resolution
    machinery Incident tickets already use via resolve_ticket (this module).
    Confirming receipt (confirm_resolution) does NOT resolve the ticket by
    itself for an asset request — it records the confirmation and returns the
    ticket to APPROVED; an agent still explicitly resolves it afterward via
    resolve_ticket, which skips its own confirmation round-trip once receipt
    is already on record (see the is_asset_request branches in both views)."""
    ticket.status = Ticket.Status.PENDING_USER
    ticket.fulfilled_at = timezone.now()
    ticket.fulfilled_by = request.user
    ticket.save(update_fields=['status', 'fulfilled_at', 'fulfilled_by'])

    TicketComment.objects.create(
        ticket=ticket,
        author=request.user,
        body=f"**Fulfilled**: {escape(summary)}. Please confirm once received.",
        visibility='PUBLIC',
        is_receipt_confirmation_prompt=True,
        is_system_generated=True,
    )

    Notification.objects.create(
        recipient=ticket.requester,
        role=role_of(ticket.requester),
        message=f'Your request {ticket.number} has been fulfilled — please confirm receipt on the ticket page.',
        url=reverse('tickets:detail', args=[ticket.pk]),
        type=Notification.Type.RESOLUTION_CONFIRMATION,
    )

    if ticket.requester.email:
        confirm_url = request.build_absolute_uri(reverse('tickets:detail', args=[ticket.pk]))
        html_message = render_to_string('emails/resolution_confirmation.html', {
            'requester_name': ticket.requester.get_full_name() or ticket.requester.email,
            'ticket_number': ticket.number,
            'ticket_title': ticket.title,
            'confirm_url': confirm_url,
            'agent_name': request.user.get_full_name() or request.user.email,
            'is_asset_request': ticket.is_asset_request,
        })
        success, result = send_email_via_brevo(
            to_email=ticket.requester.email,
            subject=f"Please confirm receipt for {ticket.number}",
            html_content=html_message,
            from_email=settings.DEFAULT_FROM_EMAIL
        )
        if not success:
            logger.error(f"Failed to send fulfillment confirmation email: {result}")


def _mobilization_still_awaiting_vendor(mobilization, exclude_pk=None):
    """True if the mobilization has any procurement request still on order
    from a vendor (not yet received or cancelled)."""
    qs = mobilization.procurement_requests.filter(
        status__in=[AssetProcurementRequest.Status.REQUESTED, AssetProcurementRequest.Status.ORDERED]
    )
    if exclude_pk:
        qs = qs.exclude(pk=exclude_pk)
    return qs.exists()


def _maybe_fulfill_mobilization_ticket(mobilization, request):
    """Fulfills the mobilization's linked ticket (if any) once every item on
    it is actually in hand — immediately if nothing on this mobilization was
    ever sourced from a vendor, or once the last open vendor procurement
    request against it clears (received or cancelled). Called from
    mobilization_create, procurement_receive, and procurement_cancel — all
    three converge here so the requester is only ever asked to confirm
    receipt once everything has actually arrived, never for a part still on
    order. Idempotent: a no-op once the ticket is already past fulfillment."""
    ticket = mobilization.ticket
    if not ticket or ticket.status not in (Ticket.Status.PENDING_FULFILLMENT, Ticket.Status.PENDING_VENDOR):
        return

    if _mobilization_still_awaiting_vendor(mobilization):
        if ticket.status != Ticket.Status.PENDING_VENDOR:
            ticket.status = Ticket.Status.PENDING_VENDOR
            ticket.save(update_fields=['status'])
        return

    if not mobilization.items.exists():
        # Everything that was on order got cancelled and nothing was ever
        # picked from stock — nothing was actually delivered, so there's
        # nothing to confirm receipt of. Leave the ticket as-is.
        return

    _mark_asset_ticket_fulfilled(ticket, request, summary=f'assets mobilized to {mobilization.destination_display}')
    TicketActivityLog.objects.create(
        ticket=ticket, action='mobilization_fulfilled', actor=request.user,
        details={'mobilization_id': mobilization.pk, 'asset_count': mobilization.items.count()}
    )


def _maybe_resolve_mobilization_receipt(ticket, actor):
    """Called after every MobilizationItem accept/dispute. Once every item
    across every mobilization linked to this ticket has been accepted or
    disputed, drives the ticket the same way confirm_resolution's
    all-or-nothing click used to for asset-request tickets: all accepted ->
    APPROVED (mirrors action='confirm'); any disputed -> PENDING_FULFILLMENT
    (mirrors action='reopen'). No-op while any item is still pending, or if
    the ticket isn't a mobilization ticket currently awaiting this at all."""
    if ticket.status != Ticket.Status.PENDING_USER or not ticket.is_mobilization_request:
        return

    items = MobilizationItem.objects.filter(mobilization__ticket=ticket).select_related('asset')
    if not items.exists() or items.filter(acknowledged_at__isnull=True, disputed_at__isnull=True).exists():
        return  # nothing to aggregate yet, or still waiting on someone

    disputed_items = list(items.filter(disputed_at__isnull=False))
    if disputed_items:
        ticket.status = Ticket.Status.PENDING_FULFILLMENT
        ticket.save(update_fields=['status'])

        names = ', '.join(i.asset.tracking_id for i in disputed_items)
        TicketActivityLog.objects.create(
            ticket=ticket, action='resolution_rejected', actor=actor,
            details={'disputed_items': names, 'source': 'mobilization_item_dispute'}
        )
        TicketComment.objects.create(
            ticket=ticket, author=actor, visibility='PUBLIC',
            body=f"**Not received**: the requester disputes receiving {escape(names)}. Sent back for review.",
            is_system_generated=True,
        )
        if ticket.fulfilled_by:
            Notification.objects.create(
                recipient=ticket.fulfilled_by, role=role_of(ticket.fulfilled_by),
                message=f"{ticket.requester.get_full_name()} disputes receiving mobilized item(s) for {ticket.number}.",
                url=reverse('tickets:conversation', args=[ticket.pk])
            )
        return

    # Every item accepted, none disputed.
    ticket.resolution_confirmed_at = timezone.now()
    ticket.resolution_confirmed_by = ticket.requester
    ticket.status = Ticket.Status.APPROVED
    ticket.save()

    TicketActivityLog.objects.create(
        ticket=ticket, action='receipt_confirmed', actor=ticket.requester,
        details={'confirmed_at': ticket.resolution_confirmed_at.isoformat(), 'source': 'mobilization_items'}
    )


def _fulfill_ticket_with_asset(ticket, asset, actor, request, comment=''):
    """Core of fulfilling an asset-request ticket by assigning `asset` to
    the requester — shared by the direct-fulfillment path (an asset was
    already in stock) and the post-receiving path (a procurement request
    for this ticket just arrived). Routes through Asset.assign_to() (the
    same method the standalone checkout flow uses) rather than setting
    assigned_to directly — this used to be a second, independent
    'who has this asset' mechanism with no checked_out_to/AssetCheckoutHistory
    of its own, which meant a fulfilled asset could then also be checked
    out to a *different* user via the ordinary checkout flow, silently
    overwriting assigned_to while the ticket still pointed at it."""
    asset.assign_to(ticket.requester, actor=actor, notes=f'Fulfilled request {ticket.number}: {comment}')

    ticket.assigned_asset = asset
    ticket.save(update_fields=['assigned_asset'])
    summary = f'{asset.name} ({asset.tracking_id}) assigned to you'
    if comment:
        summary += f'. {comment}'
    _mark_asset_ticket_fulfilled(ticket, request, summary=summary)

    TicketActivityLog.objects.create(
        ticket=ticket,
        action='asset_fulfilled',
        actor=actor,
        details={
            'asset_id': asset.pk,
            'asset_name': asset.name,
            'asset_tracking_id': asset.tracking_id,
            'assigned_to': ticket.requester.get_full_name()
        }
    )

    Notification.objects.create(
        recipient=ticket.requester,
        role=role_of(ticket.requester),
        message=f'Your asset request {ticket.number} has been fulfilled. {asset.name} assigned to you.',
        url=reverse('tickets:detail', args=[ticket.pk])
    )

    if ticket.requester.email:
        from apps.accounts.models import ClientSettings
        client = ClientSettings.objects.first()
        html_message = render_to_string('emails/asset_fulfilled.html', {
            'requester_name': ticket.requester.get_full_name() or ticket.requester.email,
            'ticket': ticket,
            'asset': asset,
            'ticket_url': request.build_absolute_uri(reverse('tickets:detail', args=[ticket.pk])),
            'logo_url': request.build_absolute_uri(client.logo.url) if client and client.logo else None,
            'client_settings': {
                'company_name': client.company_name if client else 'My Company',
            },
        })
        success, result = send_email_via_brevo(
            to_email=ticket.requester.email,
            subject=f"Your asset request {ticket.number} has been fulfilled",
            html_content=html_message,
            from_email=settings.DEFAULT_FROM_EMAIL
        )
        if not success:
            logger.error(f"Failed to send asset fulfilled email: {result}")


@login_required
@require_POST
def fulfill_asset_request(request, pk):
    """Admin action to fulfill an asset request by assigning an asset."""
    if effective_role_name(request.user) not in ['ADMIN', 'SUPERADMIN']:
        return HttpResponse(status=403)
    
    ticket = get_object_or_404(
        Ticket, pk=pk,
        status__in=[Ticket.Status.PENDING_FULFILLMENT, Ticket.Status.PENDING_VENDOR]
    )
    asset_id = request.POST.get('asset_id')
    comment = request.POST.get('comment', '').strip()
    
    if not asset_id:
        messages.error(request, 'Please select an asset to assign.')
        return redirect('tickets:conversation', pk=ticket.pk)
    
    with transaction.atomic():
        try:
            asset = Asset.objects.select_for_update().get(pk=asset_id)
        except Asset.DoesNotExist:
            messages.error(request, 'Asset not found.')
            return redirect('tickets:conversation', pk=ticket.pk)

        # Check if asset is already assigned
        if asset.assigned_to:
            messages.error(request, f'Asset {asset.name} is already assigned to {asset.assigned_to.get_full_name()}.')
            return redirect('tickets:conversation', pk=ticket.pk)

        # Check if asset is actually available (not scrapped/retired/lost/etc.)
        if not asset.is_available:
            messages.error(request, f'Asset {asset.name} is not available (status: {asset.status_display["label"]}).')
            return redirect('tickets:conversation', pk=ticket.pk)

        try:
            _fulfill_ticket_with_asset(ticket, asset, request.user, request, comment=comment)
        except ValueError as e:
            messages.error(request, str(e))
            return redirect('tickets:conversation', pk=ticket.pk)

    messages.success(request, f'Asset {asset.name} assigned to {ticket.requester.get_full_name()}.')
    return redirect('tickets:conversation', pk=ticket.pk)


@login_required
def pending_asset_fulfillment_list(request):
    """Full, paginated list of tickets awaiting asset fulfillment —
    Admin/Superadmin only. Same query the dashboard's 'Pending Asset
    Fulfillment' widget uses (apps/accounts/views/__init__.py), just
    without the [:10] cap, so 'View all' has somewhere real to go."""
    if effective_role_name(request.user) not in ['ADMIN', 'SUPERADMIN']:
        return HttpResponse(status=403)

    tickets_qs = Ticket.objects.filter(
        status=Ticket.Status.PENDING_FULFILLMENT
    ).select_related('requester', 'category').order_by('-created_at')

    paginator = Paginator(tickets_qs, 10)
    page_obj = paginator.get_page(request.GET.get('page'))

    return render(request, 'tickets/pending_asset_fulfillment_list.html', {
        'tickets': page_obj,
        'sidebar_template': get_sidebar_template(request.user),
    })


@login_required
def pending_asset_fulfillment_count(request):
    """Badge count for the sidebar 'Fulfillment' link — same filter as
    pending_asset_fulfillment_list."""
    if effective_role_name(request.user) not in ['ADMIN', 'SUPERADMIN']:
        return HttpResponse(status=403)

    count = Ticket.objects.filter(status=Ticket.Status.PENDING_FULFILLMENT).count()
    return render(request, 'partials/sidebar_count_badge.html', {'count': count})


@login_required
def pending_settings_approvals_count(request):
    """Badge count for the sidebar 'System Settings' link — total rows
    across every has_proposals resource (Vessels, Job Numbers, Vendors)
    still is_active=False after being proposed, i.e. exactly what each
    resource's pending-approval banner on that page shows."""
    if not is_admin(request.user):
        return HttpResponse(status=403)

    from .settings_registry import SETTINGS_RESOURCES
    count = 0
    for config in SETTINGS_RESOURCES.values():
        if config.has_proposals:
            count += config.model.objects.filter(is_active=False, proposed_by__isnull=False).count()
    return render(request, 'partials/sidebar_count_badge.html', {'count': count})


# ==========================================================================
# VENDOR PROCUREMENT (assets not yet in inventory)
# ==========================================================================

def _notify_new_vendor_proposed(vendor, actor, detail_url):
    """Same propose-and-approve notification shape as mobilization_create's
    third-party-vessel handling — an unrecognized vendor name becomes an
    inactive Vendor row, and Admins are notified to review/activate it.
    Runs at every point a Vendor gets proposed this way, so proposed_by is
    set here once rather than at each of those call sites."""
    if not vendor.proposed_by_id:
        vendor.proposed_by = actor
        vendor.save(update_fields=['proposed_by'])
    for admin in User.objects.filter(role=User.Role.ADMIN, is_active=True):
        Notification.objects.create(
            recipient=admin,
            role=role_of(admin),
            message=(
                f'{actor.get_full_name()} referenced "{vendor.name}", a vendor not yet in the system, '
                f'on a procurement request. Review and activate it under System Settings → Vendors if it should be added.'
            ),
            url=detail_url,
        )


@login_required
@require_POST
def procurement_request_create(request, pk):
    """Record that an asset-request ticket's needed item isn't in stock and
    is being sourced from a vendor. Ticket moves to PENDING_VENDOR —
    receiving the item later is what actually fulfills it."""
    if effective_role_name(request.user) not in ['ADMIN', 'SUPERADMIN']:
        return HttpResponse(status=403)

    ticket = get_object_or_404(Ticket, pk=pk, status=Ticket.Status.PENDING_FULFILLMENT)
    form = ProcurementRequestForm(request.POST)
    if not form.is_valid():
        for field_errors in form.errors.values():
            for error in field_errors:
                messages.error(request, error)
        return redirect('tickets:conversation', pk=ticket.pk)

    procurement_request = form.save(commit=False)
    procurement_request.vendor = form.cleaned_data.get('vendor')
    procurement_request.ticket = ticket
    procurement_request.requested_by = request.user
    procurement_request.save()

    ticket.status = Ticket.Status.PENDING_VENDOR
    ticket.save(update_fields=['status'])

    TicketComment.objects.create(
        ticket=ticket,
        author=request.user,
        body=f"**On order**: {escape(procurement_request.item_name)} x{procurement_request.quantity} requested from "
             f"{escape(procurement_request.vendor.name) if procurement_request.vendor else 'vendor (TBD)'}"
             f"{f', expected {procurement_request.expected_arrival_date}' if procurement_request.expected_arrival_date else ''}.",
        visibility='PUBLIC',
        is_system_generated=True,
    )
    TicketActivityLog.objects.create(
        ticket=ticket,
        action='procurement_requested',
        actor=request.user,
        details={'procurement_request_id': procurement_request.pk, 'item_name': procurement_request.item_name}
    )
    Notification.objects.create(
        recipient=ticket.requester,
        role=role_of(ticket.requester),
        message=f'Your asset request {ticket.number} is on order from a vendor. We\'ll notify you once it arrives.',
        url=reverse('tickets:detail', args=[ticket.pk])
    )
    if form.cleaned_data.get('_new_vendor_proposed'):
        _notify_new_vendor_proposed(procurement_request.vendor, request.user, reverse('tickets:conversation', args=[ticket.pk]))

    messages.success(request, f'{procurement_request.item_name} recorded as on order.')
    return redirect('tickets:conversation', pk=ticket.pk)


@login_required
@require_POST
def procurement_reorder_create(request, asset_pk):
    """One-click 'Reorder' from a low-stock alert / asset detail page —
    a standalone AssetProcurementRequest not tied to any ticket or
    mobilization, restocking the SKU it was raised for once received
    (same Procurement list/receive flow as every other request)."""
    if effective_role_name(request.user) not in ['ADMIN', 'SUPERADMIN']:
        return HttpResponse(status=403)

    asset = get_object_or_404(Asset, pk=asset_pk)
    if not asset.is_consumable:
        messages.error(request, 'Reordering only applies to consumable assets.')
        return redirect('tickets:asset_detail', pk=asset.pk)

    form = ProcurementRequestForm(request.POST)
    if not form.is_valid():
        for field_errors in form.errors.values():
            for error in field_errors:
                messages.error(request, error)
        return redirect('tickets:asset_detail', pk=asset.pk)

    procurement_request = form.save(commit=False)
    procurement_request.vendor = form.cleaned_data.get('vendor')
    procurement_request.requested_by = request.user
    procurement_request.save()

    if form.cleaned_data.get('_new_vendor_proposed'):
        _notify_new_vendor_proposed(procurement_request.vendor, request.user, reverse('tickets:asset_detail', args=[asset.pk]))

    messages.success(request, f'{procurement_request.item_name} recorded as on order.')
    return redirect('tickets:procurement_list')


@login_required
def procurement_list(request):
    """All vendor procurement requests, filterable by status."""
    if effective_role_name(request.user) not in ['ADMIN', 'SUPERADMIN']:
        return HttpResponse(status=403)

    status_filter = request.GET.get('status', '').strip()
    requests_qs = AssetProcurementRequest.objects.select_related(
        'category', 'vendor', 'ticket', 'mobilization', 'requested_by'
    ).order_by('-requested_at')
    if status_filter:
        requests_qs = requests_qs.filter(status=status_filter)

    paginator = Paginator(requests_qs, 25)
    page_obj = paginator.get_page(request.GET.get('page', 1))

    return render(request, 'tickets/procurement_list.html', {
        'procurement_requests': page_obj,
        'status_choices': AssetProcurementRequest.Status.choices,
        'current_status': status_filter,
        'sidebar_template': get_sidebar_template(request.user),
    })


@login_required
@require_POST
def procurement_mark_ordered(request, pk):
    if effective_role_name(request.user) not in ['ADMIN', 'SUPERADMIN']:
        return HttpResponse(status=403)

    procurement_request = get_object_or_404(AssetProcurementRequest, pk=pk, status=AssetProcurementRequest.Status.REQUESTED)
    procurement_request.status = AssetProcurementRequest.Status.ORDERED
    procurement_request.save(update_fields=['status'])
    messages.success(request, f'{procurement_request.item_name} marked as ordered.')
    return redirect(request.META.get('HTTP_REFERER') or 'tickets:procurement_list')


@login_required
@require_POST
def procurement_cancel(request, pk):
    if effective_role_name(request.user) not in ['ADMIN', 'SUPERADMIN']:
        return HttpResponse(status=403)

    procurement_request = get_object_or_404(AssetProcurementRequest, pk=pk)
    if not procurement_request.is_open:
        messages.error(request, 'Only open procurement requests can be cancelled.')
        return redirect(request.META.get('HTTP_REFERER') or 'tickets:procurement_list')

    procurement_request.status = AssetProcurementRequest.Status.CANCELLED
    procurement_request.save(update_fields=['status'])

    if procurement_request.mobilization_id:
        mobilization = procurement_request.mobilization
        if mobilization.ticket_id:
            # Sum quantities, not row counts — a consumable batch (e.g. 3
            # laptops from one vendor order) is a single MobilizationItem
            # row with quantity=3, so .count() would wrongly say 1.
            fulfilled_count = sum(mobilization.items.values_list('quantity', flat=True))
            still_open = sum(mobilization.procurement_requests.filter(
                status__in=[AssetProcurementRequest.Status.REQUESTED, AssetProcurementRequest.Status.ORDERED]
            ).exclude(pk=procurement_request.pk).values_list('quantity', flat=True))
            total = fulfilled_count + still_open
            status_line = f"All {total} items now fulfilled." if total and fulfilled_count == total else f"{fulfilled_count} of {total} items now fulfilled." if total else "Nothing left on order for this mobilization."
            TicketComment.objects.create(
                ticket=mobilization.ticket,
                author=request.user,
                visibility='PUBLIC',
                mobilization=mobilization,
                mobilization_event=TicketComment.MobilizationEvent.VENDOR_ITEM_CANCELLED,
                body=f"<strong>{escape(procurement_request.item_name)} order cancelled</strong><br>{status_line}",
                is_system_generated=True,
            )
        # Cancelling can be what clears the last thing a mobilization's
        # ticket was waiting on (e.g. 2 items ordered, 1 already received,
        # this one cancelled instead of arriving) — recheck the same way
        # receiving does.
        _maybe_fulfill_mobilization_ticket(mobilization, request)

    Notification.objects.create(
        recipient=procurement_request.requested_by,
        role=role_of(procurement_request.requested_by),
        message=f'Procurement request for {procurement_request.item_name} was cancelled.',
        url=reverse('tickets:procurement_list')
    )
    messages.success(request, f'{procurement_request.item_name} procurement request cancelled.')
    return redirect(request.META.get('HTTP_REFERER') or 'tickets:procurement_list')


@login_required
def procurement_receive_modal(request, pk):
    if effective_role_name(request.user) not in ['ADMIN', 'SUPERADMIN']:
        return HttpResponse(status=403)

    procurement_request = get_object_or_404(AssetProcurementRequest, pk=pk)
    if not procurement_request.is_open:
        return HttpResponse('<div class="p-4 text-center text-warning">This request is no longer open.</div>', status=400)

    return render(request, 'partials/procurement_receive_modal.html', {
        'procurement_request': procurement_request,
    })


@login_required
@require_POST
def procurement_receive(request, pk):
    """The item physically arrived — create the real Asset(s)/stock and,
    if this request was for a ticket or a mobilization, automatically
    finish that ticket/mobilization exactly as if the item had come from
    existing stock."""
    if effective_role_name(request.user) not in ['ADMIN', 'SUPERADMIN']:
        return HttpResponse(status=403)

    procurement_request = get_object_or_404(AssetProcurementRequest, pk=pk)
    if not procurement_request.is_open:
        messages.error(request, 'This request is no longer open.')
        return redirect(request.META.get('HTTP_REFERER') or 'tickets:procurement_list')

    with transaction.atomic():
        category = procurement_request.category
        received_assets = []

        if category.is_consumable:
            asset, _ = Asset.objects.select_for_update().get_or_create(
                name=procurement_request.item_name,
                category=category,
                defaults={'created_by': request.user, 'procurement_request': procurement_request}
            )
            asset.quantity_in_stock += procurement_request.quantity
            asset.save(update_fields=['quantity_in_stock'])
            asset.refresh_low_stock_alert()
            received_assets.append(asset)
        else:
            for _ in range(procurement_request.quantity):
                asset = Asset.objects.create(
                    name=procurement_request.item_name,
                    category=category,
                    status=Asset.Status.IN_STORE,
                    created_by=request.user,
                    procurement_request=procurement_request,
                )
                received_assets.append(asset)

        if procurement_request.ticket_id and procurement_request.ticket.status in (
            Ticket.Status.PENDING_FULFILLMENT, Ticket.Status.PENDING_VENDOR
        ):
            ticket = procurement_request.ticket
            fulfillment_asset = received_assets[0]
            try:
                _fulfill_ticket_with_asset(
                    ticket, fulfillment_asset, request.user, request,
                    comment=f'Received from vendor procurement request #{procurement_request.pk}.'
                )
            except ValueError as e:
                # e.g. a consumable-category item was tied to a ticket —
                # the stock still gets received above, it just can't be
                # auto-assigned to one requester the way an individually-
                # tracked asset can. Leave the ticket for manual fulfillment.
                messages.warning(request, f'{procurement_request.item_name} received, but could not be auto-assigned to the ticket: {e}')
        elif procurement_request.mobilization_id:
            mobilization = procurement_request.mobilization
            for asset in received_assets:
                qty = procurement_request.quantity if category.is_consumable else 1
                MobilizationItem.objects.create(mobilization=mobilization, asset=asset, quantity=qty)
                if category.is_consumable:
                    asset.quantity_in_stock -= qty
                    asset.save(update_fields=['quantity_in_stock'])
                    asset.refresh_low_stock_alert()
                else:
                    asset.status = Asset.Status.MOBILIZED
                    asset.status_updated_at = timezone.now()
                    asset.status_updated_by = request.user
                    asset.save(update_fields=['status', 'status_updated_at', 'status_updated_by'])
                AssetLog.objects.create(
                    asset=asset,
                    action=AssetLog.Action.MOBILIZED,
                    actor=request.user,
                    details={
                        'mobilization_id': mobilization.pk,
                        'destination': mobilization.destination_display,
                        'quantity': qty,
                        'procurement_request_id': procurement_request.pk,
                    }
                )

        procurement_request.status = AssetProcurementRequest.Status.RECEIVED
        procurement_request.received_at = timezone.now()
        procurement_request.received_by = request.user
        procurement_request.save(update_fields=['status', 'received_at', 'received_by'])

        if procurement_request.mobilization_id:
            mobilization = procurement_request.mobilization
            if mobilization.ticket_id:
                fulfilled_count = sum(mobilization.items.values_list('quantity', flat=True))
                still_open = sum(mobilization.procurement_requests.filter(
                    status__in=[AssetProcurementRequest.Status.REQUESTED, AssetProcurementRequest.Status.ORDERED]
                ).exclude(pk=procurement_request.pk).values_list('quantity', flat=True))
                total = fulfilled_count + still_open
                status_line = f"All {total} items now fulfilled." if fulfilled_count == total else f"{fulfilled_count} of {total} items now fulfilled."
                TicketComment.objects.create(
                    ticket=mobilization.ticket,
                    author=request.user,
                    visibility='PUBLIC',
                    mobilization=mobilization,
                    mobilization_event=TicketComment.MobilizationEvent.VENDOR_ITEM_ARRIVED,
                    body=f"<strong>{escape(procurement_request.item_name)} arrived from vendor</strong><br>{status_line}",
                    is_system_generated=True,
                )
            # Now that this one's marked RECEIVED, check whether it was the
            # last thing this mobilization's ticket (if any) was waiting on.
            _maybe_fulfill_mobilization_ticket(mobilization, request)

    Notification.objects.create(
        recipient=procurement_request.requested_by,
        role=role_of(procurement_request.requested_by),
        message=f'{procurement_request.item_name} has been received and added to inventory.',
        url=reverse('tickets:procurement_list')
    )
    messages.success(request, f'{procurement_request.item_name} received.')
    return redirect(request.META.get('HTTP_REFERER') or 'tickets:procurement_list')


@login_required
def procurement_export_pdf(request, pk):
    if effective_role_name(request.user) not in ['ADMIN', 'SUPERADMIN']:
        return HttpResponse(status=403)

    procurement_request = get_object_or_404(
        AssetProcurementRequest.objects.select_related('category', 'vendor', 'ticket', 'mobilization', 'requested_by'),
        pk=pk
    )
    from .report_exporters import export_procurement_request_pdf
    return export_procurement_request_pdf(procurement_request, request)


@login_required
def available_assets_for_fulfillment(request):
    """HTMX endpoint to get available assets for a specific request."""
    if effective_role_name(request.user) not in ['ADMIN', 'SUPERADMIN']:
        return HttpResponse(status=403)
    
    search = request.GET.get('search', '').strip()
    category = request.GET.get('category', '').strip()
    
    # Filter available assets (unassigned and active/in-store)
    assets = Asset.objects.filter(
        assigned_to__isnull=True,
        status__in=[Asset.Status.IN_STORE, Asset.Status.READY]
    ).select_related('category').order_by('name')
    
    # Filter by search term
    if search:
        assets = assets.filter(
            Q(name__icontains=search) |
            Q(tracking_id__icontains=search) |
            Q(serial_number__icontains=search) |
            Q(model__icontains=search) |
            Q(manufacturer__icontains=search)
        )

    # Optional: Filter by asset category based on the ticket's request category
    type_mapping = {
        'Hardware': ['Computer', 'Laptop', 'Printer', 'Server', 'Network Device'],
        'Software': ['Software License'],
        'Network': ['Network Device', 'Server'],
        'Printer': ['Printer'],
        'Computer': ['Computer', 'Laptop'],
    }

    if category in type_mapping:
        assets = assets.filter(category__name__in=type_mapping[category])

    # Limit results
    assets = assets[:20]
    
    return render(request, 'partials/available_assets_list.html', {
        'assets': assets,
    })


# apps/tickets/views.py - Add these new functions

# ==========================================================================
# ASSET CHECK-IN/CHECK-OUT
# ==========================================================================

@login_required
def asset_checkout_modal(request, pk):
    """Return the checkout modal for an asset."""
    if effective_role_name(request.user) not in ['ADMIN', 'SUPERADMIN']:
        return HttpResponse(status=403)
    
    asset = get_object_or_404(Asset, pk=pk)

    # Full availability check, not just "not already checked out" — matches
    # what assign_to() will actually enforce on submit.
    if asset.assignment_blocked_reason:
        return HttpResponse(f'"{asset.name}" cannot be checked out — {asset.assignment_blocked_reason}.', status=400)

    # Get users for dropdown (only active users)
    users = User.objects.filter(is_active=True).order_by('first_name', 'last_name')
    form = AssetCheckoutForm()

    return render(request, 'partials/asset_checkout_modal.html', {
        'asset': asset,
        'users': users,
        'form': form,
    })


@login_required
@require_POST
def asset_checkout(request, pk):
    """Check out an asset to a user."""
    if effective_role_name(request.user) not in ['ADMIN', 'SUPERADMIN']:
        return HttpResponse(status=403)

    asset_for_form = get_object_or_404(Asset, pk=pk)
    form = AssetCheckoutForm(request.POST)
    if not form.is_valid():
        users = User.objects.filter(is_active=True).order_by('first_name', 'last_name')
        if request.headers.get('HX-Request'):
            return render(request, 'partials/asset_checkout_modal.html', {'asset': asset_for_form, 'users': users, 'form': form})
        messages.error(request, 'Please correct the errors below.')
        return redirect('tickets:assets')

    user = form.cleaned_data['user_id']
    expected_return = form.cleaned_data['expected_return_date']
    notes = form.cleaned_data['notes'].strip()

    with transaction.atomic():
        asset = get_object_or_404(Asset.objects.select_for_update(), pk=pk)

        # assign_to() is the single source of truth for "is this asset
        # available" — it also refuses damaged/retired/scrapped/already-
        # assigned assets, not just ones already checked out.
        try:
            asset.assign_to(user, actor=request.user, expected_return_date=expected_return, notes=notes)
        except ValueError as e:
            if request.headers.get('HX-Request'):
                return HttpResponse(str(e), status=400)
            messages.error(request, str(e))
            return redirect('tickets:assets')

    # Add comment to asset notes
    checkout_note = f"**Asset checked out** by {request.user.get_full_name()} to {user.get_full_name()} on {timezone.now().strftime('%Y-%m-%d %H:%M')}"
    if notes:
        checkout_note += f"\nNotes: {notes}"
    if asset.notes:
        asset.notes = f"{asset.notes}\n\n{checkout_note}"
    else:
        asset.notes = checkout_note
    asset.save(update_fields=['notes'])
    
    # Notify user
    Notification.objects.create(
        recipient=user,
        role=role_of(user),
        message=f"Asset {asset.name} ({asset.tracking_id}) has been checked out to you.",
        url=reverse('tickets:asset_detail', args=[asset.pk])
    )
    
    messages.success(request, f'Asset "{asset.name}" checked out to {user.get_full_name()}.')
    if request.headers.get('HX-Request'):
        return HttpResponse(status=204, headers={'HX-Redirect': reverse('tickets:assets')})
    return redirect('tickets:assets')


@login_required
def asset_checkin_modal(request, pk):
    """Return the checkin modal for an asset."""
    if effective_role_name(request.user) not in ['ADMIN', 'SUPERADMIN']:
        return HttpResponse(status=403)
    
    asset = get_object_or_404(Asset, pk=pk)

    # Only allow checkin if asset is checked out
    if not asset.is_checked_out:
        return HttpResponse('This asset is not currently checked out.', status=400)

    # Pre-fill from the holder's self-reported return request, if any —
    # the admin's own return_condition assessment (not pre-filled, since
    # only the admin can actually inspect the physical item) still decides
    # the final status in release().
    open_history = asset._open_checkout_history()
    initial = {}
    if open_history and open_history.return_requested_at:
        initial = {
            'return_reason': open_history.return_requested_reason,
            'return_comment': open_history.return_requested_comment,
        }
    form = AssetCheckinForm(initial=initial)
    return render(request, 'partials/asset_checkin_modal.html', {
        'asset': asset,
        'form': form,
    })


@login_required
@require_POST
def asset_checkin(request, pk):
    """Check in an asset."""
    if effective_role_name(request.user) not in ['ADMIN', 'SUPERADMIN']:
        return HttpResponse(status=403)

    asset = get_object_or_404(Asset, pk=pk)

    # Only allow checkin if asset is checked out
    if not asset.is_checked_out:
        if request.headers.get('HX-Request'):
            return HttpResponse('This asset is not currently checked out.', status=400)
        messages.error(request, 'This asset is not currently checked out.')
        return redirect('tickets:assets')

    form = AssetCheckinForm(request.POST)
    if not form.is_valid():
        if request.headers.get('HX-Request'):
            return render(request, 'partials/asset_checkin_modal.html', {'asset': asset, 'form': form})
        messages.error(request, 'Please correct the errors below.')
        return redirect('tickets:assets')

    return_reason = form.cleaned_data['return_reason']
    return_comment = form.cleaned_data['return_comment'].strip()
    return_condition = form.cleaned_data['return_condition'].strip()
    holder = asset.checked_out_to  # captured before release() clears it

    # release() is the single source of truth for taking an asset back —
    # it decides the post-return status from BOTH condition and reason (a
    # LOST/STOLEN asset lands on that exact status, not IN_STORE), and
    # closes the open AssetCheckoutHistory row itself.
    asset.release(
        actor=request.user,
        return_reason=return_reason,
        return_comment=return_comment,
        return_condition=return_condition,
    )

    if holder:
        Notification.objects.create(
            recipient=holder, role=role_of(holder),
            message=f'Your return of "{asset.name}" ({asset.tracking_id}) has been confirmed. Thanks!',
            url='/tickets/my-assets/',
        )

    # Add comment to asset notes
    checkin_note = f"**Asset checked in** by {request.user.get_full_name()} on {timezone.now().strftime('%Y-%m-%d %H:%M')}"
    checkin_note += f"\nReason: {asset.get_return_reason_display()}"
    if return_condition:
        checkin_note += f"\nCondition: {return_condition}"
    if return_comment:
        checkin_note += f"\nComment: {return_comment}"
    
    if asset.notes:
        asset.notes = f"{asset.notes}\n\n{checkin_note}"
    else:
        asset.notes = checkin_note
    asset.save(update_fields=['notes'])
    
    messages.success(request, f'Asset "{asset.name}" checked in successfully.')
    if request.headers.get('HX-Request'):
        return HttpResponse(status=204, headers={'HX-Redirect': reverse('tickets:assets')})
    return redirect('tickets:assets')


@login_required
def asset_checkout_history(request, pk):
    """View checkout history for an asset."""
    if effective_role_name(request.user) not in ('ADMIN', 'SUPERADMIN'):
        return HttpResponse(status=403)

    asset = get_object_or_404(Asset, pk=pk)
    history = asset.checkout_history.all().order_by('-checked_out_at')

    return render(request, 'partials/asset_checkout_history.html', {
        'asset': asset,
        'history': history,
    })


# ==========================================================================
# ASSET CUSTODY TWO-STEP CONFIRMATION — recipient accepts/disputes a
# checkout or reassignment handover; holder self-initiates/cancels a
# return, which an admin then confirms via the existing asset_checkin.
# Reachable from My Assets, not the admin-only asset inventory.
# ==========================================================================

@login_required
@require_POST
def asset_checkout_accept(request, pk):
    asset = get_object_or_404(Asset, pk=pk)
    try:
        asset.acknowledge_checkout(actor=request.user)
    except ValueError as e:
        if request.headers.get('HX-Request'):
            return HttpResponse(str(e), status=400)
        messages.error(request, str(e))
        return redirect('tickets:my_assets')
    messages.success(request, f'Confirmed receipt of "{asset.name}".')
    if request.headers.get('HX-Request'):
        return HttpResponse(status=204, headers={'HX-Redirect': reverse('tickets:my_assets')})
    return redirect('tickets:my_assets')


@login_required
@require_POST
def asset_checkout_dispute(request, pk):
    asset = get_object_or_404(Asset, pk=pk)
    reason = request.POST.get('reason', '').strip()
    try:
        asset.dispute_checkout(actor=request.user, reason=reason)
    except ValueError as e:
        if request.headers.get('HX-Request'):
            return HttpResponse(str(e), status=400)
        messages.error(request, str(e))
        return redirect('tickets:my_assets')
    messages.success(request, f'Reported — an admin will follow up on "{asset.name}".')
    if request.headers.get('HX-Request'):
        return HttpResponse(status=204, headers={'HX-Redirect': reverse('tickets:my_assets')})
    return redirect('tickets:my_assets')


# ==========================================================================
# MOBILIZATION RECEIPT TWO-STEP CONFIRMATION — the requester-facing
# counterpart to the asset checkout accept/dispute above: the ticket's
# requester (not the admin who mobilized the assets) confirms or disputes
# receipt of each MobilizationItem individually. Reachable only via the
# receipt_confirm_modal below, embedded in the ticket conversation page.
# ==========================================================================

@login_required
@require_POST
def mobilization_item_accept(request, item_pk):
    item = get_object_or_404(MobilizationItem.objects.select_related('mobilization__ticket'), pk=item_pk)
    try:
        item.acknowledge_receipt(actor=request.user)
    except ValueError as e:
        messages.error(request, str(e))
        return redirect('tickets:detail', pk=item.mobilization.ticket_id or item.mobilization_id)
    if item.mobilization.ticket_id:
        _maybe_resolve_mobilization_receipt(item.mobilization.ticket, request.user)
    messages.success(request, f'Confirmed receipt of "{item.asset.name}".')
    return redirect('tickets:detail', pk=item.mobilization.ticket_id)


@login_required
@require_POST
def mobilization_item_dispute(request, item_pk):
    item = get_object_or_404(MobilizationItem.objects.select_related('mobilization__ticket'), pk=item_pk)
    reason = request.POST.get('reason', '').strip()
    try:
        item.dispute_receipt(actor=request.user, reason=reason)
    except ValueError as e:
        messages.error(request, str(e))
        return redirect('tickets:detail', pk=item.mobilization.ticket_id or item.mobilization_id)
    if item.mobilization.ticket_id:
        _maybe_resolve_mobilization_receipt(item.mobilization.ticket, request.user)
    messages.success(request, f'Reported — an admin will follow up on "{item.asset.name}".')
    return redirect('tickets:detail', pk=item.mobilization.ticket_id)


@login_required
def receipt_confirm_modal(request, pk):
    """Modal shown from the ticket conversation page's compact "confirm
    receipt" signifier — the single UI entry point for both the per-item
    mobilization handshake and the single-asset asset-request confirmation.
    Reachable only by the ticket's own requester while it's actually
    awaiting their response."""
    ticket = get_object_or_404(Ticket, pk=pk)
    if request.user != ticket.requester:
        return HttpResponse(status=403)
    if ticket.status != Ticket.Status.PENDING_USER or not ticket.is_asset_request:
        return HttpResponse("There's nothing awaiting your confirmation for this ticket.", status=400)

    context = {'ticket': ticket}
    if ticket.is_mobilization_request:
        items = MobilizationItem.objects.filter(mobilization__ticket=ticket).select_related('asset', 'mobilization')
        context['pending_items'] = items.filter(acknowledged_at__isnull=True, disputed_at__isnull=True)
        context['actioned_items'] = items.exclude(acknowledged_at__isnull=True, disputed_at__isnull=True)
    return render(request, 'partials/receipt_confirm_modal.html', context)


@login_required
@require_POST
def mobilization_items_confirm_batch(request, pk):
    """Single submit for the receipt-confirm modal's item checklist. The
    requester marks each item Accept/Dispute locally in the modal (no
    per-click page reload, no per-item message) and only this one submit —
    behind a single 'Done' confirmation — actually applies every decision,
    in one batch, then runs the usual aggregation once at the end."""
    ticket = get_object_or_404(Ticket, pk=pk)
    if request.user != ticket.requester:
        return HttpResponse(status=403)

    accept_ids = request.POST.getlist('accept_ids')
    dispute_ids = request.POST.getlist('dispute_ids')
    items = MobilizationItem.objects.filter(
        pk__in=accept_ids + dispute_ids, mobilization__ticket=ticket
    ).select_related('asset', 'mobilization')
    items_by_pk = {str(item.pk): item for item in items}

    accepted_count = disputed_count = 0
    for item_pk in accept_ids:
        item = items_by_pk.get(item_pk)
        if not item:
            continue
        try:
            item.acknowledge_receipt(actor=request.user)
            accepted_count += 1
        except ValueError:
            pass
    for item_pk in dispute_ids:
        item = items_by_pk.get(item_pk)
        if not item:
            continue
        try:
            item.dispute_receipt(actor=request.user, reason=request.POST.get(f'reason_{item_pk}', '').strip())
            disputed_count += 1
        except ValueError:
            pass

    _maybe_resolve_mobilization_receipt(ticket, request.user)

    parts = []
    if accepted_count:
        parts.append(f'{accepted_count} confirmed')
    if disputed_count:
        parts.append(f'{disputed_count} disputed')
    messages.success(request, (', '.join(parts) + '.') if parts else 'No changes recorded.')
    return redirect('tickets:detail', pk=ticket.pk)


@login_required
def asset_request_return_modal(request, pk):
    """Confirmation modal for relinquishing an asset — reachable only by
    whoever the asset is actually checked out to."""
    asset = get_object_or_404(Asset, pk=pk)
    if asset.checked_out_to_id != request.user.id:
        return HttpResponse("This asset isn't checked out to you.", status=403)
    open_history = asset._open_checkout_history()
    if open_history and open_history.return_requested_at:
        return HttpResponse('A return has already been requested for this asset.', status=400)
    form = AssetReturnRequestForm()
    return render(request, 'partials/asset_return_request_modal.html', {'asset': asset, 'form': form})


@login_required
@require_POST
def asset_request_return(request, pk):
    asset = get_object_or_404(Asset, pk=pk)
    if asset.checked_out_to_id != request.user.id:
        return HttpResponse("This asset isn't checked out to you.", status=403)

    form = AssetReturnRequestForm(request.POST)
    if not form.is_valid():
        if request.headers.get('HX-Request'):
            return render(request, 'partials/asset_return_request_modal.html', {'asset': asset, 'form': form})
        messages.error(request, 'Please correct the errors below.')
        return redirect('tickets:my_assets')

    try:
        asset.request_return(
            actor=request.user,
            reason=form.cleaned_data['return_reason'],
            comment=form.cleaned_data['return_comment'].strip(),
        )
    except ValueError as e:
        if request.headers.get('HX-Request'):
            return HttpResponse(str(e), status=400)
        messages.error(request, str(e))
        return redirect('tickets:my_assets')

    messages.success(request, f'Return requested for "{asset.name}" — an admin will arrange pickup.')
    if request.headers.get('HX-Request'):
        return HttpResponse(status=204, headers={'HX-Redirect': reverse('tickets:my_assets')})
    return redirect('tickets:my_assets')


@login_required
@require_POST
def asset_cancel_return_request(request, pk):
    asset = get_object_or_404(Asset, pk=pk)
    if asset.checked_out_to_id != request.user.id:
        return HttpResponse("This asset isn't checked out to you.", status=403)
    try:
        asset.cancel_return_request(actor=request.user)
    except ValueError as e:
        if request.headers.get('HX-Request'):
            return HttpResponse(str(e), status=400)
        messages.error(request, str(e))
        return redirect('tickets:my_assets')
    messages.success(request, f'Return request for "{asset.name}" cancelled.')
    if request.headers.get('HX-Request'):
        return HttpResponse(status=204, headers={'HX-Redirect': reverse('tickets:my_assets')})
    return redirect('tickets:my_assets')


@login_required
def pending_asset_returns_list(request):
    """Full, paginated list of assets whose holder has requested a return
    but an admin hasn't yet confirmed physical retrieval — the admin-side
    queue this two-step flow needs, parallel to pending_asset_fulfillment_list."""
    if effective_role_name(request.user) not in ['ADMIN', 'SUPERADMIN']:
        return HttpResponse(status=403)

    history_qs = AssetCheckoutHistory.objects.filter(
        return_requested_at__isnull=False, checked_in_at__isnull=True
    ).select_related('asset', 'checked_out_to').order_by('return_requested_at')

    paginator = Paginator(history_qs, 10)
    page_obj = paginator.get_page(request.GET.get('page'))

    return render(request, 'tickets/pending_asset_returns_list.html', {
        'returns': page_obj,
        'sidebar_template': get_sidebar_template(request.user),
    })


@login_required
def pending_asset_returns_count(request):
    """Badge count for the sidebar 'Returns' link — same filter as
    pending_asset_returns_list."""
    if effective_role_name(request.user) not in ['ADMIN', 'SUPERADMIN']:
        return HttpResponse(status=403)

    count = AssetCheckoutHistory.objects.filter(
        return_requested_at__isnull=False, checked_in_at__isnull=True
    ).count()
    return render(request, 'partials/sidebar_count_badge.html', {'count': count})


# ==========================================================================
# MOBILIZATION / DEMOBILIZATION
# ==========================================================================

@login_required
def mobilizations(request):
    if effective_role_name(request.user) not in ['ADMIN', 'SUPERADMIN', 'AGENT', 'TEAM_LEAD'] or request.user.department != 'IT':
        return HttpResponse(status=403)

    tab = request.GET.get('tab', 'all')
    job_filter = request.GET.get('filter_job', '')
    vessel_filter = request.GET.get('filter_vessel', '')
    system_filter = request.GET.get('filter_system', '')
    status_filter = request.GET.get('filter_status', '')

    mobilizations_list = Mobilization.objects.select_related('job_number', 'mobilized_by').prefetch_related('vessels', 'dive_systems', 'items__asset')

    if job_filter:
        mobilizations_list = mobilizations_list.filter(job_number_id=job_filter)
    if vessel_filter:
        mobilizations_list = mobilizations_list.filter(vessels__id=vessel_filter)
    if system_filter:
        mobilizations_list = mobilizations_list.filter(dive_systems__id=system_filter)
    if status_filter:
        mobilizations_list = mobilizations_list.filter(status=status_filter)

    mobilizations_list = mobilizations_list.annotate(
        pending_demob_count=Count(
            'items',
            filter=Q(items__return_requested_at__isnull=False, items__demobilized_at__isnull=True),
            distinct=True,
        )
    ).distinct()

    # Counted before the tab split so the tab label always reflects the
    # true total, not just what's on the current tab's page.
    needs_confirmation_count = mobilizations_list.filter(pending_demob_count__gt=0).count()

    if tab == 'needs_confirmation':
        mobilizations_list = mobilizations_list.filter(pending_demob_count__gt=0)

    mobilizations_list = mobilizations_list.order_by('-mobilized_at')

    paginator = Paginator(mobilizations_list, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    context = {
        'mobilizations': page_obj,
        'tab': tab,
        'needs_confirmation_count': needs_confirmation_count,
        'job_numbers': JobNumber.objects.filter(is_active=True),
        'vessels': Vessel.objects.filter(is_active=True),
        'dive_systems': DiveSystem.objects.filter(is_active=True),
        'status_choices': Mobilization.Status.choices,
        'selected_job': job_filter,
        'selected_vessel': vessel_filter,
        'selected_system': system_filter,
        'selected_status': status_filter,
        'sidebar_template': get_sidebar_template(request.user),
    }

    if request.headers.get('HX-Request'):
        return render(request, 'partials/mobilization_table.html', context)

    return render(request, 'tickets/mobilization_list.html', context)


@login_required
def mobilization_detail(request, pk):
    if effective_role_name(request.user) not in ['ADMIN', 'SUPERADMIN', 'AGENT', 'TEAM_LEAD'] or request.user.department != 'IT':
        return HttpResponse(status=403)

    mobilization = get_object_or_404(
        Mobilization.objects.select_related('job_number', 'mobilized_by', 'ticket').prefetch_related(
            'vessels', 'dive_systems', 'items__asset', 'date_extensions__extended_by'
        ),
        pk=pk
    )

    return render(request, 'tickets/mobilization_detail.html', {
        'mobilization': mobilization,
        'items': mobilization.items.select_related('asset', 'demobilized_by').all(),
        'date_extensions': mobilization.date_extensions.all(),
        'sidebar_template': get_sidebar_template(request.user),
    })


@login_required
def mobilization_audit_report(request, pk):
    """Full mobilize-to-demobilize audit trail for one mobilization —
    linked from mobilization_detail, same access control. Built for an
    IT admin/agent auditing how a specific batch of assets moved out and
    back into inventory (not the requester-facing demobilization page,
    which only covers self-report status)."""
    if effective_role_name(request.user) not in ['ADMIN', 'SUPERADMIN', 'AGENT', 'TEAM_LEAD'] or request.user.department != 'IT':
        return HttpResponse(status=403)

    mobilization = get_object_or_404(
        Mobilization.objects.select_related('job_number', 'mobilized_by', 'ticket', 'ticket__requester').prefetch_related(
            'vessels', 'dive_systems'
        ),
        pk=pk
    )

    from .report_registry import mobilization_audit_sections
    context = mobilization_audit_sections(mobilization)
    context['sidebar_template'] = get_sidebar_template(request.user)
    return render(request, 'reports/mobilization_audit_report.html', context)


@login_required
def mobilization_audit_export(request, pk):
    if effective_role_name(request.user) not in ['ADMIN', 'SUPERADMIN', 'AGENT', 'TEAM_LEAD'] or request.user.department != 'IT':
        return HttpResponse(status=403)

    mobilization = get_object_or_404(
        Mobilization.objects.select_related('job_number', 'mobilized_by', 'ticket', 'ticket__requester'),
        pk=pk
    )
    from .report_exporters import export_mobilization_audit_pdf
    return export_mobilization_audit_pdf(mobilization, request)


@login_required
def mobilization_create_page(request):
    if effective_role_name(request.user) not in ['ADMIN', 'SUPERADMIN']:
        return HttpResponse(status=403)

    ticket_id = request.GET.get('ticket_id')
    ticket = Ticket.objects.filter(pk=ticket_id).first() if ticket_id else None
    if not ticket or not ticket.is_mobilization_request:
        messages.error(request, 'Mobilizing assets requires a linked mobilization request ticket.')
        return redirect('tickets:mobilizations')

    form = MobilizationForm(initial={
        'job_number': ticket.job_number_id,
        'vessels': ticket.vessels.values_list('id', flat=True),
        'dive_systems': ticket.dive_systems.values_list('id', flat=True),
        'notes': ticket.purpose,
    })

    # Carry over what the request itself already said it needed — the
    # ASSET field group's asset_type/number_of_assets — into the Quick
    # Add by Quantity picker, so the admin isn't retyping what the
    # requester already specified. No expected_return_date prefill:
    # nothing on the request captures a target return date today.
    prefill_category = None
    prefill_quantity = None
    if ticket.service_category and ticket.service_category.field_group == ServiceCategory.FieldGroup.ASSET:
        details = ticket.service_request_details or {}
        asset_type = details.get('asset_type')
        if asset_type:
            for f in fields_for_group(ServiceCategory.FieldGroup.ASSET):
                if f.key == 'asset_type':
                    label = display_value_for_field(f, asset_type)
                    prefill_category = AssetCategory.objects.filter(name__iexact=label).first()
                    break
        try:
            prefill_quantity = int(details.get('number_of_assets') or 0) or None
        except (TypeError, ValueError):
            prefill_quantity = None

    return render(request, 'tickets/mobilization_create.html', {
        'form': form,
        'ticket': ticket,
        'trackable_categories': AssetCategory.objects.filter(is_consumable=False).order_by('name'),
        'asset_categories': AssetCategory.objects.all().order_by('name'),
        'vendors': Vendor.objects.filter(is_active=True).prefetch_related('categories').order_by('name'),
        'prefill_category': prefill_category,
        'prefill_quantity': prefill_quantity,
        'sidebar_template': get_sidebar_template(request.user),
    })


@login_required
@require_POST
def mobilization_create(request):
    if effective_role_name(request.user) not in ['ADMIN', 'SUPERADMIN']:
        return HttpResponse(status=403)

    ticket_id = request.POST.get('ticket_id')
    ticket = Ticket.objects.filter(pk=ticket_id).first() if ticket_id else None
    if not ticket or not ticket.is_mobilization_request:
        messages.error(request, 'Mobilizing assets requires a linked mobilization request ticket.')
        return redirect('tickets:mobilizations')

    form = MobilizationForm(request.POST)
    asset_ids = request.POST.getlist('asset_ids')

    # Vendor-request rows: item name / category / quantity / vendor /
    # expected date, one set of parallel arrays per row — same
    # multi-value-under-one-name shape as third_party_vessels above.
    procurement_item_names = request.POST.getlist('procurement_item_name')
    procurement_category_ids = request.POST.getlist('procurement_category_id')
    procurement_quantities = request.POST.getlist('procurement_quantity')
    procurement_vendor_ids = request.POST.getlist('procurement_vendor_id')
    procurement_vendor_names = request.POST.getlist('procurement_vendor_name')
    procurement_dates = request.POST.getlist('procurement_expected_date')
    has_procurement_rows = any(name.strip() for name in procurement_item_names)

    if not asset_ids and not has_procurement_rows:
        messages.error(request, 'Select at least one asset to mobilize, or add a vendor-request line item.')
        return redirect('tickets:mobilizations')

    # Category is required per vendor-request row (it's what the receiving
    # step uses to file the item once it arrives) — checked up front, before
    # the mobilization is created below, so a row missing it is rejected
    # outright instead of being silently dropped after the rest of the
    # mobilization has already been saved.
    missing_category_items = []
    for i, item_name in enumerate(procurement_item_names):
        item_name = item_name.strip()
        if not item_name:
            continue
        category_id = procurement_category_ids[i] if i < len(procurement_category_ids) else None
        if not category_id or not AssetCategory.objects.filter(pk=category_id).exists():
            missing_category_items.append(item_name)
    if missing_category_items:
        names = ', '.join(missing_category_items)
        messages.error(request, f'Select a category for: {names}.')
        return redirect('tickets:mobilizations')

    if not form.is_valid():
        for field_errors in form.errors.values():
            for error in field_errors:
                messages.error(request, error)
        return redirect('tickets:mobilizations')

    with transaction.atomic():
        assets_qs = Asset.objects.select_for_update().filter(pk__in=asset_ids)
        # assigned_to check is separate from is_available (which is
        # deliberately assigned_to-agnostic, shared with checkout) — an
        # asset permanently assigned to someone must not be pickable for an
        # unrelated job mobilization, mirroring the guard already applied
        # at ticket-fulfillment time (asset_id branch above).
        unavailable = [a for a in assets_qs if not a.is_available or a.assigned_to_id]
        if unavailable:
            names = ', '.join(a.tracking_id for a in unavailable)
            messages.error(request, f'These assets are not available to mobilize: {names}.')
            return redirect('tickets:mobilizations')

        # Quantity per asset — only meaningful for consumable/bulk stock
        # (e.g. "5 units of cable ties"). Individually-tracked assets are
        # always exactly 1, regardless of what's posted.
        quantities = {}
        insufficient_stock = []
        for asset in assets_qs:
            if asset.is_consumable:
                try:
                    requested = int(request.POST.get(f'quantity_{asset.pk}', 1))
                except (TypeError, ValueError):
                    requested = 1
                requested = max(1, requested)
                if requested > asset.quantity_in_stock:
                    insufficient_stock.append(asset)
                quantities[asset.pk] = requested
            else:
                quantities[asset.pk] = 1
        if insufficient_stock:
            names = ', '.join(a.name for a in insufficient_stock)
            messages.error(request, f'Not enough stock to mobilize the requested quantity for: {names}.')
            return redirect('tickets:mobilizations')

        mobilization = form.save(commit=False)
        mobilization.mobilized_by = request.user
        mobilization.mobilized_at = timezone.now()
        mobilization.ticket = ticket
        mobilization.original_expected_return_date = mobilization.expected_return_date
        mobilization.save()
        form.save_m2m()

        # Third-party vessels: reuse an existing proposal case-insensitively,
        # else create one pending approval (is_active=False) — same
        # propose-and-approve shape as Job Number. The proposer's own
        # mobilization can reference it immediately even while pending.
        new_third_party_vessels = []
        for name in form.cleaned_data.get('third_party_vessels', []):
            vessel = Vessel.objects.filter(name__iexact=name).first()
            if not vessel:
                vessel = Vessel.objects.create(name=name, is_active=False, proposed_by=request.user)
                new_third_party_vessels.append(vessel)
            mobilization.vessels.add(vessel)

        # Vendor-request line items: items wanted for this job that aren't
        # in stock — recorded against this mobilization, fulfilled later via
        # the Procurement "Receive" step, same as an asset-request ticket.
        new_procurement_vendors = []
        for i, item_name in enumerate(procurement_item_names):
            item_name = item_name.strip()
            if not item_name:
                continue
            category_id = procurement_category_ids[i] if i < len(procurement_category_ids) else None
            category = AssetCategory.objects.filter(pk=category_id).first() if category_id else None
            if not category:
                continue
            try:
                qty = max(1, int(procurement_quantities[i])) if i < len(procurement_quantities) else 1
            except (TypeError, ValueError):
                qty = 1

            vendor = None
            vendor_id = procurement_vendor_ids[i] if i < len(procurement_vendor_ids) else ''
            vendor_name = procurement_vendor_names[i].strip() if i < len(procurement_vendor_names) else ''
            if vendor_id:
                vendor = Vendor.objects.filter(pk=vendor_id).first()
            elif vendor_name:
                vendor, created = Vendor.objects.get_or_create(
                    name__iexact=vendor_name, defaults={'name': vendor_name, 'is_active': False}
                )
                if created:
                    new_procurement_vendors.append(vendor)

            expected_date = procurement_dates[i] if i < len(procurement_dates) else ''

            AssetProcurementRequest.objects.create(
                item_name=item_name,
                category=category,
                quantity=qty,
                vendor=vendor,
                expected_arrival_date=expected_date or None,
                mobilization=mobilization,
                requested_by=request.user,
            )

        for asset in assets_qs:
            qty = quantities[asset.pk]
            MobilizationItem.objects.create(mobilization=mobilization, asset=asset, quantity=qty)

            if asset.is_consumable:
                # SKU-level record stays IN_STORE throughout — only the
                # remaining count changes, since other units of the same
                # stock may still be available.
                asset.quantity_in_stock -= qty
                asset.save(update_fields=['quantity_in_stock'])
                asset.refresh_low_stock_alert()
            else:
                asset.status = Asset.Status.MOBILIZED
                asset.status_updated_at = timezone.now()
                asset.status_updated_by = request.user
                asset.save(update_fields=['status', 'status_updated_at', 'status_updated_by'])

            AssetLog.objects.create(
                asset=asset,
                action=AssetLog.Action.MOBILIZED,
                actor=request.user,
                details={
                    'mobilization_id': mobilization.pk,
                    'destination': mobilization.destination_display,
                    'quantity': qty,
                }
            )

    if ticket:
        # One itemized summary per line (stock pick or vendor order),
        # grouped by name+quantity — not per-asset-per-tracking-ID, this is
        # a lightweight manifest, distinct from the confirm-receipt card
        # which still lists each physical unit individually. Mixed
        # mobilizations (some stock, some vendor) get one coherent comment
        # instead of two disconnected ones.
        stock_groups = {}
        for asset in assets_qs:
            stock_groups[asset.name] = stock_groups.get(asset.name, 0) + quantities[asset.pk]
        procurement_rows = list(mobilization.procurement_requests.all())

        if stock_groups or procurement_rows:
            lines = [f"{qty}× {escape(name)} — mobilized from stock" for name, qty in stock_groups.items()]
            lines += [f"{pr.quantity}× {escape(pr.item_name)} — ordered from vendor" for pr in procurement_rows]
            total = sum(stock_groups.values()) + sum(pr.quantity for pr in procurement_rows)
            title = f"{total} item{'s' if total != 1 else ''} requested for {escape(mobilization.destination_display)}"
            TicketComment.objects.create(
                ticket=ticket,
                author=request.user,
                visibility='PUBLIC',
                mobilization=mobilization,
                mobilization_event=TicketComment.MobilizationEvent.CREATED,
                body=f"<strong>{title}</strong><br>" + "<br>".join(lines),
                is_system_generated=True,
            )
        # Only actually fulfills the ticket (and prompts the requester to
        # confirm receipt) once nothing on this mobilization is still on
        # order from a vendor — a procurement-only mobilization instead
        # routes the ticket to PENDING_VENDOR here and gets fulfilled later,
        # from procurement_receive/procurement_cancel, once that clears.
        _maybe_fulfill_mobilization_ticket(mobilization, request)

    for vessel in new_third_party_vessels:
        for admin in User.objects.filter(role=User.Role.ADMIN, is_active=True):
            Notification.objects.create(
                recipient=admin,
                role=role_of(admin),
                message=(
                    f'{request.user.get_full_name()} mobilized assets to "{vessel.name}", a third-party vessel '
                    f'not yet in the system. Review and activate it under System Settings → Vessels if it should be added.'
                ),
                url=reverse('tickets:mobilization_detail', args=[mobilization.pk]),
            )

    for vendor in new_procurement_vendors:
        _notify_new_vendor_proposed(vendor, request.user, reverse('tickets:mobilization_detail', args=[mobilization.pk]))

    summary_parts = []
    if assets_qs.count():
        summary_parts.append(f'{assets_qs.count()} asset(s) mobilized')
    procurement_count = mobilization.procurement_requests.count()
    if procurement_count:
        summary_parts.append(f'{procurement_count} item(s) recorded as on order')
    messages.success(request, f'{" and ".join(summary_parts)} to {mobilization.destination_display}.')
    return redirect('tickets:mobilization_detail', pk=mobilization.pk)


@login_required
def mobilization_available_assets(request):
    """HTMX endpoint: assets available to add to a new mobilization."""
    if effective_role_name(request.user) not in ['ADMIN', 'SUPERADMIN']:
        return HttpResponse(status=403)

    search = request.GET.get('search', '').strip()

    assets_qs = Asset.objects.filter(
        status__in=[Asset.Status.IN_STORE, Asset.Status.READY],
        checked_out_to__isnull=True,
        assigned_to__isnull=True,
    ).exclude(
        category__is_consumable=True, quantity_in_stock__lte=0,
    ).select_related('category').order_by('name')

    if search:
        assets_qs = assets_qs.filter(
            Q(name__icontains=search) |
            Q(tracking_id__icontains=search) |
            Q(serial_number__icontains=search) |
            Q(model__icontains=search) |
            Q(manufacturer__icontains=search)
        )

    assets_qs = assets_qs[:20]

    return render(request, 'partials/mobilization_available_assets_list.html', {
        'assets': assets_qs,
    })


@login_required
def mobilization_autopick_assets(request):
    """HTMX/JSON endpoint: given a category + quantity, return up to that
    many available individually-tracked assets in that category. This only
    speeds up *selection* — the returned IDs still go through the exact same
    `asset_ids` list and select_for_update() re-validation in
    mobilization_create as a manually-checked pick, so a race between pick
    and submit just surfaces the existing "not available" error there."""
    if effective_role_name(request.user) not in ['ADMIN', 'SUPERADMIN']:
        return HttpResponse(status=403)

    category_id = request.GET.get('category_id')
    try:
        quantity = int(request.GET.get('quantity', 0))
    except (TypeError, ValueError):
        quantity = 0
    quantity = max(0, min(quantity, 50))

    if not category_id or quantity == 0:
        return JsonResponse({'assets': []})

    assets_qs = Asset.objects.filter(
        category_id=category_id,
        category__is_consumable=False,
        status__in=[Asset.Status.IN_STORE, Asset.Status.READY],
        checked_out_to__isnull=True,
        assigned_to__isnull=True,
    ).order_by('tracking_id')[:quantity]

    return JsonResponse({
        'assets': [
            {'id': asset.pk, 'name': asset.name, 'tracking_id': asset.tracking_id}
            for asset in assets_qs
        ]
    })


@login_required
def mobilization_item_demobilize_modal(request, item_pk):
    if effective_role_name(request.user) not in ['ADMIN', 'SUPERADMIN']:
        return HttpResponse(status=403)

    item = get_object_or_404(MobilizationItem.objects.select_related('asset', 'mobilization'), pk=item_pk)
    if not item.is_active:
        return HttpResponse('<div class="p-4 text-center text-warning">This asset has already been demobilized.</div>', status=400)

    has_ticket = bool(item.mobilization.ticket_id)
    return render(request, 'partials/mobilization_demobilize_modal.html', {
        'item': item,
        'condition_choices': Asset.Condition.choices,
        'has_ticket': has_ticket,
        'is_blocked': has_ticket and not item.return_requested_at,
    })


def _demobilize_item(item, return_condition, return_notes, actor, return_quantity=None, override_reason=''):
    """Core of returning one MobilizationItem's asset to inventory — shared
    by the single-item demobilize view and the batch "Demobilize All" view.
    Does not call item.mobilization.refresh_status(); callers do that once
    after all items in a batch are processed, to avoid redundant recomputes.

    Requires the requester to have self-reported this item as returned
    (item.return_requested_at) whenever the mobilization has a linked
    ticket — the whole point of the demobilization handshake is that the
    return leg stays traceable through the system, not just admin say-so.
    The only carve-out is a legacy mobilization with no linked ticket (no
    requester ever could self-report it), which requires a typed
    override_reason instead. Both call sites validate this up front too;
    raising here is a safety net, not the primary guard."""
    if item.mobilization.ticket_id:
        if not item.return_requested_at:
            raise ValueError(f'"{item.asset.name}" has not been reported returned by the requester yet.')
    else:
        if not override_reason.strip():
            raise ValueError(f'"{item.asset.name}" needs a reason — this mobilization has no linked ticket/requester to report a return.')

    asset = item.asset

    item.demobilized_at = timezone.now()
    item.demobilized_by = actor
    item.return_condition = return_condition
    item.return_notes = return_notes

    if asset.is_consumable:
        # Partial returns: only the portion actually coming back to usable
        # stock is restored. Damaged/unusable consumables aren't restocked
        # at all — they're treated as consumed, not returned.
        if return_quantity is None:
            return_quantity = item.quantity
        return_quantity = max(0, min(return_quantity, item.quantity))
        item.return_quantity = return_quantity
        item.save()

        if return_condition not in [Asset.Condition.DAMAGED, Asset.Condition.UNUSABLE]:
            asset.quantity_in_stock += return_quantity
            asset.save(update_fields=['quantity_in_stock'])
            asset.refresh_low_stock_alert()
    else:
        item.save()
        # Damaged/unusable gear goes to maintenance instead of straight back onto the shelf.
        if return_condition in [Asset.Condition.DAMAGED, Asset.Condition.UNUSABLE]:
            asset.status = Asset.Status.MAINTENANCE
        else:
            asset.status = Asset.Status.IN_STORE
        asset.condition = return_condition
        asset.status_updated_at = timezone.now()
        asset.status_updated_by = actor
        asset.save(update_fields=['status', 'condition', 'status_updated_at', 'status_updated_by'])

    details = {
        'mobilization_id': item.mobilization_id,
        'condition': return_condition,
        'notes': return_notes,
    }
    if item.return_requested_at:
        details['self_reported_at'] = item.return_requested_at.isoformat()
        details['self_reported_by'] = item.return_requested_by.get_full_name() if item.return_requested_by else None
        details['self_reported_notes'] = item.return_requested_notes
    if override_reason:
        details['override_reason'] = override_reason

    AssetLog.objects.create(
        asset=asset,
        action=AssetLog.Action.DEMOBILIZED,
        actor=actor,
        details=details,
    )


@login_required
@require_POST
def mobilization_item_demobilize(request, item_pk):
    if effective_role_name(request.user) not in ['ADMIN', 'SUPERADMIN']:
        return HttpResponse(status=403)

    return_condition = request.POST.get('return_condition')
    return_notes = request.POST.get('return_notes', '').strip()
    override_reason = request.POST.get('override_reason', '').strip()

    with transaction.atomic():
        item = get_object_or_404(
            MobilizationItem.objects.select_for_update().select_related('asset', 'mobilization'),
            pk=item_pk
        )
        if not item.is_active:
            messages.error(request, 'This asset has already been demobilized.')
            return redirect('tickets:mobilization_detail', pk=item.mobilization_id)

        if not return_condition:
            messages.error(request, 'Please select the returned condition.')
            return redirect('tickets:mobilization_detail', pk=item.mobilization_id)

        try:
            return_quantity = int(request.POST.get('return_quantity', item.quantity))
        except (TypeError, ValueError):
            return_quantity = item.quantity

        asset_name = item.asset.name
        mobilization_id = item.mobilization_id
        try:
            _demobilize_item(
                item, return_condition, return_notes, request.user,
                return_quantity=return_quantity, override_reason=override_reason,
            )
        except ValueError as e:
            messages.error(request, str(e))
            return redirect('tickets:mobilization_detail', pk=mobilization_id)
        item.mobilization.refresh_status()

    messages.success(request, f'Asset "{asset_name}" demobilized and returned to inventory.')
    return redirect('tickets:mobilization_detail', pk=mobilization_id)


@login_required
def mobilization_demobilize_all_modal(request, pk):
    if effective_role_name(request.user) not in ['ADMIN', 'SUPERADMIN']:
        return HttpResponse(status=403)

    mobilization = get_object_or_404(Mobilization, pk=pk)
    active_items = mobilization.items.filter(demobilized_at__isnull=True).select_related('asset')
    if not active_items.exists():
        return HttpResponse('<div class="p-4 text-center text-warning">Every asset on this mobilization has already been demobilized.</div>', status=400)

    has_ticket = bool(mobilization.ticket_id)
    return render(request, 'partials/mobilization_demobilize_all_modal.html', {
        'mobilization': mobilization,
        'items': active_items,
        'condition_choices': Asset.Condition.choices,
        'has_ticket': has_ticket,
        'unreported_items': active_items.filter(return_requested_at__isnull=True) if has_ticket else [],
    })


@login_required
@require_POST
def mobilization_demobilize_all(request, pk):
    if effective_role_name(request.user) not in ['ADMIN', 'SUPERADMIN']:
        return HttpResponse(status=403)

    mobilization = get_object_or_404(Mobilization, pk=pk)
    return_condition = request.POST.get('return_condition')
    return_notes = request.POST.get('return_notes', '').strip()
    override_reason = request.POST.get('override_reason', '').strip()

    if not return_condition:
        messages.error(request, 'Please select the returned condition.')
        return redirect('tickets:mobilization_detail', pk=pk)

    with transaction.atomic():
        items = list(
            mobilization.items.select_for_update().filter(demobilized_at__isnull=True).select_related('asset')
        )
        if not items:
            messages.warning(request, 'Every asset on this mobilization has already been demobilized.')
            return redirect('tickets:mobilization_detail', pk=pk)

        # Validate the whole batch up front — all-or-nothing, so we never
        # demobilize some items and then bail out partway through.
        if mobilization.ticket_id:
            unreported = [i for i in items if not i.return_requested_at]
            if unreported:
                names = ', '.join(i.asset.tracking_id for i in unreported)
                messages.error(request, f'Not yet reported returned by the requester: {names}.')
                return redirect('tickets:mobilization_detail', pk=pk)
        elif not override_reason:
            messages.error(request, 'Please provide a reason — this mobilization has no linked ticket/requester.')
            return redirect('tickets:mobilization_detail', pk=pk)

        for item in items:
            _demobilize_item(item, return_condition, return_notes, request.user, override_reason=override_reason)

        mobilization.refresh_status()

    messages.success(request, f'{len(items)} asset(s) demobilized and returned to inventory.')
    return redirect('tickets:mobilization_detail', pk=pk)


@login_required
def mobilization_extend_date_modal(request, pk):
    if effective_role_name(request.user) not in ['ADMIN', 'SUPERADMIN']:
        return HttpResponse(status=403)

    mobilization = get_object_or_404(Mobilization, pk=pk)
    if mobilization.status != Mobilization.Status.ACTIVE:
        return HttpResponse('<div class="p-4 text-center text-warning">Only an active mobilization\'s return date can be extended.</div>', status=400)

    return render(request, 'partials/mobilization_extend_date_modal.html', {
        'mobilization': mobilization,
    })


@login_required
@require_POST
def mobilization_extend_date(request, pk):
    if effective_role_name(request.user) not in ['ADMIN', 'SUPERADMIN']:
        return HttpResponse(status=403)

    mobilization = get_object_or_404(Mobilization, pk=pk)
    if mobilization.status != Mobilization.Status.ACTIVE:
        messages.error(request, 'Only an active mobilization\'s return date can be extended.')
        return redirect('tickets:mobilization_detail', pk=pk)

    new_date_raw = request.POST.get('new_date', '').strip()
    reason = request.POST.get('reason', '').strip()

    if not new_date_raw:
        messages.error(request, 'Please choose a new return date.')
        return redirect('tickets:mobilization_detail', pk=pk)

    try:
        new_date = date.fromisoformat(new_date_raw)
    except ValueError:
        messages.error(request, 'Invalid date.')
        return redirect('tickets:mobilization_detail', pk=pk)

    previous_date = mobilization.expected_return_date
    if previous_date and new_date <= previous_date:
        messages.error(request, 'The new return date must be after the current return date.')
        return redirect('tickets:mobilization_detail', pk=pk)

    MobilizationDateExtension.objects.create(
        mobilization=mobilization,
        previous_date=previous_date,
        new_date=new_date,
        reason=reason,
        extended_by=request.user,
    )
    mobilization.expected_return_date = new_date
    mobilization.save(update_fields=['expected_return_date'])

    messages.success(request, f'Return date extended to {new_date.strftime("%b %d, %Y")}.')
    return redirect('tickets:mobilization_detail', pk=pk)


@login_required
def job_mobilization_lookup(request):
    """HTMX endpoint used from the asset-request fulfillment screen: shows
    what's currently mobilized (out) for a given job/vessel/dive system, so
    an admin can check existing issue before approving more."""
    if effective_role_name(request.user) not in ['ADMIN', 'SUPERADMIN']:
        return HttpResponse(status=403)

    job_id = request.GET.get('job_number')
    vessel_ids = request.GET.getlist('vessel')
    system_ids = request.GET.getlist('dive_system')

    items = MobilizationItem.objects.filter(demobilized_at__isnull=True).select_related('asset', 'mobilization')

    filters = Q()
    matched = False
    if job_id:
        filters |= Q(mobilization__job_number_id=job_id)
        matched = True
    if vessel_ids:
        filters |= Q(mobilization__vessels__id__in=vessel_ids)
        matched = True
    if system_ids:
        filters |= Q(mobilization__dive_systems__id__in=system_ids)
        matched = True

    items = items.filter(filters).distinct() if matched else items.none()

    return render(request, 'partials/job_mobilization_lookup.html', {
        'items': items,
    })